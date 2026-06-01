"""Excel Structure Defect Detector.

Scan a financial model Excel file and find structural defects:
- Cells that should be formulas but are static values
- SUMIF/VLOOKUP blocks with inconsistent formula coverage
- Parameter propagation chains with breaks

Usage:
    python -m financial_kg.diagnostic.structure_checker file.xlsx
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from ._shared import is_formula_cell, extract_formula_refs


@dataclass
class Block:
    """A group of structurally similar rows (e.g., 16 SUMIF blocks)."""
    sheet: str
    start_row: int
    end_row: int
    key_col: int = 3  # Column where SUMIF appears (usually C)
    sumif_row: int = 0
    date_row: int = 0
    year_row: int = 0
    month_row: int = 0
    cols: list[str] = field(default_factory=list)
    formula_cells: dict[str, bool] = field(default_factory=dict)  # ref -> has_formula

    @property
    def ref(self) -> str:
        return f"{self.sheet} 第 {self.start_row}-{self.end_row} 行"


@dataclass
class Defect:
    type: str          # "static_should_be_formula" | "propagation_break" | "inconsistent_block"
    severity: str      # "critical" | "warning" | "info"
    sheet: str
    cells: list[str]   # Defective cell references
    description: str
    context: dict = field(default_factory=dict)

    def __str__(self) -> str:
        cells_str = ", ".join(self.cells[:5])
        if len(self.cells) > 5:
            cells_str += f" ... ({len(self.cells)} total)"
        return f"[{self.severity.upper()}] {self.type}: {self.description}\n  Cells: {cells_str}"


@dataclass
class CheckReport:
    file_path: str
    total_blocks: int
    defects: list[Defect]

    def summary(self) -> str:
        lines = [
            f"Excel Structure Check: {self.file_path}",
            f"Blocks found: {self.total_blocks}",
            f"Defects found: {len(self.defects)}",
        ]
        crit = sum(1 for d in self.defects if d.severity == "critical")
        warn = sum(1 for d in self.defects if d.severity == "warning")
        info = sum(1 for d in self.defects if d.severity == "info")
        lines.append(f"  Critical: {crit}, Warning: {warn}, Info: {info}")
        lines.append("")
        for i, d in enumerate(self.defects, 1):
            lines.append(f"Defect #{i}:")
            lines.append(f"  {d}")
        return "\n".join(lines)


class ExcelStructureChecker:
    """Scan Excel file for structural defects in formula coverage."""

    def __init__(self, max_rows: int = 1000, max_cols: int = 50):
        self.max_rows = max_rows
        self.max_cols = max_cols

    def check(self, file_path: str) -> CheckReport:
        """Run full structure check on the given Excel file."""
        from datetime import datetime

        wb = openpyxl.load_workbook(file_path)

        all_defects: list[Defect] = []
        total_blocks = 0
        checked_date_rows: dict[str, set[int]] = {}  # sheet -> set of date rows

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Step 1: Find SUMIF blocks
            blocks = self._find_sumif_blocks(ws, sheet_name)
            total_blocks += len(blocks)

            # Step 2: Check formula coverage in each block
            for block in blocks:
                defects = self._check_formula_coverage(ws, block)
                all_defects.extend(defects)
                checked_date_rows.setdefault(sheet_name, set()).add(block.date_row)

            # Step 3: Check for inconsistent block patterns
            for block in blocks:
                defects = self._check_block_consistency(ws, block)
                all_defects.extend(defects)

            # Step 4: Scan ALL rows for formula-datetime gaps (catches rows
            # not associated with any SUMIF block, e.g. standalone date rows)
            skip = checked_date_rows.get(sheet_name, set())
            defects = self._scan_date_row_gaps(ws, sheet_name, skip)
            all_defects.extend(defects)

        wb.close()

        return CheckReport(
            file_path=file_path,
            total_blocks=total_blocks,
            defects=sorted(all_defects, key=lambda d: (
                {"critical": 0, "warning": 1, "info": 2}.get(d.severity, 3)
            )),
        )

    def _find_sumif_blocks(self, ws, sheet_name: str) -> list[Block]:
        """Identify SUMIF-containing row groups (blocks with similar structure)."""
        sumif_rows: list[int] = []

        for row_idx in range(1, self.max_rows + 1):
            cell = ws[f"C{row_idx}"]
            if cell.value and isinstance(cell.value, str) and "SUMIF" in cell.value:
                sumif_rows.append(row_idx)

        if not sumif_rows:
            return []

        # Group consecutive SUMIF rows into blocks
        # Each SUMIF row is part of a 4-row group: sumif_row, year_row=date_row-1,
        # date_row=sumif_row-7 (approximate), month_row=date_row+1
        blocks = []
        for sumif_row in sumif_rows:
            # Detect date_row by looking for date-valued cells nearby
            date_row = self._find_date_row(ws, sumif_row)
            if date_row is None:
                continue

            year_row = date_row - 1
            month_row = date_row + 1

            # Find the column range (C to last non-empty col in date_row)
            cols = []
            for ci in range(1, self.max_cols + 1):
                cl = get_column_letter(ci)
                cell = ws[f"{cl}{date_row}"]
                if cell.value is not None:
                    cols.append(cl)

            if len(cols) < 3:
                continue

            block = Block(
                sheet=sheet_name,
                start_row=year_row,
                end_row=sumif_row,
                sumif_row=sumif_row,
                date_row=date_row,
                year_row=year_row,
                month_row=month_row,
                cols=cols,
            )

            # Record formula status for each cell in date_row
            for cl in cols:
                cell = ws[f"{cl}{date_row}"]
                block.formula_cells[f"{cl}{date_row}"] = is_formula_cell(cell)

            blocks.append(block)

        return blocks

    def _find_date_row(self, ws, sumif_row: int) -> int | None:
        """Find the date row associated with a SUMIF row.

        The date row is typically 1-10 rows above the SUMIF row and contains
        date/datetime values or date-related formulas (YEAR, DATE).
        """
        from datetime import datetime

        for offset in range(1, 15):
            row = sumif_row - offset
            if row < 1:
                break

            # Check multiple columns to identify date row
            has_date = False
            has_formula = False
            for cl in ("C", "D", "E"):
                cell = ws[f"{cl}{row}"]
                if cell.value is None:
                    continue
                if isinstance(cell.value, datetime):
                    has_date = True
                    break
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    has_formula = True
                    # Check if formula references a date cell
                    upper = cell.value.upper()
                    if any(k in upper for k in ("YEAR", "DATE", "MIN", "DATEDIF")):
                        has_date = True
                        break

            if has_date or (has_formula and not has_date):
                # Verify: this row has at least 3 non-empty cells
                non_empty = sum(
                    1 for ci in range(1, 50)
                    if ws.cell(row=row, column=ci).value is not None
                )
                if non_empty >= 3:
                    return row

        return None

    def _check_formula_coverage(
        self, ws, block: Block,
    ) -> list[Defect]:
        """Check if all cells in the date row are formulas.

        Detects cells that should be formulas but are static values.
        """
        defects: list[Defect] = []

        # Find formula cells and static cells in date_row
        formula_cols = []
        static_cols = []

        for cl in block.cols:
            cell = ws[f"{cl}{block.date_row}"]
            if is_formula_cell(cell):
                formula_cols.append(cl)
            elif cell.value is not None:
                # Skip label columns (A, B) — they're not date values
                if cl in ("A", "B"):
                    continue
                # Keep datetime values — these are the static dates we want to flag
                # Skip plain string labels
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    continue
                static_cols.append(cl)

        # If there are both formula and static cells in the same row,
        # the static cells are likely structural defects
        if formula_cols and static_cols:
            static_refs = [f"{cl}{block.date_row}" for cl in static_cols]
            formula_refs = [f"{cl}{block.date_row}" for cl in formula_cols]

            # Check if static cells are BETWEEN formula cells (structural defect)
            # Pattern: first col is formula, last col is formula, middle are static
            has_head = any(
                ws[f"{cl}{block.date_row}"].value
                and is_formula_cell(ws[f"{cl}{block.date_row}"])
                for cl in block.cols[:2]
            )
            has_tail = any(
                is_formula_cell(ws[f"{cl}{block.date_row}"])
                for cl in block.cols[-2:]
            )

            if has_head and static_cols:
                # Build readable position description
                static_str = "、".join(static_cols)
                formula_str = "、".join(formula_cols[:6])
                if len(formula_cols) > 6:
                    formula_str += f" 等 {len(formula_cols)} 列"

                defects.append(Defect(
                    type="static_should_be_formula",
                    severity="critical",
                    sheet=block.sheet,
                    cells=static_refs,
                    description=(
                        f"第 {block.date_row} 行（日期行）中，"
                        f"{static_str} 列为静态值，但相邻的 {formula_str} 列均为公式。"
                        f"当参数（如折旧年限）变化时，静态值不会响应，"
                        f"导致 SUMIF 汇总结果错误。"
                    ),
                    context={
                        "问题": f"第 {block.date_row} 行的 {static_str} 列应为公式但当前是静态值",
                        "影响": "参数变化时这些列不会自动更新，SUMIF 汇总将使用旧日期",
                        "所在块": block.ref,
                        "公式列": formula_refs[:10],
                        "静态值列": static_refs,
                    },
                ))

        return defects

    def _check_block_consistency(self, ws, block: Block) -> list[Defect]:
        """Check if the SUMIF range matches the actual data range."""
        defects: list[Defect] = []

        # Check SUMIF formula range
        for cl in block.cols[:1]:  # Check first SUMIF cell
            cell = ws[f"{cl}{block.sumif_row}"]
            if cell.value and isinstance(cell.value, str) and "SUMIF" in cell.value:
                formula = cell.value
                # Extract range from SUMIF
                range_match = re.search(r"\$([A-Z])\$(\d+):\$([A-Z])\$(\d+)", formula)
                if range_match:
                    range_end_col = range_match.group(3)
                    data_end_col = block.cols[-1] if block.cols else "C"
                    if range_end_col != data_end_col:
                        defects.append(Defect(
                            type="inconsistent_block",
                            severity="warning",
                            sheet=block.sheet,
                            cells=[f"{cl}{block.sumif_row}"],
                            description=(
                                f"SUMIF 公式的求和范围截止到 {range_end_col} 列，"
                                f"但实际数据延伸到 {data_end_col} 列，"
                                f"部分数据未被纳入汇总。所在块：{block.ref}"
                            ),
                        ))

        return defects

    def _scan_date_row_gaps(
        self, ws, sheet_name: str, skip_rows: set[int],
    ) -> list[Defect]:
        """Scan all rows for formula-datetime-formula gaps.

        Catches date rows not associated with any SUMIF block.
        Pattern: col N has formula, cols N+1..M-1 have static datetime,
        col M has formula → static datetimes should be formulas.
        """
        from datetime import datetime

        defects: list[Defect] = []

        for row in range(1, self.max_rows + 1):
            if row in skip_rows:
                continue

            formula_cols: list[str] = []
            datetime_cols: list[str] = []

            for ci in range(1, self.max_cols + 1):
                cl = get_column_letter(ci)
                cell = ws[f"{cl}{row}"]
                if cell.value is None:
                    continue
                if cl in ("A", "B"):
                    continue
                if is_formula_cell(cell):
                    formula_cols.append(cl)
                elif isinstance(cell.value, datetime):
                    datetime_cols.append(cl)

            # Need at least 2 formula cols and datetime cols between them
            if len(formula_cols) < 2 or not datetime_cols:
                continue

            # Find datetime cols that sit between the first and last formula col
            first_ci = column_index_from_string(formula_cols[0])
            last_ci = column_index_from_string(formula_cols[-1])

            gap_cols = [
                dc for dc in datetime_cols
                if first_ci < column_index_from_string(dc) < last_ci
            ]

            if len(gap_cols) < 2:
                continue

            gap_refs = [f"{cl}{row}" for cl in gap_cols]
            formula_refs = [f"{cl}{row}" for cl in formula_cols]

            gap_str = "、".join(gap_cols[:6])
            if len(gap_cols) > 6:
                gap_str += f" 等 {len(gap_cols)} 列"

            defects.append(Defect(
                type="static_should_be_formula",
                severity="critical",
                sheet=sheet_name,
                cells=gap_refs,
                description=(
                    f"第 {row} 行中，{gap_str} 为静态日期值，"
                    f"但首列 {formula_cols[0]} 和末列 {formula_cols[-1]} 均为公式。"
                    f"当参数变化时，中间的静态日期不会更新，导致依赖这些日期的公式计算错误。"
                ),
                context={
                    "问题": f"第 {row} 行的 {gap_str} 应为公式但当前是静态日期值",
                    "影响": "参数变化时这些日期列不会自动更新，下游公式将使用旧日期",
                    "公式边界": f"{formula_cols[0]}{row}（公式）→ {gap_str}（静态）→ {formula_cols[-1]}{row}（公式）",
                    "公式列": formula_refs,
                    "静态值列": gap_refs,
                },
            ))

        return defects


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m financial_kg.diagnostic.structure_checker file.xlsx")
        sys.exit(1)

    checker = ExcelStructureChecker()
    report = checker.check(sys.argv[1])
    print(report.summary())
