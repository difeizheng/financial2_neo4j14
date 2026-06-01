"""Shared utilities for Excel diagnostic tools."""
from __future__ import annotations

import re
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook


def load_dual_mode(path: str) -> tuple[Workbook, Workbook]:
    """Load an Excel file in both formula and data_only modes.

    Returns (wb_formula, wb_values).
    """
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    return wb_formula, wb_values


def is_formula_cell(cell) -> bool:
    """Check if an openpyxl cell contains a formula."""
    if cell.data_type == "f":
        return True
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return True
    return False


def get_cell_value_safe(ws, col: str, row: int) -> Any:
    """Safely get cell value, handling out-of-range and None."""
    try:
        cell = ws[f"{col}{row}"]
        return cell.value
    except Exception:
        return None


# Regex to extract cell references from formulas
# Matches: A1, $A$1, Sheet!A1, 'Sheet Name'!$A$1
_REF_RE = re.compile(
    r"(?:'[^']+'!)?\$?([A-Z]{1,3})\$?(\d+)",
    re.IGNORECASE,
)


def extract_formula_refs(formula: str) -> list[tuple[str, int]]:
    """Extract all cell references (col, row) from a formula string.

    Returns list of (column_letter, row_number) tuples.
    Does NOT expand ranges (A1:B3) — only extracts the literal references.
    """
    if not formula or not isinstance(formula, str):
        return []
    refs = []
    for m in _REF_RE.finditer(formula):
        col = m.group(1).upper()
        row = int(m.group(2))
        refs.append((col, row))
    return refs


def values_equal(a: Any, b: Any, rel_tol: float = 1e-9, abs_tol: float = 1e-9) -> bool:
    """Compare two cell values with tolerance for floating point."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if a == 0 and b == 0:
            return True
        diff = abs(a - b)
        if diff <= abs_tol:
            return True
        denom = max(abs(a), abs(b))
        if denom > 0 and diff / denom <= rel_tol:
            return True
        return False
    from datetime import datetime
    if isinstance(a, datetime) and isinstance(b, datetime):
        return a == b
    return a == b
