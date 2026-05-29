"""Excel Diff Root Cause Analyzer.

Compare two Excel files with linked formulas and automatically find
the root cause of value differences. Detects:
- Formula text changes
- openpyxl cache loss (roundtrip test)
- Static value changes
- Propagation chain breaks

Usage:
    python -m financial_kg.diagnostic.diff_analyzer file_a.xlsx file_b.xlsx
"""
from __future__ import annotations

import os
import sys
import tempfile
from dataclasses import dataclass, field
from typing import Any

import openpyxl

from ._shared import (
    extract_formula_refs,
    is_formula_cell,
    load_dual_mode,
    values_equal,
)


@dataclass
class RootCause:
    type: str          # "formula_changed" | "cache_loss" | "value_changed" | "propagation_break"
    source: str        # Root cause cell reference (e.g. "参数输入表!I217")
    cells: list[str]   # All affected cell references
    path: list[str]    # Propagation path from source
    detail: str        # Human-readable explanation

    def __str__(self) -> str:
        cells_str = ", ".join(self.cells[:10])
        if len(self.cells) > 10:
            cells_str += f" ... ({len(self.cells)} total)"
        return f"[{self.type}] {self.source} → {cells_str}\n  {self.detail}"


@dataclass
class DiffReport:
    file_a: str
    file_b: str
    total_differences: int
    root_causes: list[RootCause]
    roundtrip_lost: int
    tool_is_cause: bool
    differences: list[dict] = field(default_factory=list)

    def summary(self) -> str:
        lines = [
            f"Excel Diff Analysis: {os.path.basename(self.file_a)} vs {os.path.basename(self.file_b)}",
            f"Total differences: {self.total_differences}",
            f"Root causes found: {len(self.root_causes)}",
        ]
        if self.roundtrip_lost > 0:
            lines.append(
                f"WARNING: openpyxl roundtrip loses {self.roundtrip_lost} cached values"
            )
            lines.append("  → Tool (openpyxl) is likely the root cause of differences")
        lines.append("")
        for i, rc in enumerate(self.root_causes, 1):
            lines.append(f"Root Cause #{i}:")
            lines.append(f"  {rc}")
        return "\n".join(lines)


class ExcelDiffAnalyzer:
    """Compare two Excel files and find the root cause of differences."""

    def __init__(self, max_rows: int = 1000, max_cols: int = 50):
        self.max_rows = max_rows
        self.max_cols = max_cols

    def analyze(self, file_a: str, file_b: str) -> DiffReport:
        """Run full analysis comparing file_a (reference) with file_b (modified)."""
        # Load both files in dual mode
        wb_a_f, wb_a_v = load_dual_mode(file_a)
        wb_b_f, wb_b_v = load_dual_mode(file_b)

        # Step 1: Find all value differences
        differences = self._find_value_differences(wb_a_v, wb_b_v)

        # Step 2: Roundtrip test on file_a
        roundtrip_lost = self._roundtrip_test(file_a)

        # Step 3: Classify differences and find root causes
        root_causes = self._classify_differences(
            differences, wb_a_f, wb_b_f, wb_a_v, wb_b_v, roundtrip_lost
        )

        wb_a_f.close()
        wb_a_v.close()
        wb_b_f.close()
        wb_b_v.close()

        return DiffReport(
            file_a=file_a,
            file_b=file_b,
            total_differences=len(differences),
            root_causes=root_causes,
            roundtrip_lost=roundtrip_lost,
            tool_is_cause=roundtrip_lost > 0,
            differences=differences,
        )

    def _find_value_differences(
        self, wb_a_v: openpyxl.Workbook, wb_b_v: openpyxl.Workbook,
    ) -> list[dict]:
        """Find all cells with different cached values."""
        diffs = []
        common_sheets = set(wb_a_v.sheetnames) & set(wb_b_v.sheetnames)

        for sheet_name in common_sheets:
            ws_a = wb_a_v[sheet_name]
            ws_b = wb_b_v[sheet_name]

            for row in ws_a.iter_rows(
                min_row=1, max_row=self.max_rows,
                min_col=1, max_col=self.max_cols,
            ):
                for cell in row:
                    ref = cell.coordinate
                    val_a = cell.value
                    val_b = ws_b[ref].value

                    if val_a is not None and not values_equal(val_a, val_b):
                        diffs.append({
                            "sheet": sheet_name,
                            "ref": ref,
                            "val_a": val_a,
                            "val_b": val_b,
                        })

        return diffs

    def _roundtrip_test(self, file_path: str) -> int:
        """Load-save-load and count cells that lost cached values."""
        wb_orig = openpyxl.load_workbook(file_path, data_only=True)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            wb = openpyxl.load_workbook(file_path)
            wb.save(tmp_path)
            wb.close()

            wb_rt = openpyxl.load_workbook(tmp_path, data_only=True)

            lost = 0
            for sn in wb_orig.sheetnames:
                if sn not in wb_rt.sheetnames:
                    continue
                ws_o = wb_orig[sn]
                ws_r = wb_rt[sn]
                for row in ws_o.iter_rows(
                    min_row=1, max_row=self.max_rows,
                    min_col=1, max_col=self.max_cols,
                ):
                    for cell in row:
                        if cell.value is not None:
                            rt_val = ws_r[cell.coordinate].value
                            if rt_val is None:
                                lost += 1

            wb_rt.close()
        finally:
            os.unlink(tmp_path)

        wb_orig.close()
        return lost

    def _classify_differences(
        self,
        differences: list[dict],
        wb_a_f: openpyxl.Workbook, wb_b_f: openpyxl.Workbook,
        wb_a_v: openpyxl.Workbook, wb_b_v: openpyxl.Workbook,
        roundtrip_lost: int,
    ) -> list[RootCause]:
        """Classify each difference and group into root causes."""
        formula_changed_group: dict[str, list[str]] = {}
        cache_loss_cells: list[str] = []
        value_changed_cells: list[str] = []
        propagation_roots: dict[str, list[str]] = {}

        for diff in differences:
            sheet = diff["sheet"]
            ref = diff["ref"]
            full_ref = f"{sheet}!{ref}"

            ws_a_f = wb_a_f[sheet]
            ws_b_f = wb_b_f[sheet]

            cell_a_f = ws_a_f[ref]
            cell_b_f = ws_b_f[ref]

            is_formula_a = is_formula_cell(cell_a_f)
            is_formula_b = is_formula_cell(cell_b_f)

            # Case 1: formula text changed
            if is_formula_a and is_formula_b:
                if cell_a_f.value != cell_b_f.value:
                    formula_changed_group.setdefault(
                        sheet, []
                    ).append(full_ref)
                    continue

            # Case 2: both are formulas, same text, different cached values
            if is_formula_a and is_formula_b and cell_a_f.value == cell_b_f.value:
                if roundtrip_lost > 0:
                    cache_loss_cells.append(full_ref)
                else:
                    # Trace dependency chain to find root
                    root = self._trace_root(
                        sheet, ref, wb_a_f, wb_a_v, wb_b_v, set(),
                    )
                    propagation_roots.setdefault(root, []).append(full_ref)
                continue

            # Case 3: one is formula, one is value (structure change)
            if is_formula_a != is_formula_b:
                formula_changed_group.setdefault(sheet, []).append(full_ref)
                continue

            # Case 4: both are plain values, different
            value_changed_cells.append(full_ref)

        # Build root cause list
        causes: list[RootCause] = []

        for sheet, cells in formula_changed_group.items():
            causes.append(RootCause(
                type="formula_changed",
                source=f"{sheet} ({len(cells)} cells)",
                cells=cells,
                path=[],
                detail=(
                    f"Formula text changed in {len(cells)} cells on sheet '{sheet}'. "
                    f"These are intentional modifications to the sheet."
                ),
            ))

        if cache_loss_cells:
            causes.append(RootCause(
                type="cache_loss",
                source="openpyxl",
                cells=cache_loss_cells[:20],
                path=[],
                detail=(
                    f"openpyxl strips formula cached values on save. "
                    f"{len(cache_loss_cells)} cells affected. "
                    f"Roundtrip test lost {roundtrip_lost} cached values. "
                    f"Use XML-level manipulation (zip+lxml) to preserve caches."
                ),
            ))

        if value_changed_cells:
            causes.append(RootCause(
                type="value_changed",
                source=value_changed_cells[0],
                cells=value_changed_cells[:20],
                path=[],
                detail=f"Static values changed in {len(value_changed_cells)} cells",
            ))

        for root, cells in propagation_roots.items():
            causes.append(RootCause(
                type="propagation_break",
                source=root,
                cells=cells,
                path=[root],
                detail=f"Difference propagates from {root}",
            ))

        return causes

    def _trace_root(
        self,
        sheet: str, ref: str,
        wb_a_f: openpyxl.Workbook,
        wb_a_v: openpyxl.Workbook,
        wb_b_v: openpyxl.Workbook,
        visited: set[str],
        depth: int = 0,
    ) -> str:
        """Recursively trace formula chain to find the root cause cell."""
        key = f"{sheet}!{ref}"
        if key in visited or depth > 5:
            return key
        visited.add(key)

        ws_f = wb_a_f[sheet]
        cell = ws_f[ref]

        if not is_formula_cell(cell) or not isinstance(cell.value, str):
            return key

        # Extract references from formula
        refs = extract_formula_refs(cell.value)

        # Check each referenced cell
        for col, row in refs:
            ref_key = f"{sheet}!{col}{row}"
            if ref_key in visited:
                continue

            val_a = wb_a_v[sheet][f"{col}{row}"].value
            val_b = wb_b_v[sheet][f"{col}{row}"].value

            if not values_equal(val_a, val_b):
                # Found a deeper difference — recurse
                return self._trace_root(
                    sheet, f"{col}{row}", wb_a_f, wb_a_v, wb_b_v,
                    visited, depth + 1,
                )

        # No deeper difference found — this cell is the root
        return key


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python -m financial_kg.diagnostic.diff_analyzer file_a.xlsx file_b.xlsx")
        sys.exit(1)

    analyzer = ExcelDiffAnalyzer()
    report = analyzer.analyze(sys.argv[1], sys.argv[2])
    print(report.summary())
