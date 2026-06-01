"""Auto-fix for Excel structure defects (static datetime → formula).

Replaces static datetime values in date rows with
MIN(DATE(YEAR(prev),12,31), end_ref) formulas, using XML-level
manipulation to preserve all other cells' cached values.
"""
from __future__ import annotations

import re
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from ._shared import extract_formula_refs
from .structure_checker import Defect

_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_MAP = {"s": _NS}


def fix_structure_defects(file_path: str, defects: list[Defect]) -> str:
    """Fix static-should-be-formula defects. Returns path to fixed file."""
    wb_f = load_workbook(file_path)
    wb_v = load_workbook(file_path, data_only=True)

    sheet_fixes: dict[str, list[tuple[int, int, str]]] = defaultdict(list)
    for defect in defects:
        if defect.type != "static_should_be_formula":
            continue
        fixes = _generate_fix_formulas(wb_f, wb_v, defect)
        if fixes:
            sheet_fixes[defect.sheet].extend(fixes)

    sheet_names = list(wb_f.sheetnames)
    wb_f.close()
    wb_v.close()

    if not sheet_fixes:
        raise ValueError("没有可修复的缺陷")

    return _apply_xml_fixes(file_path, sheet_names, sheet_fixes)


def _generate_fix_formulas(
    wb_f, wb_v, defect: Defect,
) -> list[tuple[int, int, str]]:
    """Generate (row, col_idx, formula_text) fixes for a defect."""
    sheet = defect.sheet
    ws_f = wb_f[sheet]
    ws_v = wb_v[sheet]

    ctx = defect.context

    # Parse row from "第 N 行" in description
    row_match = re.search(r"第 (\d+) 行", defect.description)
    if not row_match:
        return []
    row = int(row_match.group(1))

    # Parse column letters from references
    formula_refs = ctx.get("公式列", [])
    static_refs = ctx.get("静态值列", [])
    if not formula_refs or not static_refs:
        return []

    formula_cols = sorted(
        {re.match(r"([A-Z]+)", r).group(1) for r in formula_refs if re.match(r"([A-Z]+)", r)},
        key=lambda c: column_index_from_string(c),
    )
    static_cols = sorted(
        {re.match(r"([A-Z]+)", r).group(1) for r in static_refs if re.match(r"([A-Z]+)", r)},
        key=lambda c: column_index_from_string(c),
    )
    if not formula_cols or not static_cols:
        return []

    # Find end date reference from rightmost formula cell
    rightmost_col = formula_cols[-1]
    rightmost_formula = ws_f[f"{rightmost_col}{row}"].value
    if not rightmost_formula:
        return []
    end_ref = _find_end_ref(rightmost_formula, row)
    if not end_ref:
        return []

    # Generate formula for each static column
    fixes = []
    is_prev_year_end = False

    for i, col in enumerate(static_cols):
        ci = column_index_from_string(col)
        prev_col = get_column_letter(ci - 1)

        if i == 0:
            # First gap: check left boundary cell's cached date
            prev_val = ws_v[f"{prev_col}{row}"].value
            if isinstance(prev_val, datetime):
                is_prev_year_end = (prev_val.month == 12 and prev_val.day == 31)

        if is_prev_year_end:
            year_expr = f"YEAR({prev_col}{row})+1"
        else:
            year_expr = f"YEAR({prev_col}{row})"

        formula = f"MIN(DATE({year_expr},12,31),{end_ref})"
        fixes.append((row, ci, formula))

        # After this fix, the cell will be Dec 31 (year-end)
        is_prev_year_end = True

    return fixes


def _find_end_ref(formula: str, current_row: int) -> str | None:
    """Extract end date reference (cross-row) from rightmost formula."""
    text = formula.lstrip("=")
    refs = extract_formula_refs(text)

    for col, row in refs:
        if row != current_row:
            return f"{col}${row}"

    if refs:
        col, row = refs[-1]
        return f"{col}${row}"

    return None


# ── XML-level fix application ─────────────────────────────────────────────


def _apply_xml_fixes(
    file_path: str,
    sheet_names: list[str],
    sheet_fixes: dict[str, list[tuple[int, int, str]]],
) -> str:
    """Apply formula fixes via XML manipulation, preserving all cached values."""
    suffix = Path(file_path).suffix
    with NamedTemporaryFile(suffix=suffix, delete=False) as f:
        output_path = f.name

    with zipfile.ZipFile(file_path, "r") as zin:
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                sheet_match = re.match(r"xl/worksheets/sheet(\d+)\.xml", item.filename)
                if sheet_match:
                    idx = int(sheet_match.group(1)) - 1
                    if idx < len(sheet_names):
                        name = sheet_names[idx]
                        if name in sheet_fixes:
                            data = _fix_sheet_xml(data, sheet_fixes[name])

                zout.writestr(item, data)

    return output_path


def _fix_sheet_xml(xml_data: bytes, fixes: list[tuple[int, int, str]]) -> bytes:
    """Apply fixes to a single sheet's XML."""
    root = etree.fromstring(xml_data)

    # Build fix lookup: row -> {col_letter: formula}
    fix_map: dict[int, dict[str, str]] = defaultdict(dict)
    for row, col_idx, formula in fixes:
        fix_map[row][get_column_letter(col_idx)] = formula

    for row_elem in root.iter(f"{{{_NS}}}row"):
        row_num = int(row_elem.get("r", "0"))
        if row_num not in fix_map:
            continue

        row_fixes = fix_map[row_num]

        # Collect style from an adjacent formula cell
        template_style = _find_formula_style(row_elem)

        for cell_elem in row_elem:
            ref = cell_elem.get("r", "")
            ref_match = re.match(r"([A-Z]+)(\d+)", ref)
            if not ref_match:
                continue
            col_letter = ref_match.group(1)

            if col_letter in row_fixes:
                _set_cell_formula(cell_elem, row_fixes[col_letter], template_style)

    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _find_formula_style(row_elem) -> str | None:
    """Find style ID from any formula cell in this row."""
    for cell_elem in row_elem:
        f_child = cell_elem.find(f"{{{_NS}}}f")
        if f_child is not None:
            style = cell_elem.get("s")
            if style:
                return style
    return None


def _set_cell_formula(cell_elem, formula: str, style_id: str | None = None) -> None:
    """Replace a cell's content with a formula."""
    # Remove existing <f> and <v> elements
    for tag in ("f", "v"):
        for child in list(cell_elem):
            if child.tag == f"{{{_NS}}}{tag}":
                cell_elem.remove(child)

    # Add formula element
    f_elem = etree.SubElement(cell_elem, f"{{{_NS}}}f")
    f_elem.text = formula

    # Set style
    if style_id is not None and "s" not in cell_elem.attrib:
        cell_elem.set("s", style_id)

    # Remove type attribute for numeric/date formula results
    if cell_elem.get("t") in ("n", None):
        if "t" in cell_elem.attrib:
            del cell_elem.attrib["t"]
