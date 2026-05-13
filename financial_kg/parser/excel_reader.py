from __future__ import annotations
import re
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from ..models.cell import CellData


# Detect if a cell value looks like a formula string stored as value
_FORMULA_RE = re.compile(r"^=")


def _infer_data_type(cell) -> str:
    """Infer data type from an openpyxl cell."""
    if cell.data_type == "n":
        return "number"
    if cell.data_type == "s":
        return "string"
    if cell.data_type == "d":
        return "date"
    if cell.data_type == "b":
        return "bool"
    if cell.data_type == "f" or (
        isinstance(cell.value, str) and _FORMULA_RE.match(cell.value)
    ):
        return "formula"
    return "string"


def read_excel(filepath: str) -> dict[str, list[CellData]]:
    """Read all non-empty cells from an Excel file.

    Returns a dict mapping sheet_name -> list[CellData].
    Merged cells: the top-left cell carries the value; all other cells in the
    merge group record is_merged=True and merge_parent_id pointing to the
    top-left cell.
    """
    # Load twice: once for formulas, once for computed values
    wb_formula = openpyxl.load_workbook(filepath, data_only=False)
    wb_values = openpyxl.load_workbook(filepath, data_only=True)

    result: dict[str, list[CellData]] = {}

    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]

        # Build merge map: cell_coord -> parent_coord (top-left of merge group)
        merge_map: dict[str, str] = {}
        # Also track merge range info for the top-left cell
        merge_info: dict[str, tuple[int, str]] = {}  # parent_coord -> (end_row, end_col)
        for merge_range in ws_formula.merged_cells.ranges:
            min_row, min_col = merge_range.min_row, merge_range.min_col
            max_row, max_col = merge_range.max_row, merge_range.max_col
            parent_coord = f"{get_column_letter(min_col)}{min_row}"
            merge_info[parent_coord] = (max_row, get_column_letter(max_col))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    coord = f"{get_column_letter(col)}{row}"
                    if coord != parent_coord:
                        merge_map[coord] = parent_coord

        cells: list[CellData] = []

        for row in ws_formula.iter_rows():
            for cell_f in row:
                if cell_f.value is None:
                    continue

                col_letter = get_column_letter(cell_f.column)
                coord = f"{col_letter}{cell_f.row}"

                # Get computed value from the values workbook
                cell_v = ws_values[coord]
                computed_value = cell_v.value

                # Determine formula vs value
                formula_raw: Optional[str] = None
                if isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    formula_raw = cell_f.value
                    data_type = "formula"
                else:
                    data_type = _infer_data_type(cell_f)
                    computed_value = cell_f.value  # use formula-wb value for non-formula cells

                is_merged = coord in merge_map
                merge_parent_id: Optional[str] = None
                if is_merged:
                    parent_coord = merge_map[coord]
                    p_col, p_row = coordinate_from_string(parent_coord)
                    merge_parent_id = f"{sheet_name}_{p_row}_{p_col}"

                # Merge range info for top-left cell
                merge_end_row: Optional[int] = None
                merge_end_col: Optional[str] = None
                if coord in merge_info:
                    merge_end_row, merge_end_col = merge_info[coord]

                cells.append(CellData(
                    sheet=sheet_name,
                    row=cell_f.row,
                    col=col_letter,
                    value=computed_value,
                    formula_raw=formula_raw,
                    data_type=data_type,
                    is_merged=is_merged,
                    merge_parent_id=merge_parent_id,
                    merge_end_row=merge_end_row,
                    merge_end_col=merge_end_col,
                    number_format=cell_f.number_format or None,
                ))

        result[sheet_name] = cells

    wb_formula.close()
    wb_values.close()
    return result
