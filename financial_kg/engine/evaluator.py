"""Formula evaluator: re-evaluates a cell's formula using the formulas library.

The `formulas` library expects inputs keyed by raw Excel reference strings
(e.g. 'F5', '$I$250', 'F5:BE5', '参数输入表!I250').  Our cell IDs use the
format "{sheet}_{row}_{col}".  This module bridges the two representations.

Performance notes
-----------------
- Compiled formula objects are cached **persistently** across recalculation
  sessions (keyed by formula string).
- Input plans (structural mapping: which cell_ids a formula references) are
  cached per cell_id.  Only values are re-read on each evaluation — no
  repeated string parsing, sheet normalization, or range expansion.
- Simple formula patterns (binops, SUM, negation) use a fast regex path
  that bypasses the formulas library entirely.
- clear_formula_cache() clears both formula and input plan caches.  Call
  only when a new workbook is loaded.
"""
from __future__ import annotations

import operator
import re
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Optional

import numpy as np

try:
    from formulas import Parser as _FormulaParser
    _PARSER = _FormulaParser()
    _FORMULAS_AVAILABLE = True
except ImportError:
    _FORMULAS_AVAILABLE = False


from financial_kg.models.graph import FinancialGraph


# ── Sheet name mapping for Excel formula quirks ───────────────────────────────
_SHEET_NAME_ALIASES = {
    '表1-式样及构造信息表': '表1-资金筹措及还本付息表',
    '表1-资金筹措及还本付息信息表': '表1-资金筹措及还本付息表',
}

# ── Cached sheet-name lookup ─────────────────────────────────────────────────
_SHEETS_CACHE: list = [-1, set()]


def _get_actual_sheets(graph: FinancialGraph) -> set[str]:
    gid = id(graph)
    if _SHEETS_CACHE[0] == gid:
        return _SHEETS_CACHE[1]
    sheets: set[str] = set()
    for cid in graph.cells.keys():
        parts = cid.rsplit('_', 2)
        if len(parts) == 3:
            sheets.add(parts[0])
    _SHEETS_CACHE[0] = gid
    _SHEETS_CACHE[1] = sheets
    return sheets


def _normalize_sheet_name(sheet: str, graph: FinancialGraph) -> str:
    if sheet in _SHEET_NAME_ALIASES:
        return _SHEET_NAME_ALIASES[sheet]
    actual_sheets = _get_actual_sheets(graph)
    if sheet in actual_sheets:
        return sheet
    for actual in actual_sheets:
        if sheet.split('-')[0] == actual.split('-')[0] if '-' in sheet else False:
            return actual
    return sheet


# ── Address helpers ───────────────────────────────────────────────────────────

def _split_col_row(addr: str):
    m = re.match(r"([A-Za-z]+)(\d+)", addr)
    if not m:
        raise ValueError(f"Cannot parse cell address: {addr!r}")
    return m.group(1).upper(), m.group(2)


def _addr_to_cell_id(sheet: str, addr: str) -> str:
    col, row = _split_col_row(addr)
    return f"{sheet}_{row}_{col}"


def _coerce(val: Any) -> Any:
    if val is None:
        return 0.0
    if isinstance(val, str) and val in ('#NUM!', '#VALUE!', '#DIV/0!', '#REF!', '#N/A'):
        return val
    if isinstance(val, str) and 'T00:00:00' in val:
        try:
            dt = datetime.fromisoformat(val.replace('T00:00:00', ''))
            excel_epoch = datetime(1899, 12, 30)
            return float((dt - excel_epoch).days)
        except Exception:
            pass
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        return val
    return str(val)


def _cell_id_to_ref(cell_id: str, formula_sheet: str) -> str:
    parts = cell_id.rsplit("_", 2)
    if len(parts) != 3:
        return cell_id
    sheet, row, col = parts
    ref = f"{col}{row}"
    if sheet != formula_sheet:
        ref = f"{sheet}!{ref}"
    return ref


# ══════════════════════════════════════════════════════════════════════════════
# Input Plan: cached structural mapping (which cells a formula references)
# ══════════════════════════════════════════════════════════════════════════════

_SENTINEL = object()  # marker for missing cells


@dataclass
class _RangePlan:
    cell_ids: list[str]
    shape: tuple[int, int]  # (rows, cols)


@dataclass
class _InputPlan:
    scalar_inputs: list[tuple[str, str]]          # [(input_key, cell_id), ...]
    range_inputs: list[tuple[str, _RangePlan]]    # [(input_key, plan), ...]
    # Pre-allocated numpy array templates for ranges (filled by _build_fast)
    _range_arrays: dict[int, np.ndarray] = field(default_factory=dict, repr=False)

    def __post_init__(self):
        # Pre-allocate numpy arrays with zeros for each range
        for i, (_, rp) in enumerate(self.range_inputs):
            self._range_arrays[i] = np.zeros(rp.shape, dtype=np.float64)


_input_plans: dict[str, _InputPlan] = {}


def _build_input_plan(
    cell_id: str,
    func_inputs,
    formula_sheet: str,
    graph: FinancialGraph,
) -> _InputPlan:
    if cell_id in _input_plans:
        return _input_plans[cell_id]

    scalars: list[tuple[str, str]] = []
    ranges: list[tuple[str, _RangePlan]] = []

    for raw_key in func_inputs:
        if "!" in raw_key:
            sheet_part, addr_part = raw_key.split("!", 1)
            sheet_part = sheet_part.strip("'")
            sheet_part = _normalize_sheet_name(sheet_part, graph)
        else:
            sheet_part = formula_sheet
            addr_part = raw_key

        addr_part = addr_part.replace("$", "")

        if ":" in addr_part:
            start, end = addr_part.split(":", 1)
            start_col, start_row = _split_col_row(start)
            end_col, end_row = _split_col_row(end)

            from openpyxl.utils import column_index_from_string, get_column_letter
            sc = column_index_from_string(start_col)
            ec = column_index_from_string(end_col)
            sr, er = int(start_row), int(end_row)

            cell_ids = []
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    cell_ids.append(f"{sheet_part}_{r}_{get_column_letter(c)}")

            ranges.append((raw_key, _RangePlan(
                cell_ids=cell_ids,
                shape=(er - sr + 1, ec - sc + 1),
            )))
        else:
            cid = _addr_to_cell_id(sheet_part, addr_part)
            scalars.append((raw_key, cid))

    plan = _InputPlan(scalar_inputs=scalars, range_inputs=ranges)
    _input_plans[cell_id] = plan
    return plan


def _build_input_map_fast(plan: _InputPlan, graph: FinancialGraph) -> dict[str, np.ndarray]:
    kwargs: dict[str, np.ndarray] = {}

    for key, cell_id in plan.scalar_inputs:
        cell = graph.cells.get(cell_id)
        kwargs[key] = np.array([[_coerce(cell.value if cell else None)]])

    for i, (key, rp) in enumerate(plan.range_inputs):
        arr = plan._range_arrays[i]
        idx = 0
        for r in range(rp.shape[0]):
            for c in range(rp.shape[1]):
                cell = graph.cells.get(rp.cell_ids[idx])
                v = _coerce(cell.value if cell else None)
                arr[r, c] = float(v) if isinstance(v, (int, float)) else 0.0
                idx += 1
        kwargs[key] = arr

    return kwargs


# ── Legacy resolver (used by non-plan paths) ──────────────────────────────────

def _resolve_input_key(raw_key: str, formula_sheet: str, graph: FinancialGraph) -> np.ndarray:
    if "!" in raw_key:
        sheet_part, addr_part = raw_key.split("!", 1)
        sheet_part = sheet_part.strip("'")
        sheet_part = _normalize_sheet_name(sheet_part, graph)
    else:
        sheet_part = formula_sheet
        addr_part = raw_key

    addr_part = addr_part.replace("$", "")

    if ":" in addr_part:
        return _resolve_range(sheet_part, addr_part, graph)

    cell_id = _addr_to_cell_id(sheet_part, addr_part)
    cell = graph.cells.get(cell_id)
    return np.array([[_coerce(cell.value if cell else None)]])


def _resolve_range(sheet: str, addr: str, graph: FinancialGraph) -> np.ndarray:
    sheet = _normalize_sheet_name(sheet, graph)
    start, end = addr.split(":", 1)
    start_col, start_row = _split_col_row(start)
    end_col, end_row = _split_col_row(end)

    from openpyxl.utils import column_index_from_string, get_column_letter
    sc = column_index_from_string(start_col)
    ec = column_index_from_string(end_col)
    sr, er = int(start_row), int(end_row)

    rows = []
    for r in range(sr, er + 1):
        row_vals = []
        for c in range(sc, ec + 1):
            col_letter = get_column_letter(c)
            cell_id = f"{sheet}_{r}_{col_letter}"
            cell = graph.cells.get(cell_id)
            row_vals.append(_coerce(cell.value if cell else None))
        rows.append(row_vals)
    return np.array(rows)


# ══════════════════════════════════════════════════════════════════════════════
# Fast path: simple formula patterns bypass formulas library
# ══════════════════════════════════════════════════════════════════════════════

_MISS = object()

_SIMPLE_BINOPS = {
    '+': operator.add, '-': operator.sub,
    '*': operator.mul, '/': operator.truediv,
}

# Pre-compiled regex patterns for fast-path matching
_RE_NEGATION = re.compile(r'^=-([A-Za-z一-鿿][\w一-鿿]*!?\$?[A-Z]+\$?\d+)$')
_RE_BINOP_CELLS = re.compile(
    r'^=([A-Za-z一-鿿][\w一-鿿]*!?\$?[A-Z]+\$?\d+)'
    r'([+\-*/])'
    r'([A-Za-z一-鿿][\w一-鿿]*!?\$?[A-Z]+\$?\d+)$'
)
_RE_BINOP_SCALAR_R = re.compile(
    r'^=([A-Za-z一-鿿][\w一-鿿]*!?\$?[A-Z]+\$?\d+)'
    r'([*/])(-?[\d.]+)$'
)
_RE_BINOP_SCALAR_L = re.compile(
    r'^=(-?[\d.]+)'
    r'([*/])'
    r'([A-Za-z一-鿿][\w一-鿿]*!?\$?[A-Z]+\$?\d+)$'
)
_RE_SUM_RANGE = re.compile(r'^=SUM\(([A-Z]+\$?\d+:[A-Z]+\$?\d+)\)$', re.IGNORECASE)
_RE_SUM_SHEET_RANGE = re.compile(
    r"^=SUM\('?([^!']+)'?!([A-Z]+\$?\d+:[A-Z]+\$?\d+)\)$", re.IGNORECASE
)

# Additional fast-path patterns (v4.3.0)
# _REF: matches A1, $A$1, Sheet!A1, '表1'!$A$1
_REF = r'(?:[\w一-鿝][\w一-鿝]*!)?\$?[A-Z]+\$?\d+'
_RE_YEAR = re.compile(r'^=YEAR\((' + _REF + r')\)$', re.IGNORECASE)
_RE_ABS = re.compile(r'^=ABS\((' + _REF + r')\)$', re.IGNORECASE)
_RE_ROUND = re.compile(r'^=ROUND\((' + _REF + r'),\s*(-?\d+)\)$', re.IGNORECASE)
# MAX(0, ref-ref): =MAX(0, F65-F75) — common for max-with-zero pattern
_RE_MAX_ZERO_REF = re.compile(
    r'^=MAX\(0,\s*(' + _REF + r')-(' + _REF + r')\)$', re.IGNORECASE
)

# ── DATEDIF fast path: formulas library returns #NUM!/#VALUE! for it ──────────
# The library's xdatedif fails on lowercase unit codes ("d" vs "D") and
# wrap_ufunc converts the string unit argument to float.  Handle in fast path.
_DATEDIF_EPOCH = datetime(1899, 12, 30)

_RE_DATEDIF = re.compile(
    r'^=DATEDIF\(\s*(' + _REF + r')\s*,\s*(' + _REF
    + r')\s*,\s*"([DdMmYy]{1,2})"\s*\)$',
    re.IGNORECASE,
)
# =ROUND((DATEDIF(C19,D19,"d"))/30,0)  or  =ROUND(DATEDIF(I5,I7,"D")/365*12,0)
_RE_DATEDIF_ROUNDED = re.compile(
    r'^=ROUND\(\s*\(?\s*DATEDIF\(\s*(' + _REF + r')\s*,\s*(' + _REF
    + r')\s*,\s*"([DdMmYy]{1,2})"\s*\)\s*\)?\s*/\s*(-?[\d.]+)'
    r'(?:\s*\*\s*(-?[\d.]+))?\s*,\s*(-?\d+)\s*\)$',
    re.IGNORECASE,
)


def _datedif_calc(sd_serial: float, ed_serial: float, unit: str) -> float | None:
    """Compute DATEDIF between two Excel serial dates."""
    unit = unit.upper()
    if sd_serial > ed_serial:
        return None
    if unit == 'D':
        return ed_serial - sd_serial
    dt_s = _DATEDIF_EPOCH + timedelta(days=int(sd_serial))
    dt_e = _DATEDIF_EPOCH + timedelta(days=int(ed_serial))
    sy, sm, sd_ = dt_s.year, dt_s.month, dt_s.day
    ey, em, ed_ = dt_e.year, dt_e.month, dt_e.day
    if unit == 'Y':
        return ey - sy - int((em, ed_) < (sm, sd_))
    if unit == 'M':
        return (ey - sy) * 12 + (em - sm) - int(ed_ < sd_)
    if unit == 'MD':
        if ed_ < sd_:
            prev_month = dt_e.replace(day=1) - timedelta(days=1)
            return (ed_ + prev_month.day) - sd_
        return ed_ - sd_
    if unit == 'YM':
        return (em - sm - int(ed_ < sd_)) % 12
    if unit == 'YD':
        try:
            return (dt_e.replace(year=sy) - dt_s).days
        except ValueError:
            return (dt_e.replace(year=sy, month=2, day=28) - dt_s).days
    return None
# AVERAGE with quoted sheet prefix: =AVERAGE('表名'!H12:H20)
_RE_AVERAGE_QUOTED = re.compile(
    r"^=AVERAGE\('([^']+)'!([A-Z]+\$?\d+:[A-Z]+\$?\d+)\)$", re.IGNORECASE
)
# IF: comparison between cell ref and value/ref, then two value/ref branches
_RE_IF_SIMPLE = re.compile(
    r'^=IF\((' + _REF + r')([><=!]+)(-?[\d.]+|' + _REF + r'),\s*'
    r'(-?[\d.]+|' + _REF + r'),\s*'
    r'(-?[\d.]+|' + _REF + r')\)$',
    re.IGNORECASE
)


def _excel_serial_to_year(serial: float) -> int | None:
    """Convert Excel serial date to year. Excel epoch = 1899-12-30."""
    try:
        excel_epoch = datetime(1899, 12, 30)
        dt = excel_epoch + timedelta(days=int(serial))
        return dt.year
    except (ValueError, OverflowError):
        return None


def _read_cell_value(ref: str, formula_sheet: str, graph: FinancialGraph) -> Any:
    """Read a single cell value by Excel reference string."""
    if "!" in ref:
        sheet, addr = ref.split("!", 1)
        sheet = sheet.strip("'")
        sheet = _normalize_sheet_name(sheet, graph)
    else:
        sheet = formula_sheet
        addr = ref
    addr = addr.replace("$", "")
    cell_id = _addr_to_cell_id(sheet, addr)
    cell = graph.cells.get(cell_id)
    return _coerce(cell.value if cell else None)


def _try_range_agg(
    formula_raw: str, plan: _InputPlan, graph: FinancialGraph, op: str,
) -> Any:
    """Aggregate numeric values from all range inputs in the plan."""
    vals: list[float] = []
    if not plan.range_inputs:
        return _MISS
    for _, rp in plan.range_inputs:
        for cid in rp.cell_ids:
            cell = graph.cells.get(cid)
            v = _coerce(cell.value if cell else None)
            if isinstance(v, (int, float)):
                vals.append(float(v))
            else:
                return _MISS  # non-numeric in range → fallback
    if not vals:
        return _MISS
    if op == 'max':
        return max(vals)
    elif op == 'min':
        return min(vals)
    elif op == 'average':
        return sum(vals) / len(vals)
    return _MISS


def _try_range_agg_sheet(
    sheet_name: str, range_str: str, graph: FinancialGraph, op: str,
) -> Any:
    """Aggregate values from an explicit sheet!range reference."""
    from openpyxl.utils import column_index_from_string, get_column_letter
    start, end = range_str.split(":", 1)
    sc_letter, sr = _split_col_row(start)
    ec_letter, er = _split_col_row(end)
    sc = column_index_from_string(sc_letter)
    ec = column_index_from_string(ec_letter)
    sr, er = int(sr), int(er)

    vals: list[float] = []
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            col = get_column_letter(c)
            cid = f"{sheet_name}_{r}_{col}"
            cell = graph.cells.get(cid)
            v = _coerce(cell.value if cell else None)
            if isinstance(v, (int, float)):
                vals.append(float(v))
            else:
                return _MISS
    if not vals:
        return _MISS
    if op == 'max':
        return max(vals)
    elif op == 'min':
        return min(vals)
    elif op == 'average':
        return sum(vals) / len(vals)
    return _MISS


def _try_fast_eval(
    formula_raw: str, formula_sheet: str, graph: FinancialGraph, plan: _InputPlan,
) -> Any:
    """Evaluate simple formula patterns without the formulas library.

    Returns _MISS if the pattern doesn't match any fast path.
    """
    f = formula_raw.lstrip('=')

    # ── Negation: =-REF ──
    m = _RE_NEGATION.match(formula_raw)
    if m:
        val = _read_cell_value(m.group(1), formula_sheet, graph)
        if isinstance(val, (int, float)):
            return -float(val)
        return _MISS

    # ── Binary op: =REF+REF, =REF-REF, =REF*REF, =REF/REF ──
    m = _RE_BINOP_CELLS.match(formula_raw)
    if m:
        a = _read_cell_value(m.group(1), formula_sheet, graph)
        b = _read_cell_value(m.group(3), formula_sheet, graph)
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            try:
                return _SIMPLE_BINOPS[m.group(2)](float(a), float(b))
            except ZeroDivisionError:
                return _MISS
        return _MISS

    # ── Binary op: =REF*scalar, =REF/scalar ──
    m = _RE_BINOP_SCALAR_R.match(formula_raw)
    if m:
        a = _read_cell_value(m.group(1), formula_sheet, graph)
        if isinstance(a, (int, float)):
            try:
                return _SIMPLE_BINOPS[m.group(2)](float(a), float(m.group(3)))
            except ZeroDivisionError:
                return _MISS
        return _MISS

    # ── Binary op: =scalar*REF, =scalar/REF ──
    m = _RE_BINOP_SCALAR_L.match(formula_raw)
    if m:
        b = _read_cell_value(m.group(3), formula_sheet, graph)
        if isinstance(b, (int, float)):
            try:
                return _SIMPLE_BINOPS[m.group(2)](float(m.group(1)), float(b))
            except ZeroDivisionError:
                return _MISS
        return _MISS

    # ── SUM(range): =SUM(F5:BA5) ──
    m = _RE_SUM_RANGE.match(formula_raw)
    if m:
        if plan.range_inputs:
            total = 0.0
            for _, rp in plan.range_inputs:
                for cid in rp.cell_ids:
                    cell = graph.cells.get(cid)
                    v = _coerce(cell.value if cell else None)
                    if not isinstance(v, (int, float)):
                        return _MISS
                    total += float(v)
            return total
        return _MISS

    # ── SUM(sheet!range): =SUM('表1'!F5:BA5) ──
    m = _RE_SUM_SHEET_RANGE.match(formula_raw)
    if m:
        if plan.range_inputs:
            total = 0.0
            for _, rp in plan.range_inputs:
                for cid in rp.cell_ids:
                    cell = graph.cells.get(cid)
                    v = _coerce(cell.value if cell else None)
                    if not isinstance(v, (int, float)):
                        return _MISS
                    total += float(v)
            return total
        return _MISS

    # ── YEAR(ref): =YEAR(A1) ──
    m = _RE_YEAR.match(formula_raw)
    if m:
        val = _read_cell_value(m.group(1), formula_sheet, graph)
        if isinstance(val, (int, float)):
            year = _excel_serial_to_year(float(val))
            return year if year else _MISS
        return _MISS

    # ── ABS(ref): =ABS(A1) ──
    m = _RE_ABS.match(formula_raw)
    if m:
        val = _read_cell_value(m.group(1), formula_sheet, graph)
        if isinstance(val, (int, float)):
            return abs(float(val))
        return _MISS

    # ── ROUND(ref, ndigits): =ROUND(A1, 2) ──
    m = _RE_ROUND.match(formula_raw)
    if m:
        val = _read_cell_value(m.group(1), formula_sheet, graph)
        if isinstance(val, (int, float)):
            return round(float(val), int(m.group(2)))
        return _MISS

    # ── MAX(0, ref-ref): =MAX(0, F65-F75) ──
    m = _RE_MAX_ZERO_REF.match(formula_raw)
    if m:
        a = _read_cell_value(m.group(1), formula_sheet, graph)
        b = _read_cell_value(m.group(2), formula_sheet, graph)
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            return max(0.0, float(a) - float(b))
        return _MISS

    # ── AVERAGE('sheet'!range): =AVERAGE('投产&达产比例'!H12:H20) ──
    m = _RE_AVERAGE_QUOTED.match(formula_raw)
    if m:
        sheet_name = m.group(1)
        range_str = m.group(2)
        result = _try_range_agg_sheet(sheet_name, range_str, graph, 'average')
        if result is not _MISS:
            return result

    # ── ROUND(DATEDIF(ref,ref,"d")/N, 0): very common pattern for month counting ──
    m = _RE_DATEDIF_ROUNDED.match(formula_raw)
    if m:
        a = _read_cell_value(m.group(1), formula_sheet, graph)
        b = _read_cell_value(m.group(2), formula_sheet, graph)
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            result = _datedif_calc(float(a), float(b), m.group(3))
            if result is not None:
                try:
                    val = result / float(m.group(4))
                    if m.group(5):  # optional multiplier (e.g. *12)
                        val *= float(m.group(5))
                    return round(val, int(m.group(6)))
                except ZeroDivisionError:
                    pass
        return _MISS

    # ── DATEDIF(ref, ref, "unit"): bare DATEDIF ──
    m = _RE_DATEDIF.match(formula_raw)
    if m:
        a = _read_cell_value(m.group(1), formula_sheet, graph)
        b = _read_cell_value(m.group(2), formula_sheet, graph)
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            result = _datedif_calc(float(a), float(b), m.group(3))
            if result is not None:
                return result
        return _MISS

    # ── IF simple: =IF(A1>0, B1, 0) ──
    m = _RE_IF_SIMPLE.match(formula_raw)
    if m:
        left_ref = m.group(1)
        op = m.group(2)
        right_str = m.group(3)
        true_val = m.group(4)
        false_val = m.group(5)

        left = _read_cell_value(left_ref, formula_sheet, graph)
        if not isinstance(left, (int, float)):
            return _MISS
        left = float(left)

        # Right side: literal number or cell ref
        try:
            right = float(right_str)
        except ValueError:
            right_cell = _read_cell_value(right_str, formula_sheet, graph)
            if not isinstance(right_cell, (int, float)):
                return _MISS
            right = float(right_cell)

        # Comparison
        _OPS = {'>': lambda a, b: a > b, '<': lambda a, b: a < b,
                '>=': lambda a, b: a >= b, '<=': lambda a, b: a <= b,
                '=': lambda a, b: a == b, '<>': lambda a, b: a != b,
                '!=': lambda a, b: a != b}
        cmp_fn = _OPS.get(op)
        if cmp_fn is None:
            return _MISS
        cond = cmp_fn(left, right)

        chosen = true_val if cond else false_val
        try:
            return float(chosen)
        except ValueError:
            return _read_cell_value(chosen, formula_sheet, graph)
        except ValueError:
            return _read_cell_value(chosen, formula_sheet, graph)

    return _MISS


# ══════════════════════════════════════════════════════════════════════════════
# Formula compilation cache
# ══════════════════════════════════════════════════════════════════════════════

_compiled_cache: dict[str, Any] = {}

# ── Performance stats (for profiling, disabled by default) ────────────────────
_perf_stats: dict[str, float] = {
    "build_plan": 0.0,
    "build_input": 0.0,
    "eval_func": 0.0,
    "fast_path": 0.0,
    "count": 0,
    "fast_hits": 0,
}
_perf_enabled: bool = False


def enable_perf_stats(enabled: bool = True) -> None:
    global _perf_enabled
    _perf_enabled = enabled
    if enabled:
        for k in _perf_stats:
            _perf_stats[k] = 0.0


def get_perf_stats() -> dict[str, float]:
    return dict(_perf_stats)


def clear_formula_cache() -> None:
    _compiled_cache.clear()
    _input_plans.clear()


def _compile_formula(formula: str):
    if formula in _compiled_cache:
        return _compiled_cache[formula]

    if not formula.startswith("="):
        formula = "=" + formula

    try:
        ast_result = _PARSER.ast(formula)
        func = ast_result[1].compile()
        _compiled_cache[formula] = func
        return func
    except Exception:
        _compiled_cache[formula] = None
        return None


# ══════════════════════════════════════════════════════════════════════════════
# Main entry point
# ══════════════════════════════════════════════════════════════════════════════

def evaluate_cell(cell_id: str, graph: FinancialGraph) -> Optional[Any]:
    if not _FORMULAS_AVAILABLE:
        return None

    cell = graph.cells.get(cell_id)
    if cell is None or not cell.formula_raw:
        return None

    if _perf_enabled:
        _perf_stats["count"] += 1

    # ── Build or fetch cached input plan ──
    func = _compile_formula(cell.formula_raw)
    if func is None:
        return None

    if _perf_enabled:
        t0 = time.perf_counter()

    plan = _build_input_plan(cell_id, func.inputs, cell.sheet, graph)

    if _perf_enabled:
        _perf_stats["build_plan"] += time.perf_counter() - t0

    # ── Try fast path for simple patterns ──
    if _perf_enabled:
        t0 = time.perf_counter()

    fast_result = _try_fast_eval(cell.formula_raw, cell.sheet, graph, plan)

    if _perf_enabled:
        _perf_stats["fast_path"] += time.perf_counter() - t0

    if fast_result is not _MISS:
        if _perf_enabled:
            _perf_stats["fast_hits"] += 1
        return fast_result

    # ── Full evaluation via formulas library ──
    try:
        if _perf_enabled:
            t0 = time.perf_counter()

        kwargs = _build_input_map_fast(plan, graph)

        if _perf_enabled:
            _perf_stats["build_input"] += time.perf_counter() - t0
            t0 = time.perf_counter()

        result = func(**kwargs)

        if _perf_enabled:
            _perf_stats["eval_func"] += time.perf_counter() - t0
    except Exception:
        return None

    return _extract_scalar(result)


def _extract_scalar(result: Any) -> Any:
    if isinstance(result, np.ndarray):
        flat = result.flatten()
        if flat.size == 0:
            return None
        val = flat[0]
        if isinstance(val, float) and np.isnan(val):
            return None
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            return float(val)
        if isinstance(val, (np.bool_,)):
            return bool(val)
        return val
    return result
