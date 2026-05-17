"""Batch Q&A processing engine."""
from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Any

import openpyxl

from financial_kg.models.graph import FinancialGraph
from financial_kg.engine.derived_metrics import compute_derived_metrics, serialize_metrics


@dataclass
class BatchResult:
    question: str
    answer_text: str
    confidence: int
    metrics: list[dict]
    sources: list[dict]
    error: str | None = None


def read_questions_from_excel(filepath: str) -> list[str]:
    """Read questions from Excel file. Column A = question text."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet_name = "questions" if "questions" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    questions = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).strip():
            questions.append(str(row[0]).strip())

    wb.close()
    return questions


def process_batch_questions(
    graph: FinancialGraph,
    questions: list[str],
    derived_fn,
    whatif_fn,
    llm_fn,
) -> list[BatchResult]:
    """Process all questions with 3-tier fallback: what-if -> derived -> LLM.

    Args:
        graph: FinancialGraph for computation.
        questions: List of question strings.
        derived_fn: Callable(question, dm_data, graph) -> dict|None
        whatif_fn: Callable(graph, question) -> WhatIfAnswer|None
        llm_fn: Callable(question) -> dict (text, confidence, metrics, sources)
    """
    dm_data = serialize_metrics(compute_derived_metrics(graph))
    results: list[BatchResult] = []

    for q in questions:
        try:
            # Tier 1: what-if
            whatif_answer = whatif_fn(graph, q)
            if whatif_answer:
                results.append(_whatif_to_batch(whatif_answer, q))
                continue

            # Tier 2: derived metrics
            derived = derived_fn(q, dm_data, graph)
            if derived:
                results.append(BatchResult(
                    question=q,
                    answer_text=derived.get("text", ""),
                    confidence=derived.get("confidence", 0),
                    metrics=derived.get("metrics", []),
                    sources=derived.get("sources", []),
                ))
                continue

            # Tier 3: LLM
            llm_result = llm_fn(q)
            results.append(BatchResult(
                question=q,
                answer_text=llm_result.get("text", ""),
                confidence=llm_result.get("confidence", 0),
                metrics=llm_result.get("metrics", []),
                sources=llm_result.get("sources", []),
            ))

        except Exception as e:
            results.append(BatchResult(
                question=q, answer_text="", confidence=0,
                metrics=[], sources=[], error=str(e),
            ))

    return results


def _whatif_to_batch(answer, question: str) -> BatchResult:
    """Convert WhatIfAnswer to BatchResult."""
    metrics = []
    for m in answer.metrics:
        if m.after is None:
            continue
        metrics.append({
            "name": m.name, "before": m.before,
            "after": m.after, "delta": m.delta, "unit": m.unit,
        })

    return BatchResult(
        question=question,
        answer_text=answer.text,
        confidence=answer.confidence,
        metrics=metrics,
        sources=[{
            "name": answer.param_name, "sheet": "参数输入",
            "value": f"{answer.param_before:,.2f} → {answer.param_after:,.2f}",
        }],
    )


def export_batch_results_excel(results: list[BatchResult], output_path: str) -> str:
    """Export batch results to Excel with 3 sheets."""
    wb = openpyxl.Workbook()
    header_fill = openpyxl.styles.PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "问答汇总"
    for col_idx, h in enumerate(["序号", "问题", "回答摘要", "置信度", "错误"], 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, r in enumerate(results, 2):
        ws1.cell(row=row_idx, column=1, value=row_idx - 1)
        ws1.cell(row=row_idx, column=2, value=r.question)
        ws1.cell(row=row_idx, column=3, value=r.answer_text[:300])
        ws1.cell(row=row_idx, column=4, value=r.confidence)
        ws1.cell(row=row_idx, column=5, value=r.error or "")

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 60
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 30

    # Sheet 2: Metrics
    ws2 = wb.create_sheet("指标明细")
    for col_idx, h in enumerate(["序号", "指标名称", "Before", "After", "Delta", "单位"], 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    for i, r in enumerate(results, 1):
        for m in r.metrics:
            ws2.cell(row=row_idx, column=1, value=i)
            ws2.cell(row=row_idx, column=2, value=m.get("name", ""))
            ws2.cell(row=row_idx, column=3, value=m.get("before"))
            ws2.cell(row=row_idx, column=4, value=m.get("after"))
            ws2.cell(row=row_idx, column=5, value=m.get("delta"))
            ws2.cell(row=row_idx, column=6, value=m.get("unit", ""))
            row_idx += 1

    # Sheet 3: Sources
    ws3 = wb.create_sheet("数据来源")
    for col_idx, h in enumerate(["序号", "来源名称", "Sheet", "值"], 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    for i, r in enumerate(results, 1):
        for s in r.sources:
            ws3.cell(row=row_idx, column=1, value=i)
            ws3.cell(row=row_idx, column=2, value=s.get("name", ""))
            ws3.cell(row=row_idx, column=3, value=s.get("sheet", ""))
            ws3.cell(row=row_idx, column=4, value=s.get("value", ""))
            row_idx += 1

    wb.save(output_path)
    return output_path
