"""对比可视化组件: 汇总分析、热力图、指标聚合图、差异报告导出。"""
from __future__ import annotations

import json
import os
from typing import Any

from financial_kg.models.graph import FinancialGraph
from financial_kg.engine.snapshot import SnapshotDiff


# ── 汇总分析 ─────────────────────────────────────────────────────────────────

def compute_change_summary(
    diff: SnapshotDiff,
    graph: FinancialGraph,
) -> dict:
    """计算变更汇总统计。"""
    cells = diff.changed_cells
    if not cells:
        return {
            "total_increase": 0,
            "total_decrease": 0,
            "max_magnitude": 0,
            "max_magnitude_cell": "",
            "top_indicators": [],
            "sheets_ranking": [],
            "critical_count": 0,
            "normal_count": 0,
        }

    increases = [c for c in cells if c.get("direction") == "increase"]
    decreases = [c for c in cells if c.get("direction") == "decrease"]
    total_increase = sum(c.get("change_magnitude", 0) for c in increases if isinstance(c.get("change_magnitude"), (int, float)))
    total_decrease = sum(c.get("change_magnitude", 0) for c in decreases if isinstance(c.get("change_magnitude"), (int, float)))

    max_cell = max(cells, key=lambda c: c.get("change_magnitude", 0))

    # Sheet 变更排行
    sheet_counts: dict[str, int] = {}
    for c in cells:
        sheet = c.get("sheet", "")
        sheet_counts[sheet] = sheet_counts.get(sheet, 0) + 1
    sheets_ranking = sorted(sheet_counts.items(), key=lambda x: x[1], reverse=True)

    # Top 10 indicator
    ind_counts: dict[str, int] = {}
    for c in cells:
        name = c.get("indicator_name", "") or ""
        if name:
            ind_counts[name] = ind_counts.get(name, 0) + 1
    top_indicators = sorted(ind_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    # 影响评级
    critical = sum(1 for c in cells if c.get("indicator_name"))
    normal = len(cells) - critical

    return {
        "total_increase": total_increase,
        "total_decrease": total_decrease,
        "max_magnitude": max_cell.get("change_magnitude", 0),
        "max_magnitude_cell": max_cell.get("id", ""),
        "top_indicators": top_indicators,
        "sheets_ranking": sheets_ranking,
        "critical_count": critical,
        "normal_count": normal,
    }


# ── 指标聚合图 ───────────────────────────────────────────────────────────────

def build_indicator_change_chart(
    diff: SnapshotDiff,
    graph: FinancialGraph,
) -> dict:
    """构建指标级变化聚合图数据。"""
    ind_data: dict[str, dict] = {}
    for c in diff.changed_cells:
        name = c.get("indicator_name", "") or ""
        if not name:
            continue
        if name not in ind_data:
            ind_data[name] = {"name": name, "cell_count": 0, "total_magnitude": 0, "increases": 0, "decreases": 0}
        ind_data[name]["cell_count"] += 1
        ind_data[name]["total_magnitude"] += c.get("change_magnitude", 0)
        if c.get("direction") == "increase":
            ind_data[name]["increases"] += 1
        elif c.get("direction") == "decrease":
            ind_data[name]["decreases"] += 1

    items = sorted(ind_data.values(), key=lambda x: x["cell_count"], reverse=True)
    return {"items": items}


def render_indicator_chart_html(
    chart_data: dict,
    echarts_cdn: str = "https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js",
) -> str:
    """渲染 ECharts 指标变化条形图。"""
    items = chart_data.get("items", [])
    if not items:
        return "<p>无关联 indicator 的变化单元格</p>"

    names = [i["name"][:30] for i in items]
    counts = [i["cell_count"] for i in items]
    magnitudes = [round(i["total_magnitude"], 2) for i in items]

    items_json = json.dumps(items, ensure_ascii=False, default=str)
    names_json = json.dumps(names, ensure_ascii=False)
    counts_json = json.dumps(counts)
    magnitudes_json = json.dumps(magnitudes)

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<script src="{echarts_cdn}"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #1a1a2e; color: #e0e0e0; padding: 12px; }}
.chart-container {{ width: 100%; height: 500px; }}
</style>
</head>
<body>
<div id="chart" class="chart-container"></div>
<script>
const chart = echarts.init(document.getElementById('chart'), null, {{ renderer: 'canvas' }});
const names = {names_json};
const counts = {counts_json};
const magnitudes = {magnitudes_json};
const items = {items_json};

chart.setOption({{
  tooltip: {{
    trigger: 'axis',
    axisPointer: {{ type: 'shadow' }},
    formatter: function(params) {{
      const idx = params[0].dataIndex;
      const item = items[idx];
      return `<b>${{item.name}}</b><br/>` +
        `变化单元格: ${{item.cell_count}}<br/>` +
        `总变化量: ${{item.total_magnitude.toFixed(2)}}<br/>` +
        `↑ 增加: ${{item.increases}} / ↓ 减少: ${{item.decreases}}`;
    }}
  }},
  grid: {{ left: 150, right: 60, top: 20, bottom: 30 }},
  xAxis: {{ type: 'value', name: '变化单元格数', axisLine: {{ lineStyle: {{ color: '#555' }} }}, splitLine: {{ lineStyle: {{ color: '#333' }} }} }},
  yAxis: {{ type: 'category', data: names, inverse: true, axisLabel: {{ color: '#ccc', fontSize: 11 }}, axisLine: {{ lineStyle: {{ color: '#555' }} }} }},
  series: [{{
    type: 'bar',
    data: counts.map((v, i) => ({{
      value: v,
      itemStyle: {{
        color: magnitudes[i] >= 0 ?
          new echarts.graphic.LinearGradient(0, 0, 1, 0, [{{ offset: 0, color: '#66bb6a' }}, {{ offset: 1, color: '#2e7d32' }}]) :
          new echarts.graphic.LinearGradient(0, 0, 1, 0, [{{ offset: 0, color: '#ef5350' }}, {{ offset: 1, color: '#c62828' }}])
      }}
    }})),
    barWidth: 20,
    label: {{ show: true, position: 'right', color: '#ccc' }}
  }}]
}});
window.addEventListener('resize', () => chart.resize());
</script>
</body>
</html>"""


# ── 单元格热力图 ─────────────────────────────────────────────────────────────

def build_heatmap_data(
    graph: FinancialGraph,
    diff: SnapshotDiff,
    sheet_name: str | None = None,
) -> list[dict]:
    """构建单元格级热力图数据。

    返回 [{row, col_num, magnitude, old, new, direction, indicator_name, cell_id, sheet}, ...]
    按 sheet 过滤, 如果 sheet_name 为 None 则返回全部。
    """
    result = []
    for c in diff.changed_cells:
        sheet = c.get("sheet", "")
        if sheet_name and sheet != sheet_name:
            continue
        cell = graph.cells.get(c["id"])
        col_num = _col_to_num(cell.col) if cell and cell.col else 0
        result.append({
            "row": cell.row if cell else 0,
            "col_num": col_num,
            "col_letter": cell.col if cell else "",
            "magnitude": c.get("change_magnitude", 0),
            "old": c.get("old"),
            "new": c.get("new"),
            "direction": c.get("direction", "unchanged"),
            "indicator_name": c.get("indicator_name", ""),
            "cell_id": c["id"],
            "sheet": sheet,
        })
    return result


def _col_to_num(col: str) -> int:
    """Excel 列字母转数字 (A=1, B=2, ..., AA=27)。"""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def render_heatmap_html(
    heatmap_data: list[dict],
    sheet_name: str = "",
    echarts_cdn: str = "https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js",
) -> str:
    """渲染 ECharts 单元格热力图。"""
    if not heatmap_data:
        return f"<p>Sheet「{sheet_name}」无变化单元格</p>"

    rows = sorted({d["row"] for d in heatmap_data})
    cols = sorted({d["col_num"] for d in heatmap_data})
    min_row, max_row = rows[0], rows[-1]
    min_col, max_col = cols[0], cols[-1]

    # 构建 (col_num, row, magnitude) 数据点
    data_points = []
    for d in heatmap_data:
        data_points.append([d["col_num"], d["row"], d.get("magnitude", 0)])

    max_mag = max(abs(d[2]) for d in data_points) if data_points else 1

    data_json = json.dumps(data_points)
    items_json = json.dumps(heatmap_data, ensure_ascii=False, default=str)

    title = f"Sheet: {sheet_name} — {len(heatmap_data)} 个变化单元格"

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<script src="{echarts_cdn}"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #1a1a2e; color: #e0e0e0; padding: 12px; }}
.chart-container {{ width: 100%; height: 600px; }}
</style>
</head>
<body>
<div id="chart" class="chart-container"></div>
<script>
const chart = echarts.init(document.getElementById('chart'), null, {{ renderer: 'canvas' }});
const data = {data_json};
const items = {items_json};
const maxMag = {max_mag} || 1;

chart.setOption({{
  title: {{ text: '{title}', left: 'center', textStyle: {{ color: '#e0e0e0', fontSize: 14 }} }},
  tooltip: {{
    formatter: function(params) {{
      if (!params.data) return '';
      const item = items[params.dataIndex];
      if (!item) return '';
      return `<b>${{item.cell_id}}</b><br/>` +
        `Sheet: ${{item.sheet}}<br/>` +
        `旧值: ${{item.old}}<br/>` +
        `新值: ${{item.new}}<br/>` +
        `变化量: ${{item.magnitude.toFixed(2)}}<br/>` +
        `方向: ${{item.direction === 'increase' ? '↑ 增加' : '↓ 减少'}}<br/>` +
        `Indicator: ${{item.indicator_name || '—'}}`;
    }}
  }},
  grid: {{ left: 60, right: 80, top: 50, bottom: 40 }},
  xAxis: {{
    type: 'value', min: {min_col - 1}, max: {max_col + 1},
    axisLabel: {{ color: '#888', formatter: v => v > 0 ? String.fromCharCode(64 + v) : '' }},
    splitLine: {{ show: false }},
    axisLine: {{ lineStyle: {{ color: '#444' }} }}
  }},
  yAxis: {{
    type: 'value', min: {min_row - 1}, max: {max_row + 1},
    inverse: true,
    axisLabel: {{ color: '#888' }},
    splitLine: {{ show: false }},
    axisLine: {{ lineStyle: {{ color: '#444' }} }}
  }},
  visualMap: {{
    min: -maxMag, max: maxMag,
    orient: 'vertical', right: 10, top: 'center',
    calculable: true,
    inRange: {{ color: ['#c62828', '#ffcdd2', '#f5f5f5', '#c8e6c9', '#2e7d32'] }},
    textStyle: {{ color: '#aaa' }},
    seriesIndex: [0]
  }},
  series: [{{
    type: 'heatmap',
    data: data,
    emphasis: {{
      itemStyle: {{
        shadowBlur: 10,
        shadowColor: 'rgba(0,0,0,0.5)',
        borderColor: '#fff',
        borderWidth: 2
      }}
    }}
  }}]
}});
window.addEventListener('resize', () => chart.resize());
</script>
</body>
</html>"""


# ── 差异报告 Excel 导出 ──────────────────────────────────────────────────────

def export_diff_report_excel(
    diff: SnapshotDiff,
    graph: FinancialGraph,
    output_path: str,
) -> str:
    """导出差异报告 Excel。

    Sheet1: 汇总指标
    Sheet2: 变化明细 (带颜色标记)
    Sheet3: 受影响 Indicator
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Sheet 1: 汇总
    ws1 = wb.active
    ws1.title = "汇总"
    header_font = Font(bold=True, size=12)
    ws1["A1"] = "快照对比报告"
    ws1["A1"].font = Font(bold=True, size=16)
    ws1["A2"] = f"基准: {diff.snapshot_a}"
    ws1["A3"] = f"对比: {diff.snapshot_b}"
    ws1["A5"] = "指标"; ws1["B5"] = "数值"; ws1["A5"].font = header_font; ws1["B5"].font = header_font
    ws1["A6"] = "变化单元格数"; ws1["B6"] = diff.summary.get("total_changed_cells", 0)
    ws1["A7"] = "受影响 Indicator 数"; ws1["B7"] = diff.summary.get("total_changed_indicators", 0)
    ws1["A8"] = "涉及 Sheet"; ws1["B8"] = ", ".join(diff.summary.get("sheets_affected", []))
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40

    # Sheet 2: 变化明细
    ws2 = wb.create_sheet("变化明细")
    headers = ["Cell ID", "Sheet", "旧值", "新值", "变化量", "方向", "公式", "Indicator"]
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, c in enumerate(diff.changed_cells, 2):
        direction = c.get("direction", "unchanged")
        fill = green_fill if direction == "increase" else red_fill
        ws2.cell(row=row_idx, column=1, value=c["id"])
        ws2.cell(row=row_idx, column=2, value=c.get("sheet", ""))
        ws2.cell(row=row_idx, column=3, value=c.get("old"))
        ws2.cell(row=row_idx, column=4, value=c.get("new"))
        ws2.cell(row=row_idx, column=5, value=c.get("change_magnitude", 0))
        dir_cell = ws2.cell(row=row_idx, column=6, value="↑ 增加" if direction == "increase" else "↓ 减少")
        dir_cell.fill = fill
        ws2.cell(row=row_idx, column=7, value=c.get("formula") or "")
        ws2.cell(row=row_idx, column=8, value=c.get("indicator_name") or "")
        for col_idx in range(1, 9):
            ws2.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx in range(1, 9):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    # Sheet 3: 受影响 Indicator
    ws3 = wb.create_sheet("受影响 Indicator")
    ind_headers = ["Indicator", "Sheet", "旧汇总值", "新汇总值", "变化单元格数"]
    for col_idx, h in enumerate(ind_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, ind in enumerate(diff.affected_indicators, 2):
        ws3.cell(row=row_idx, column=1, value=ind["name"])
        ws3.cell(row=row_idx, column=2, value=ind.get("sheet", ""))
        ws3.cell(row=row_idx, column=3, value=ind.get("old_summary"))
        ws3.cell(row=row_idx, column=4, value=ind.get("new_summary"))
        ws3.cell(row=row_idx, column=5, value=ind.get("changed_cell_count", 0))

    for col_idx in range(1, 6):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30

    wb.save(output_path)
    return output_path
