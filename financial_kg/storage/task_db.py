"""SQLite-backed task and snapshot registry."""
from __future__ import annotations
import json
import sqlite3
import os
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from typing import Optional

_DEFAULT_DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "tasks.db")


@dataclass
class TaskRecord:
    id: str
    filename: str
    status: str          # pending | running | done | error
    created_at: str
    cell_count: int = 0
    indicator_count: int = 0
    table_count: int = 0
    output_dir: str = ""
    error_msg: str = ""


@dataclass
class SnapshotRecord:
    id: str
    task_id: str
    name: str
    description: str
    created_at: str
    filepath: str


class TaskDB:
    def __init__(self, db_path: str = _DEFAULT_DB):
        self.db_path = os.path.abspath(db_path)
        self._init()

    @contextmanager
    def _conn(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()

    def _init(self):
        with self._conn() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS tasks (
                    id TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'pending',
                    created_at TEXT NOT NULL,
                    cell_count INTEGER DEFAULT 0,
                    indicator_count INTEGER DEFAULT 0,
                    table_count INTEGER DEFAULT 0,
                    output_dir TEXT DEFAULT '',
                    error_msg TEXT DEFAULT ''
                );
                CREATE TABLE IF NOT EXISTS snapshots (
                    id TEXT PRIMARY KEY,
                    task_id TEXT NOT NULL,
                    name TEXT NOT NULL,
                    description TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    filepath TEXT NOT NULL,
                    FOREIGN KEY (task_id) REFERENCES tasks(id)
                );
                CREATE TABLE IF NOT EXISTS qa_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id TEXT NOT NULL,
                    messages TEXT NOT NULL DEFAULT '[]',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY (task_id) REFERENCES tasks(id)
                );
                CREATE INDEX IF NOT EXISTS idx_qa_history_task ON qa_history(task_id);
                CREATE TABLE IF NOT EXISTS sensitivity_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id TEXT NOT NULL,
                    run_name TEXT NOT NULL,
                    params TEXT NOT NULL DEFAULT '[]',
                    perturbations TEXT NOT NULL DEFAULT '[]',
                    base_metrics TEXT NOT NULL DEFAULT '{}',
                    scenarios TEXT NOT NULL DEFAULT '[]',
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (task_id) REFERENCES tasks(id)
                );
                CREATE INDEX IF NOT EXISTS idx_sensitivity_task ON sensitivity_history(task_id);
                CREATE TABLE IF NOT EXISTS scenario_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id TEXT NOT NULL,
                    run_name TEXT NOT NULL,
                    params TEXT NOT NULL DEFAULT '[]',
                    base_metrics TEXT NOT NULL DEFAULT '{}',
                    scenarios TEXT NOT NULL DEFAULT '[]',
                    comparison_table TEXT NOT NULL DEFAULT '[]',
                    delta_table TEXT NOT NULL DEFAULT '[]',
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (task_id) REFERENCES tasks(id)
                );
                CREATE INDEX IF NOT EXISTS idx_scenario_task ON scenario_history(task_id);
                CREATE TABLE IF NOT EXISTS monte_carlo_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id TEXT NOT NULL,
                    run_name TEXT NOT NULL,
                    mode TEXT NOT NULL DEFAULT 'fast',
                    iterations INTEGER NOT NULL,
                    params TEXT NOT NULL DEFAULT '[]',
                    base_irr REAL DEFAULT 0,
                    statistics TEXT NOT NULL DEFAULT '{}',
                    probability_table TEXT NOT NULL DEFAULT '[]',
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (task_id) REFERENCES tasks(id)
                );
                CREATE INDEX IF NOT EXISTS idx_monte_carlo_task ON monte_carlo_history(task_id);
                CREATE TABLE IF NOT EXISTS qa_answers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id TEXT NOT NULL,
                    question TEXT NOT NULL,
                    answer_data TEXT NOT NULL DEFAULT '{}',
                    confidence INTEGER DEFAULT 0,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (task_id) REFERENCES tasks(id)
                );
                CREATE INDEX IF NOT EXISTS idx_qa_answers_task ON qa_answers(task_id);
                CREATE INDEX IF NOT EXISTS idx_qa_answers_time ON qa_answers(created_at);
            """)

    # ── Tasks ────────────────────────────────────────────────────────────────

    def create_task(self, task_id: str, filename: str, output_dir: str = "") -> TaskRecord:
        rec = TaskRecord(
            id=task_id,
            filename=filename,
            status="pending",
            created_at=datetime.now().isoformat(),
            output_dir=output_dir,
        )
        with self._conn() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO tasks VALUES (?,?,?,?,?,?,?,?,?)",
                (rec.id, rec.filename, rec.status, rec.created_at,
                 rec.cell_count, rec.indicator_count, rec.table_count,
                 rec.output_dir, rec.error_msg),
            )
        return rec

    def update_task(self, task_id: str, **kwargs) -> None:
        allowed = {"status", "cell_count", "indicator_count", "table_count",
                   "output_dir", "error_msg"}
        fields = {k: v for k, v in kwargs.items() if k in allowed}
        if not fields:
            return
        set_clause = ", ".join(f"{k}=?" for k in fields)
        with self._conn() as conn:
            conn.execute(
                f"UPDATE tasks SET {set_clause} WHERE id=?",
                (*fields.values(), task_id),
            )

    def get_task(self, task_id: str) -> Optional[TaskRecord]:
        with self._conn() as conn:
            row = conn.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
        return _row_to_task(row) if row else None

    def list_tasks(self) -> list[TaskRecord]:
        with self._conn() as conn:
            rows = conn.execute("SELECT * FROM tasks ORDER BY created_at DESC").fetchall()
        return [_row_to_task(r) for r in rows]

    # ── Snapshots ────────────────────────────────────────────────────────────

    def save_snapshot(
        self,
        snap_id: str,
        task_id: str,
        name: str,
        filepath: str,
        description: str = "",
    ) -> SnapshotRecord:
        rec = SnapshotRecord(
            id=snap_id,
            task_id=task_id,
            name=name,
            description=description,
            created_at=datetime.now().isoformat(),
            filepath=filepath,
        )
        with self._conn() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO snapshots VALUES (?,?,?,?,?,?)",
                (rec.id, rec.task_id, rec.name, rec.description,
                 rec.created_at, rec.filepath),
            )
        return rec

    def list_snapshots(self, task_id: str) -> list[SnapshotRecord]:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM snapshots WHERE task_id=? ORDER BY created_at DESC",
                (task_id,),
            ).fetchall()
        return [_row_to_snapshot(r) for r in rows]

    def get_snapshot(self, snap_id: str) -> Optional[SnapshotRecord]:
        with self._conn() as conn:
            row = conn.execute("SELECT * FROM snapshots WHERE id=?", (snap_id,)).fetchone()
        return _row_to_snapshot(row) if row else None

    # ── QA History ────────────────────────────────────────────────────────────

    def save_qa_history(self, task_id: str, messages: list[dict]) -> None:
        now = datetime.now().isoformat()
        with self._conn() as conn:
            existing = conn.execute(
                "SELECT id FROM qa_history WHERE task_id=?", (task_id,)
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE qa_history SET messages=?, updated_at=? WHERE task_id=?",
                    (json.dumps(messages, ensure_ascii=False), now, task_id),
                )
            else:
                conn.execute(
                    "INSERT INTO qa_history (task_id, messages, created_at, updated_at) VALUES (?,?,?,?)",
                    (task_id, json.dumps(messages, ensure_ascii=False), now, now),
                )

    def load_qa_history(self, task_id: str) -> list[dict]:
        with self._conn() as conn:
            row = conn.execute(
                "SELECT messages FROM qa_history WHERE task_id=?", (task_id,)
            ).fetchone()
        if row:
            try:
                return json.loads(row["messages"])
            except (json.JSONDecodeError, TypeError):
                return []
        return []

    def clear_qa_history(self, task_id: str) -> None:
        with self._conn() as conn:
            conn.execute("DELETE FROM qa_history WHERE task_id=?", (task_id,))

    # ── Sensitivity History ──────────────────────────────────────────────────

    def save_sensitivity(
        self,
        task_id: str,
        run_name: str,
        params: list[dict],
        perturbations: list[float],
        base_metrics: dict,
        scenarios: list[dict],
    ) -> int:
        now = datetime.now().isoformat()
        with self._conn() as conn:
            cur = conn.execute(
                "INSERT INTO sensitivity_history (task_id, run_name, params, perturbations, base_metrics, scenarios, created_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (task_id, run_name, json.dumps(params, ensure_ascii=False),
                 json.dumps(perturbations), json.dumps(base_metrics, ensure_ascii=False),
                 json.dumps(scenarios, ensure_ascii=False), now),
            )
            return cur.lastrowid

    def list_sensitivity(self, task_id: str) -> list[dict]:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT id, run_name, params, perturbations, base_metrics, scenarios, created_at "
                "FROM sensitivity_history WHERE task_id=? ORDER BY created_at DESC",
                (task_id,),
            ).fetchall()
        return [
            {
                "id": r["id"],
                "run_name": r["run_name"],
                "params": json.loads(r["params"]),
                "perturbations": json.loads(r["perturbations"]),
                "base_metrics": json.loads(r["base_metrics"]),
                "scenarios": json.loads(r["scenarios"]),
                "created_at": r["created_at"],
            }
            for r in rows
        ]

    def load_sensitivity(self, record_id: int) -> Optional[dict]:
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM sensitivity_history WHERE id=?", (record_id,)
            ).fetchone()
        if not row:
            return None
        return {
            "id": row["id"],
            "task_id": row["task_id"],
            "run_name": row["run_name"],
            "params": json.loads(row["params"]),
            "perturbations": json.loads(row["perturbations"]),
            "base_metrics": json.loads(row["base_metrics"]),
            "scenarios": json.loads(row["scenarios"]),
            "created_at": row["created_at"],
        }

    def delete_sensitivity(self, record_id: int) -> None:
        with self._conn() as conn:
            conn.execute("DELETE FROM sensitivity_history WHERE id=?", (record_id,))

    # ── QA Answers (granular, for comparison) ────────────────────────────────

    def save_qa_answer(
        self,
        task_id: str,
        question: str,
        answer_data: dict,
        confidence: int = 0,
    ) -> int:
        """Save a single QA answer, return answer_id."""
        now = datetime.now().isoformat()
        with self._conn() as conn:
            cur = conn.execute(
                "INSERT INTO qa_answers (task_id, question, answer_data, confidence, created_at) VALUES (?,?,?,?,?)",
                (task_id, question, json.dumps(answer_data, ensure_ascii=False), confidence, now),
            )
            return cur.lastrowid

    def list_qa_answers(self, task_id: str, limit: int = 50) -> list[dict]:
        """List QA answers for a task, newest first."""
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT id, question, answer_data, confidence, created_at FROM qa_answers WHERE task_id=? ORDER BY created_at DESC LIMIT ?",
                (task_id, limit),
            ).fetchall()
        return [
            {
                "id": r["id"],
                "question": r["question"],
                "answer_data": json.loads(r["answer_data"]),
                "confidence": r["confidence"],
                "created_at": r["created_at"],
            }
            for r in rows
        ]

    def get_qa_answer(self, answer_id: int) -> Optional[dict]:
        """Get a single QA answer by ID."""
        with self._conn() as conn:
            row = conn.execute(
                "SELECT id, task_id, question, answer_data, confidence, created_at FROM qa_answers WHERE id=?",
                (answer_id,),
            ).fetchone()
        if not row:
            return None
        return {
            "id": row["id"],
            "task_id": row["task_id"],
            "question": row["question"],
            "answer_data": json.loads(row["answer_data"]),
            "confidence": row["confidence"],
            "created_at": row["created_at"],
        }

    def delete_qa_answer(self, answer_id: int) -> None:
        """Delete a single QA answer."""
        with self._conn() as conn:
            conn.execute("DELETE FROM qa_answers WHERE id=?", (answer_id,))

    def clear_qa_answers(self, task_id: str) -> None:
        """Delete all QA answers for a task."""
        with self._conn() as conn:
            conn.execute("DELETE FROM qa_answers WHERE task_id=?", (task_id,))

    def export_qa_history_excel(
        self,
        task_id: str,
        output_path: str,
    ) -> str:
        """Export QA history to Excel with 3 sheets: Summary, Metrics, Sources."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        answers = self.list_qa_answers(task_id, limit=200)

        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "问答汇总"
        headers = ["ID", "问题", "回答摘要", "置信度", "时间"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, a in enumerate(answers, 2):
            answer_text = a["answer_data"].get("text", "")[:200] if a["answer_data"] else ""
            ws1.cell(row=row_idx, column=1, value=a["id"])
            ws1.cell(row=row_idx, column=2, value=a["question"])
            ws1.cell(row=row_idx, column=3, value=answer_text)
            ws1.cell(row=row_idx, column=4, value=a["confidence"])
            ws1.cell(row=row_idx, column=5, value=a["created_at"][:19])

        ws1.column_dimensions["A"].width = 8
        ws1.column_dimensions["B"].width = 40
        ws1.column_dimensions["C"].width = 60
        ws1.column_dimensions["D"].width = 10
        ws1.column_dimensions["E"].width = 20

        # Sheet 2: Metrics
        ws2 = wb.create_sheet("指标明细")
        m_headers = ["问答ID", "指标名称", "值", "单位", "匹配原因"]
        for col_idx, h in enumerate(m_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font

        row_idx = 2
        for a in answers:
            for m in a["answer_data"].get("metrics", []):
                ws2.cell(row=row_idx, column=1, value=a["id"])
                ws2.cell(row=row_idx, column=2, value=m.get("name", ""))
                ws2.cell(row=row_idx, column=3, value=m.get("value", ""))
                ws2.cell(row=row_idx, column=4, value=m.get("unit", ""))
                ws2.cell(row=row_idx, column=5, value=m.get("match_reason", ""))
                row_idx += 1

        # Sheet 3: Sources
        ws3 = wb.create_sheet("数据来源")
        s_headers = ["问答ID", "来源名称", "Sheet", "值", "单位", "评分"]
        for col_idx, h in enumerate(s_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font

        row_idx = 2
        for a in answers:
            for s in a["answer_data"].get("sources", []):
                ws3.cell(row=row_idx, column=1, value=a["id"])
                ws3.cell(row=row_idx, column=2, value=s.get("name", ""))
                ws3.cell(row=row_idx, column=3, value=s.get("sheet", ""))
                ws3.cell(row=row_idx, column=4, value=s.get("value", ""))
                ws3.cell(row=row_idx, column=5, value=s.get("unit", ""))
                ws3.cell(row=row_idx, column=6, value=s.get("score", 0))
                row_idx += 1

        wb.save(output_path)
        return output_path

    # ── Scenario History ────────────────────────────────────────────────────────

    def save_scenario(
        self,
        task_id: str,
        run_name: str,
        params: list[dict],
        base_metrics: dict,
        scenarios: list[dict],
        comparison_table: list[dict],
        delta_table: list[dict],
    ) -> int:
        """Save scenario analysis result, return record_id."""
        now = datetime.now().isoformat()
        with self._conn() as conn:
            cur = conn.execute(
                "INSERT INTO scenario_history (task_id, run_name, params, base_metrics, scenarios, comparison_table, delta_table, created_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (task_id, run_name, json.dumps(params, ensure_ascii=False),
                 json.dumps(base_metrics, ensure_ascii=False),
                 json.dumps(scenarios, ensure_ascii=False),
                 json.dumps(comparison_table, ensure_ascii=False),
                 json.dumps(delta_table, ensure_ascii=False), now),
            )
            return cur.lastrowid

    def list_scenario(self, task_id: str) -> list[dict]:
        """List scenario analysis history for a task."""
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT id, run_name, params, base_metrics, scenarios, comparison_table, delta_table, created_at "
                "FROM scenario_history WHERE task_id=? ORDER BY created_at DESC",
                (task_id,),
            ).fetchall()
        return [
            {
                "id": r["id"],
                "run_name": r["run_name"],
                "params": json.loads(r["params"]),
                "base_metrics": json.loads(r["base_metrics"]),
                "scenarios": json.loads(r["scenarios"]),
                "comparison_table": json.loads(r["comparison_table"]),
                "delta_table": json.loads(r["delta_table"]),
                "created_at": r["created_at"],
            }
            for r in rows
        ]

    def load_scenario(self, record_id: int) -> Optional[dict]:
        """Load a single scenario analysis record."""
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM scenario_history WHERE id=?", (record_id,)
            ).fetchone()
        if not row:
            return None
        return {
            "id": row["id"],
            "task_id": row["task_id"],
            "run_name": row["run_name"],
            "params": json.loads(row["params"]),
            "base_metrics": json.loads(row["base_metrics"]),
            "scenarios": json.loads(row["scenarios"]),
            "comparison_table": json.loads(row["comparison_table"]),
            "delta_table": json.loads(row["delta_table"]),
            "created_at": row["created_at"],
        }

    def delete_scenario(self, record_id: int) -> None:
        """Delete a scenario analysis record."""
        with self._conn() as conn:
            conn.execute("DELETE FROM scenario_history WHERE id=?", (record_id,))

    # ── Monte Carlo History ────────────────────────────────────────────────────────

    def save_monte_carlo(
        self,
        task_id: str,
        run_name: str,
        mode: str,
        iterations: int,
        params: list[dict],
        base_irr: float,
        statistics: dict,
        probability_table: list[dict],
        irr_values: list[float] | None = None,  # Optional: store raw IRR values
    ) -> int:
        """Save monte carlo result, return record_id."""
        now = datetime.now().isoformat()
        with self._conn() as conn:
            # Add irr_values column if not exists (migration)
            try:
                conn.execute("ALTER TABLE monte_carlo_history ADD COLUMN irr_values TEXT DEFAULT '[]'")
            except Exception:
                pass  # Column already exists

            cur = conn.execute(
                "INSERT INTO monte_carlo_history (task_id, run_name, mode, iterations, params, base_irr, statistics, probability_table, irr_values, created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (task_id, run_name, mode, iterations,
                 json.dumps(params, ensure_ascii=False),
                 base_irr,
                 json.dumps(statistics, ensure_ascii=False),
                 json.dumps(probability_table, ensure_ascii=False),
                 json.dumps(irr_values or [], ensure_ascii=False),  # Store IRR values
                 now),
            )
            return cur.lastrowid

    def list_monte_carlo(self, task_id: str) -> list[dict]:
        """List monte carlo history for a task."""
        with self._conn() as conn:
            # Ensure irr_values column exists
            try:
                conn.execute("ALTER TABLE monte_carlo_history ADD COLUMN irr_values TEXT DEFAULT '[]'")
            except Exception:
                pass

            rows = conn.execute(
                "SELECT id, run_name, mode, iterations, params, base_irr, statistics, probability_table, irr_values, created_at "
                "FROM monte_carlo_history WHERE task_id=? ORDER BY created_at DESC",
                (task_id,),
            ).fetchall()
        return [
            {
                "id": r["id"],
                "run_name": r["run_name"],
                "mode": r["mode"],
                "iterations": r["iterations"],
                "params": json.loads(r["params"]),
                "base_irr": r["base_irr"],
                "statistics": json.loads(r["statistics"]),
                "probability_table": json.loads(r["probability_table"]),
                "irr_values": json.loads(r["irr_values"] if "irr_values" in r.keys() else "[]"),  # Load IRR values
                "created_at": r["created_at"],
            }
            for r in rows
        ]

    def load_monte_carlo(self, record_id: int) -> Optional[dict]:
        """Load a single monte carlo record."""
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM monte_carlo_history WHERE id=?", (record_id,)
            ).fetchone()
        if not row:
            return None
        return {
            "id": row["id"],
            "task_id": row["task_id"],
            "run_name": row["run_name"],
            "mode": row["mode"],
            "iterations": row["iterations"],
            "params": json.loads(row["params"]),
            "base_irr": row["base_irr"],
            "statistics": json.loads(row["statistics"]),
            "probability_table": json.loads(row["probability_table"]),
            "created_at": row["created_at"],
        }

    def delete_monte_carlo(self, record_id: int) -> None:
        """Delete a monte carlo record."""
        with self._conn() as conn:
            conn.execute("DELETE FROM monte_carlo_history WHERE id=?", (record_id,))

    # ── Delete ─────────────────────────────────────────────────────────────────

    def delete_task(self, task_id: str) -> Optional[TaskRecord]:
        """Delete a task, its snapshots, and return the task record for file cleanup.
        Returns None if the task doesn't exist."""
        task = self.get_task(task_id)
        if task is None:
            return None
        with self._conn() as conn:
            conn.execute("DELETE FROM snapshots WHERE task_id=?", (task_id,))
            conn.execute("DELETE FROM tasks WHERE id=?", (task_id,))
        return task

    def list_snapshot_files(self, task_id: str) -> list[str]:
        """Return filepath for all snapshots of a task."""
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT filepath FROM snapshots WHERE task_id=?", (task_id,)
            ).fetchall()
        return [r[0] for r in rows]


def _row_to_task(row) -> TaskRecord:
    return TaskRecord(
        id=row["id"], filename=row["filename"], status=row["status"],
        created_at=row["created_at"], cell_count=row["cell_count"],
        indicator_count=row["indicator_count"], table_count=row["table_count"],
        output_dir=row["output_dir"], error_msg=row["error_msg"],
    )


def _row_to_snapshot(row) -> SnapshotRecord:
    return SnapshotRecord(
        id=row["id"], task_id=row["task_id"], name=row["name"],
        description=row["description"], created_at=row["created_at"],
        filepath=row["filepath"],
    )
