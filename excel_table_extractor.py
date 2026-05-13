#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 多表格智能识别与提取工具

功能：
1. 自动识别 sheet 中的多个独立表格区域
2. 提取每个表格的坐标、范围
3. 获取表格名称、描述、数据、公式、样式
4. 支持合并单元格处理
5. 输出结构化 JSON 数据

依赖：pip install openpyxl
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import json
from typing import Dict, List, Tuple, Any, Optional
from collections import defaultdict


class DateTimeEncoder(json.JSONEncoder):
    """Custom JSON encoder for datetime objects"""
    def default(self, obj):
        import datetime
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return obj.isoformat()
        return super().default(obj)


class ExcelTableExtractor:
    """Excel 多表格提取器"""
    
    def __init__(self, file_path: str):
        """
        初始化提取器
        
        Args:
            file_path: Excel 文件路径
        """
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=False)
    
    def detect_merged_cells(self, ws) -> Dict[str, Tuple[int, int, int, int]]:
        """
        获取所有合并单元格的范围
        
        Args:
            ws: worksheet 对象
            
        Returns:
            合并单元格字典 {address: (start_row, start_col, end_row, end_col)}
        """
        merged_ranges = {}
        for merged_range in ws.merged_cells.ranges:
            coord_str = merged_range.coord
            start_addr, end_addr = coord_str.split(':')
            start_row = int(''.join(filter(str.isdigit, start_addr)))
            start_col = column_index_from_string(''.join(filter(str.isalpha, start_addr)))
            end_row = int(''.join(filter(str.isdigit, end_addr)))
            end_col = column_index_from_string(''.join(filter(str.isalpha, end_addr)))
            merged_ranges[coord_str] = (start_row, start_col, end_row, end_col)
        return merged_ranges
    
    def is_cell_in_merged_range(self, row: int, col: int, merged_ranges: Dict) -> Optional[str]:
        """
        检查单元格是否在合并范围内
        
        Args:
            row: 行号
            col: 列号
            merged_ranges: 合并单元格字典
            
        Returns:
            如果在合并范围内，返回合并区域地址，否则返回 None
        """
        for addr, (start_r, start_c, end_r, end_c) in merged_ranges.items():
            if start_r <= row <= end_r and start_c <= col <= end_c:
                return addr
        return None
    
    def detect_table_boundaries(self, ws, merged_ranges: Dict) -> List[Dict]:
        """
        检测 sheet 中的所有独立表格边界
        """
        tables = []
        visited = set()
        recognized_ranges = []
        
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        
        # 第一步：识别纵向合并单元格作为 L 型表锚点
        vertical_merges = []
        for addr, (mr, mc, er, ec) in merged_ranges.items():
            if (er - mr + 1) >= 3 and (ec - mc + 1) <= 2:
                cell = ws.cell(row=mr, column=mc)
                if cell.value is not None:
                    vertical_merges.append((mr, mc, er, ec, addr))
        
        vertical_merges.sort(key=lambda x: (x[0], x[1]))
        
        # 过滤：如果一个锚点在另一个锚点的行范围内且列附近，跳过
        filtered_merges = []
        for anchor in vertical_merges:
            mr, mc, er, ec, addr = anchor
            is_inside_other = False
            for other in vertical_merges:
                omr, omc, oer, oec, oaddr = other
                if addr == oaddr:
                    continue
                if omr <= mr <= oer and omc < mc <= omc + 3:
                    is_inside_other = True
                    break
            if not is_inside_other:
                filtered_merges.append(anchor)
        
        # 第二步：标记所有 L 型锚点区域为"保留"
        l_anchor_ranges = []
        for anchor in filtered_merges:
            mr, mc, er, ec, addr = anchor
            l_anchor_ranges.append((mr, mc, er, min(ec + 100, max_col)))
        
        # 第三步：识别普通矩形表
        for start_row in range(1, max_row + 1):
            for start_col in range(1, max_col + 1):
                if (start_row, start_col) in visited:
                    continue
                
                cell = ws.cell(row=start_row, column=start_col)
                if cell.value is None:
                    continue
                
                # 检查是否是 L 型锚点
                is_l_anchor = False
                merged_addr = self.is_cell_in_merged_range(start_row, start_col, merged_ranges)
                if merged_addr:
                    merged_info = merged_ranges[merged_addr]
                    mr, mc, er, ec = merged_info
                    if (er - mr + 1) >= 3 and (ec - mc + 1) <= 2:
                        is_l_anchor = True
                    if (start_row, start_col) != (mr, mc):
                        continue
                
                if is_l_anchor:
                    continue
                
                # 检查是否在某个 L 型锚点的行范围内
                is_in_l_range = False
                for anchor in filtered_merges:
                    amr, amc, aer, aec, aaddr = anchor
                    if amr <= start_row <= aer and amc < start_col:
                        is_in_l_range = True
                        break
                if is_in_l_range:
                    continue
                
                table_info = self._expand_rectangular_table(ws, start_row, start_col, visited, merged_ranges, max_row, max_col, recognized_ranges)
                if table_info:
                    tables.append(table_info)
                    recognized_ranges.append((table_info['start_row'], table_info['start_col'], 
                                             table_info['end_row'], table_info['end_col']))
        
        # 第四步：从每个纵向合并锚点扩展 L 型表
        for anchor in filtered_merges:
            mr, mc, er, ec, addr = anchor
            
            if (mr, mc) in visited:
                continue
            
            table_info = self._expand_l_shaped_table(ws, mr, mc, er, ec, visited, merged_ranges, max_row, max_col, recognized_ranges)
            if table_info:
                table_info['title_address'] = addr
                table_info['title_text'] = ws.cell(row=mr, column=mc).value
                
                overlaps = []
                sr, sc, er2, ec2 = table_info['start_row'], table_info['start_col'], table_info['end_row'], table_info['end_col']
                for r_range in recognized_ranges:
                    if sr <= r_range[2] and er2 >= r_range[0] and sc <= r_range[3] and ec2 >= r_range[1]:
                        overlaps.append(r_range)
                table_info['excluded_ranges'] = overlaps
                
                tables.append(table_info)
                recognized_ranges.append((table_info['start_row'], table_info['start_col'], 
                                         table_info['end_row'], table_info['end_col']))
                
                sr, sc, er2, ec2 = table_info['start_row'], table_info['start_col'], table_info['end_row'], table_info['end_col']
                for other_addr, (omr, omc, oer, oec) in merged_ranges.items():
                    if (oer - omr + 1) >= 3 and (oec - omc + 1) <= 2:
                        if sr <= omr <= er2 and sc <= omc <= ec2:
                            visited.add((omr, omc))
        
        # 第五步：合并同一行的表格（处理有空列但实际是一个表的情况）
        tables = self._merge_adjacent_tables(tables, visited, ws, max_row, max_col)
        
        # 按位置排序
        tables.sort(key=lambda t: (t['start_row'], t['start_col']))
        
        return tables
    
    def _merge_adjacent_tables(self, tables: List[Dict], visited: set, ws, max_row: int, max_col: int) -> List[Dict]:
        """
        合并同一区域内被空列分隔的表格
        """
        if not tables:
            return tables
        
        merged = []
        used = set()
        
        for i, t1 in enumerate(tables):
            if i in used:
                continue
            
            current = t1.copy()
            for j, t2 in enumerate(tables):
                if j <= i or j in used:
                    continue
                
                # 检查是否在同一行范围内且列接近（间隔不超过3列）
                col_gap = t2['start_col'] - current['end_col'] - 1
                row_overlap = (t1['start_row'] <= t2['end_row'] and t2['start_row'] <= t1['end_row'])
                
                if row_overlap and 0 < col_gap <= 3:
                    # 检查上一行是否有标题连接它们
                    title_row = min(current['start_row'], t2['start_row']) - 1
                    has_bridge = False
                    
                    # 检查上一行或当前行是否有数据连接
                    for check_row in [title_row, min(current['start_row'], t2['start_row'])]:
                        if 1 <= check_row <= max_row:
                            for c in range(current['start_col'], max(current['end_col'], t2['end_col']) + 1):
                                cell = ws.cell(row=check_row, column=c)
                                if cell.value is not None:
                                    has_bridge = True
                                    break
                        if has_bridge:
                            break
                    
                    if has_bridge:
                        current = {
                            'start_row': min(current['start_row'], t2['start_row']),
                            'end_row': max(current['end_row'], t2['end_row']),
                            'start_col': min(current['start_col'], t2['start_col']),
                            'end_col': max(current['end_col'], t2['end_col']),
                            'cells': current.get('cells', set()) | t2.get('cells', set()),
                            'address': f"{get_column_letter(min(current['start_col'], t2['start_col']))}{min(current['start_row'], t2['start_row'])}:{get_column_letter(max(current['end_col'], t2['end_col']))}{max(current['end_row'], t2['end_row'])}"
                        }
                        used.add(j)
            
            merged.append(current)
            used.add(i)
        
        return merged
    
    def _expand_l_shaped_table(self, ws, anchor_start_row, anchor_start_col, anchor_end_row, anchor_end_col,
                                visited: set, merged_ranges: Dict, max_row: int, max_col: int,
                                recognized_ranges: list) -> Optional[Dict]:
        """
        扩展 L 型表格（左侧有纵向合并单元格）
        """
        table_cells = set()
        
        # 添加锚点区域（左侧合并单元格）
        for r in range(anchor_start_row, anchor_end_row + 1):
            for c in range(anchor_start_col, anchor_end_col + 1):
                table_cells.add((r, c))
                visited.add((r, c))
        
        # 向右扩展：找到最右侧有数据的列
        max_data_col = anchor_end_col
        
        # 检查锚点行本身（包括合并单元格）
        for c in range(anchor_end_col + 1, max_col + 1):
            merged_addr = self.is_cell_in_merged_range(anchor_start_row, c, merged_ranges)
            if merged_addr:
                mr2, mc2, er2, ec2 = merged_ranges[merged_addr]
                cell = ws.cell(row=mr2, column=mc2)
                if cell.value is not None:
                    max_data_col = max(max_data_col, ec2)
            else:
                cell = ws.cell(row=anchor_start_row, column=c)
                if cell.value is not None:
                    max_data_col = max(max_data_col, c)
        
        # 采样锚点下方前50行
        for r in range(anchor_start_row + 1, min(anchor_start_row + 50, max_row + 1)):
            for c in range(anchor_end_col + 1, max_col + 1):
                merged_addr = self.is_cell_in_merged_range(r, c, merged_ranges)
                if merged_addr:
                    mr2, mc2, er2, ec2 = merged_ranges[merged_addr]
                    cell = ws.cell(row=mr2, column=mc2)
                    if cell.value is not None:
                        max_data_col = max(max_data_col, ec2)
                else:
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None:
                        max_data_col = max(max_data_col, c)
        
        # 向下扩展：从锚点起始行开始，直到遇到空行
        current_row = anchor_start_row
        while current_row <= max_row:
            row_has_data = False
            for c in range(anchor_start_col, max_data_col + 1):
                is_recognized = self._is_in_recognized_range(current_row, c, recognized_ranges)
                if not is_recognized:
                    # 检查是否是合并单元格
                    merged_addr = self.is_cell_in_merged_range(current_row, c, merged_ranges)
                    if merged_addr:
                        mr2, mc2, er2, ec2 = merged_ranges[merged_addr]
                        cell = ws.cell(row=mr2, column=mc2)
                    else:
                        cell = ws.cell(row=current_row, column=c)
                    if cell.value is not None:
                        row_has_data = True
                        break
                else:
                    row_has_data = True
            
            if not row_has_data:
                if anchor_start_row <= current_row <= anchor_end_row:
                    current_row += 1
                    continue
                break
            
            # 添加该行数据
            for c in range(anchor_start_col, max_data_col + 1):
                is_recognized = self._is_in_recognized_range(current_row, c, recognized_ranges)
                if is_recognized:
                    continue
                if (current_row, c) not in visited:
                    merged_addr = self.is_cell_in_merged_range(current_row, c, merged_ranges)
                    if merged_addr:
                        mr2, mc2, er2, ec2 = merged_ranges[merged_addr]
                        cell = ws.cell(row=mr2, column=mc2)
                    else:
                        cell = ws.cell(row=current_row, column=c)
                    if cell.value is not None:
                        table_cells.add((current_row, c))
                    visited.add((current_row, c))
            
            current_row += 1
        
        if not table_cells:
            return None
        
        # 计算边界（基于实际单元格，不基于已识别区域）
        rows = [r for r, c in table_cells]
        cols = [c for r, c in table_cells]
        
        start_row = min(rows)
        end_row = max(rows)
        start_col = min(cols)
        end_col = max(cols)
        
        return {
            'start_row': start_row,
            'end_row': end_row,
            'start_col': start_col,
            'end_col': end_col,
            'cells': table_cells,
            'address': f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        }
    
    def _is_in_recognized_range(self, row: int, col: int, recognized_ranges: list) -> bool:
        """检查单元格是否属于已识别的表格区域"""
        for sr, sc, er, ec in recognized_ranges:
            if sr <= row <= er and sc <= col <= ec:
                return True
        return False
    
    def _expand_rectangular_table(self, ws, start_row: int, start_col: int,
                                   visited: set, merged_ranges: Dict, max_row: int, max_col: int,
                                   recognized_ranges: list) -> Optional[Dict]:
        """
        扩展矩形表格（普通表）
        """
        table_cells = set()
        queue = [(start_row, start_col)]
        
        while queue:
            row, col = queue.pop(0)
            
            if (row, col) in visited:
                continue
            
            # 检查是否属于已识别表格区域（排除重叠）
            if self._is_in_recognized_range(row, col, recognized_ranges):
                visited.add((row, col))
                continue
            
            # 检查是否是合并单元格
            merged_addr = self.is_cell_in_merged_range(row, col, merged_ranges)
            if merged_addr:
                mr, mc, er, ec = merged_ranges[merged_addr]
                # 添加整个合并区域
                for r in range(mr, er + 1):
                    for c in range(mc, ec + 1):
                        if (r, c) not in visited:
                            table_cells.add((r, c))
                            visited.add((r, c))
                # 从合并区域底部继续向下扩展
                # 添加合并区域下方的单元格到队列
                for c in range(mc, ec + 1):
                    if er + 1 <= max_row and (er + 1, c) not in visited:
                        queue.append((er + 1, c))
            else:
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    table_cells.add((row, col))
                    visited.add((row, col))
                else:
                    continue
            
            # 扩展相邻单元格
            neighbors = [
                (row - 1, col), (row + 1, col),
                (row, col - 1), (row, col + 1)
            ]
            
            for nr, nc in neighbors:
                if (nr, nc) in visited:
                    continue
                if nr < 1 or nc < 1 or nr > max_row or nc > max_col:
                    continue
                
                # 检查是否是合并单元格
                neighbor_merged = self.is_cell_in_merged_range(nr, nc, merged_ranges)
                if neighbor_merged:
                    merged_info = merged_ranges[neighbor_merged]
                    mr, mc, er, ec = merged_info
                    all_visited = all((r, c) in visited for r in range(mr, er + 1) for c in range(mc, ec + 1))
                    if not all_visited:
                        queue.append((nr, nc))
                else:
                    queue.append((nr, nc))
        
        if not table_cells:
            return None
        
        rows = [r for r, c in table_cells]
        cols = [c for r, c in table_cells]
        
        return {
            'start_row': min(rows),
            'end_row': max(rows),
            'start_col': min(cols),
            'end_col': max(cols),
            'cells': table_cells,
            'address': f"{get_column_letter(min(cols))}{min(rows)}:{get_column_letter(max(cols))}{max(rows)}"
        }
    
    def _is_in_recognized_range(self, row: int, col: int, recognized_ranges: list) -> bool:
        """检查单元格是否属于已识别的表格区域"""
        for sr, sc, er, ec in recognized_ranges:
            if sr <= row <= er and sc <= col <= ec:
                return True
        return False
        
        # 添加表头行
        for c in data_cols:
            table_cells.add((header_row, c))
            visited.add((header_row, c))
        
        # 向下扩展数据行直到连续空行
        current_row = header_row + 1
        empty_row_count = 0
        
        while current_row <= max_row:
            row_has_data = False
            for c in data_cols:
                if (current_row, c) not in visited:
                    cell = ws.cell(row=current_row, column=c)
                    if cell.value is not None:
                        row_has_data = True
                        break
            
            if not row_has_data:
                empty_row_count += 1
                if empty_row_count >= 2:
                    break
            else:
                empty_row_count = 0
                # 添加该行数据
                for c in data_cols:
                    if (current_row, c) not in visited:
                        cell = ws.cell(row=current_row, column=c)
                        if cell.value is not None:
                            table_cells.add((current_row, c))
                        visited.add((current_row, c))
            
            current_row += 1
        
        # 计算边界
        rows = [r for r, c in table_cells]
        cols = [c for r, c in table_cells]
        
        return {
            'start_row': min(rows),
            'end_row': max(rows),
            'start_col': min(cols),
            'end_col': max(cols),
            'cells': table_cells,
            'address': f"{get_column_letter(min(cols))}{min(rows)}:{get_column_letter(max(cols))}{max(rows)}",
            'header_row': header_row
        }
    
    def expand_compact_table(self, ws, start_row: int, start_col: int,
                              visited: set, merged_ranges: Dict, max_row: int, max_col: int) -> Optional[Dict]:
        """
        扩展紧凑表格（使用BFS，但限制在局部区域）
        """
        queue = [(start_row, start_col)]
        table_cells = set()
        
        # 记录起始位置，用于限制扩展范围
        min_row, max_r = start_row, start_row
        min_col, max_c = start_col, start_col
        
        while queue:
            row, col = queue.pop(0)
            
            if (row, col) in visited:
                continue
            
            # 检查是否是合并单元格
            merged_addr = self.is_cell_in_merged_range(row, col, merged_ranges)
            if merged_addr:
                merged_info = merged_ranges[merged_addr]
                mr, mc, er, ec = merged_info
                for r in range(mr, er + 1):
                    for c in range(mc, ec + 1):
                        if (r, c) not in visited:
                            cell = ws.cell(row=r, column=c)
                            if cell.value is not None or self.is_cell_in_merged_range(r, c, merged_ranges):
                                table_cells.add((r, c))
                                visited.add((r, c))
                                min_row = min(min_row, r)
                                max_r = max(max_r, r)
                                min_col = min(min_col, c)
                                max_c = max(max_c, c)
            else:
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    table_cells.add((row, col))
                    visited.add((row, col))
                    min_row = min(min_row, row)
                    max_r = max(max_r, row)
                    min_col = min(min_col, col)
                    max_c = max(max_c, col)
                else:
                    continue
            
            # 扩展相邻单元格
            neighbors = [
                (row - 1, col), (row + 1, col),
                (row, col - 1), (row, col + 1)
            ]
            
            for nr, nc in neighbors:
                if (nr, nc) in visited:
                    continue
                if nr < 1 or nc < 1 or nr > max_row or nc > max_col:
                    continue
                # 限制扩展范围，避免跨越太大
                if abs(nr - start_row) > 50 or abs(nc - start_col) > 50:
                    continue
                
                neighbor_merged = self.is_cell_in_merged_range(nr, nc, merged_ranges)
                if neighbor_merged:
                    merged_info = merged_ranges[neighbor_merged]
                    mr, mc, er, ec = merged_info
                    all_visited = all((r, c) in visited for r in range(mr, er + 1) for c in range(mc, ec + 1))
                    if not all_visited:
                        queue.append((nr, nc))
                else:
                    queue.append((nr, nc))
        
        if not table_cells:
            return None
        
        return {
            'start_row': min_row,
            'end_row': max_r,
            'start_col': min_col,
            'end_col': max_c,
            'cells': table_cells,
            'address': f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_c)}{max_r}"
        }
    
    def expand_table(self, ws, start_row: int, start_col: int, 
                     visited: set, merged_ranges: Dict) -> Optional[Dict]:
        """
        从起始单元格扩展，找到完整的表格边界
        
        Args:
            ws: worksheet 对象
            start_row: 起始行
            start_col: 起始列
            visited: 已访问单元格集合
            merged_ranges: 合并单元格字典
            
        Returns:
            表格信息字典或 None
        """
        # 使用 BFS/DFS 扩展表格
        queue = [(start_row, start_col)]
        table_cells = set()
        
        while queue:
            row, col = queue.pop(0)
            
            if (row, col) in visited:
                continue
            
            # 检查是否是合并单元格
            merged_addr = self.is_cell_in_merged_range(row, col, merged_ranges)
            if merged_addr:
                merged_info = merged_ranges[merged_addr]
                mr, mc, er, ec = merged_info
                # 将合并区域内的所有单元格加入
                for r in range(mr, er + 1):
                    for c in range(mc, ec + 1):
                        if (r, c) not in visited:
                            cell = ws.cell(row=r, column=c)
                            if cell.value is not None or merged_addr == self.is_cell_in_merged_range(r, c, merged_ranges):
                                table_cells.add((r, c))
                                visited.add((r, c))
            else:
                # 普通单元格
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    table_cells.add((row, col))
                    visited.add((row, col))
                else:
                    continue
            
            # 扩展相邻单元格（上、下、左、右）
            neighbors = [
                (row - 1, col), (row + 1, col),
                (row, col - 1), (row, col + 1)
            ]
            
            for nr, nc in neighbors:
                if (nr, nc) in visited:
                    continue
                
                # 检查边界
                if nr < 1 or nc < 1:
                    continue
                
                # 检查是否是合并单元格
                neighbor_merged = self.is_cell_in_merged_range(nr, nc, merged_ranges)
                if neighbor_merged:
                    merged_info = merged_ranges[neighbor_merged]
                    mr, mc, er, ec = merged_info
                    # 检查合并区域是否已完全访问
                    all_visited = all((r, c) in visited for r in range(mr, er + 1) for c in range(mc, ec + 1))
                    if not all_visited:
                        queue.append((nr, nc))
                else:
                    queue.append((nr, nc))
        
        if not table_cells:
            return None
        
        # 计算表格边界
        rows = [r for r, c in table_cells]
        cols = [c for r, c in table_cells]
        
        start_row = min(rows)
        end_row = max(rows)
        start_col = min(cols)
        end_col = max(cols)
        
        return {
            'start_row': start_row,
            'end_row': end_row,
            'start_col': start_col,
            'end_col': end_col,
            'cells': table_cells,
            'address': f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        }
    
    def identify_table_header(self, ws, table_info: Dict) -> Tuple[int, int, int, int]:
        """
        识别表格的表头区域
        
        策略：
        1. 通常第一行是表头
        2. 如果有合并单元格，可能是标题行
        3. 通过样式判断（加粗、背景色等）
        
        Args:
            ws: worksheet 对象
            table_info: 表格信息
            
        Returns:
            (header_start_row, header_end_row, title_row, title_end_row)
        """
        start_row = table_info['start_row']
        end_row = table_info['end_row']
        start_col = table_info['start_col']
        end_col = table_info['end_col']
        
        # 默认第一行是表头
        header_start = start_row
        header_end = start_row
        
        # 检查是否有标题行（合并单元格且居中）
        title_row = None
        title_end = None
        
        # 检查第一行是否有合并单元格（可能是标题）
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col)
            if cell.alignment and cell.alignment.horizontal == 'center':
                # 可能是标题
                title_row = start_row
                title_end = start_row
                break
        
        # 如果没有找到标题，表头就是第一行
        if title_row is None:
            title_row = start_row
            title_end = start_row
            header_start = start_row + 1
            header_end = start_row + 1
        else:
            header_start = start_row + 1
            header_end = start_row + 1
        
        return header_start, header_end, title_row, title_end
    
    def extract_cell_info(self, ws, row: int, col: int, merged_ranges: Dict) -> Dict:
        """
        提取单个单元格的完整信息（数据、公式、样式）
        
        Args:
            ws: worksheet 对象
            row: 行号
            col: 列号
            merged_ranges: 合并单元格字典
            
        Returns:
            单元格信息字典
        """
        cell = ws.cell(row=row, column=col)
        merged_addr = self.is_cell_in_merged_range(row, col, merged_ranges)
        
        # 获取公式（data_only=False 时，cell.value 可能是公式字符串）
        formula = None
        if hasattr(cell, 'data_type') and cell.data_type == 'f':
            formula = cell.value
        elif isinstance(cell.value, str) and cell.value.startswith('='):
            formula = cell.value
        
        # 提取样式信息
        style_info = {
            'font': {
                'name': cell.font.name,
                'size': cell.font.size,
                'bold': cell.font.bold,
                'italic': cell.font.italic,
                'color': self._get_color_safe(cell.font.color),
                'underline': cell.font.underline
            },
            'fill': {
                'patternType': cell.fill.patternType,
                'bgColor': self._get_color_safe(cell.fill.fgColor),
                'fgColor': self._get_color_safe(cell.fill.bgColor)
            },
            'border': {
                'top': self._get_border_style(cell.border.top),
                'bottom': self._get_border_style(cell.border.bottom),
                'left': self._get_border_style(cell.border.left),
                'right': self._get_border_style(cell.border.right)
            },
            'alignment': {
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical,
                'wrapText': cell.alignment.wrapText,
                'indent': cell.alignment.indent
            },
            'number_format': cell.number_format
        }
        
        # 合并单元格信息
        merge_info = None
        if merged_addr:
            merge_info = {
                'address': merged_addr,
                'merged': True
            }
        
        return {
            'row': row,
            'col': col,
            'address': f"{get_column_letter(col)}{row}",
            'value': cell.value,
            'formula': formula,
            'style': style_info,
            'merge': merge_info
        }
    
    def _get_color_safe(self, color_obj) -> Optional[str]:
        """安全获取颜色值，避免 openpyxl 错误消息"""
        if color_obj is None:
            return None
        # Handle RGBColor objects
        if hasattr(color_obj, 'rgb'):
            rgb = color_obj.rgb
            if rgb is not None:
                s = str(rgb)
                if 'Values must be of type' not in s:
                    return s
        # Handle indexed colors
        if hasattr(color_obj, 'index'):
            return f"indexed:{color_obj.index}"
        # Handle string colors
        if isinstance(color_obj, str):
            if 'Values must be of type' not in color_obj:
                return color_obj
        return None
    
    def _get_color(self, color_obj) -> Optional[str]:
        """获取颜色值"""
        if color_obj is None:
            return None
        # Handle RGBColor objects
        if hasattr(color_obj, 'rgb'):
            rgb = color_obj.rgb
            if rgb is not None:
                return str(rgb)
        # Handle indexed colors
        if hasattr(color_obj, 'index'):
            return f"indexed:{color_obj.index}"
        # Handle string colors
        if isinstance(color_obj, str):
            return color_obj
        return None
    
    def _get_border_style(self, border_obj) -> Optional[str]:
        """获取边框样式"""
        if border_obj and border_obj.style:
            return border_obj.style.value if hasattr(border_obj.style, 'value') else str(border_obj.style)
        return None
    
    def extract_table_data(self, ws, table_info: Dict, merged_ranges: Dict) -> Dict:
        """
        提取表格的完整数据（支持 L 型表）
        """
        start_row = table_info['start_row']
        end_row = table_info['end_row']
        start_col = table_info['start_col']
        end_col = table_info['end_col']
        
        # 检测是否为 L 型表（有 title_address 和 title_text）
        is_l_shaped = 'title_address' in table_info and table_info.get('title_text') is not None
        
        if is_l_shaped:
            # L 型表处理：没有独立表头，标题来自纵向合并单元格
            title = table_info['title_text']
            
            # 找到左侧纵向合并列的范围
            title_addr = table_info['title_address']
            title_start_col = start_col
            title_end_col = start_col
            
            # 解析合并单元格范围
            for addr, (mr, mc, er, ec) in merged_ranges.items():
                if addr == title_addr:
                    title_start_col = mc
                    title_end_col = ec
                    break
            
            # 数据列从合并列右侧开始
            data_start_col = title_end_col + 1
            
            # L 型表没有表头行，所有行都是数据行
            headers = []
            header_row_data = []
            data_start_row = start_row
        else:
            # 普通表处理
            header_start, header_end, title_row, title_end = self.identify_table_header(ws, table_info)
            
            # 提取标题
            title = None
            if title_row and title_row <= end_row:
                title_cells = []
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=title_row, column=col)
                    if cell.value:
                        title_cells.append(cell.value)
                if title_cells:
                    title = ' '.join(str(v) for v in title_cells)
            
            headers = []
            header_row_data = []
            if header_start <= end_row:
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=header_start, column=col)
                    headers.append(cell.value if cell.value else f"列{get_column_letter(col)}")
                    header_row_data.append(self.extract_cell_info(ws, header_start, col, merged_ranges))
            
            data_start_row = header_start + 1
        
        # 提取数据行
        data_rows = []
        formulas = {}
        style_grid = {}
        
        # 获取排除区域（用于 L 型表排除小表区域）
        excluded_ranges = table_info.get('excluded_ranges', [])
        
        for row in range(data_start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                # 检查是否在排除区域内
                is_excluded = False
                for ex_r in excluded_ranges:
                    if ex_r[0] <= row <= ex_r[2] and ex_r[1] <= col <= ex_r[3]:
                        is_excluded = True
                        break
                if is_excluded:
                    continue
                
                cell_info = self.extract_cell_info(ws, row, col, merged_ranges)
                row_data.append(cell_info)
                
                # 记录公式
                if cell_info['formula']:
                    formulas[cell_info['address']] = cell_info['formula']
                
                # 记录样式
                style_grid[cell_info['address']] = cell_info['style']
            
            data_rows.append({
                'row': row,
                'data': row_data,
                'values': [cell['value'] for cell in row_data]
            })
        
        return {
            'address': table_info['address'],
            'range': {
                'start_row': start_row,
                'end_row': end_row,
                'start_col': start_col,
                'end_col': end_col
            },
            'title': title,
            'headers': headers,
            'header_row': header_row_data,
            'data_rows': data_rows,
            'formulas': formulas,
            'styles': style_grid,
            'merged_cells': {addr: info for addr, info in merged_ranges.items() 
                           if self._is_in_range(addr, start_row, start_col, end_row, end_col)}
        }
    
    def _is_in_range(self, addr: str, start_row: int, start_col: int, 
                     end_row: int, end_col: int) -> bool:
        """检查合并单元格是否在表格范围内"""
        start_addr, end_addr = addr.split(':')
        
        sr = int(start_addr[1:].split(':')[0]) if ':' in start_addr else int(''.join(filter(str.isdigit, start_addr)))
        sc_letter = ''.join(filter(str.isalpha, start_addr))
        er = int(''.join(filter(str.isdigit, end_addr)))
        ec_letter = ''.join(filter(str.isalpha, end_addr))
        
        start_col_num = column_index_from_string(sc_letter)
        end_col_num = column_index_from_string(ec_letter)
        
        return (start_row <= sr <= end_row and start_col <= start_col_num <= end_col and
                start_row <= er <= end_row and start_col <= end_col_num <= end_col)
    
    def extract_all_sheets(self) -> Dict[str, Any]:
        """
        提取所有 sheet 中的表格信息
        
        Returns:
            所有表格的结构化数据
        """
        result = {
            'file': self.file_path,
            'sheets': {}
        }
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            # 获取合并单元格
            merged_ranges = self.detect_merged_cells(ws)
            
            # 检测表格边界
            tables = self.detect_table_boundaries(ws, merged_ranges)
            
            sheet_data = {
                'sheet_name': sheet_name,
                'tables': [],
                'table_count': len(tables)
            }
            
            for i, table_info in enumerate(tables, 1):
                table_data = self.extract_table_data(ws, table_info, merged_ranges)
                table_data['table_index'] = i
                sheet_data['tables'].append(table_data)
            
            result['sheets'][sheet_name] = sheet_data
        
        return result
    
    def save_to_json(self, output_path: str = None):
        """
        提取所有数据并保存到 JSON 文件
        
        Args:
            output_path: 输出文件路径，默认在原文件名后加 _extracted.json
        """
        if not output_path:
            output_path = self.file_path.rsplit('.', 1)[0] + '_extracted.json'
        
        result = self.extract_all_sheets()
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, cls=DateTimeEncoder)
        
        return output_path
    
    def close(self):
        """关闭工作簿"""
        self.wb.close()


def main():
    """主函数示例"""
    import sys
    
    if len(sys.argv) < 2:
        print("用法：python excel_table_extractor.py <excel_file_path> [output_json_path]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    print(f"正在处理文件：{excel_file}")
    
    extractor = ExcelTableExtractor(excel_file)
    
    if output_file:
        result_path = extractor.save_to_json(output_file)
        print(f"结果已保存到：{result_path}")
    else:
        result = extractor.extract_all_sheets()
        print(json.dumps(result, ensure_ascii=False, indent=2, cls=DateTimeEncoder))
    
    extractor.close()


if __name__ == '__main__':
    main()
