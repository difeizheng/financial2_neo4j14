"""Fix timeline static dates in Excel via direct XML manipulation.

Problem: 时间表 has 16 SUMIF blocks where the date rows contain static
values instead of formulas. When depreciation period changes, these dates
don't shift, causing incorrect SUMIF results.

Solution (XML-level to preserve ALL cached values):
1. Parse the xlsx as a zip file
2. Directly modify only the timeline sheet XML
3. Replace static dates with MIN(DATE(...), D$start) formulas
4. Extend all blocks to cover full operational period (C to AX)
5. Copy styles from template cells
6. Re-zip without touching any other sheets

This preserves ALL formula cached values in ALL sheets, unlike openpyxl
which strips cached values on save.

Usage:
    python fix_timeline_dates.py [input.xlsx] [output.xlsx]
"""
from __future__ import annotations

import sys
import os
import re
import shutil
import tempfile
import zipfile
from copy import copy
from datetime import datetime, date, timedelta
from io import BytesIO
from openpyxl.utils import column_index_from_string, get_column_letter
from lxml import etree


NS = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_X = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


BLOCKS = [
    (22, 19, 17),
    (28, 25, 23),
    (34, 31, 29),
    (40, 37, 35),
    (50, 47, 45),
    (56, 53, 51),
    (62, 59, 57),
    (68, 65, 63),
    (78, 75, 73),
    (84, 81, 79),
    (90, 87, 85),
    (96, 93, 91),
    (106, 103, 101),
    (112, 109, 107),
    (118, 115, 113),
    (124, 121, 119),
]

MAX_COL = "AX"
MAX_COL_IDX = column_index_from_string("AX")


def _find_cell_element(row_elem, col_ref: str):
    """Find a <c> element by column letter within a <row> element."""
    for c in row_elem:
        r = c.attrib.get("r", "")
        # Match col_ref at start, followed by digits
        if r.startswith(col_ref) and r[len(col_ref):].isdigit():
            return c
    return None


def _get_or_create_cell(row_elem, col_ref: str, row_num: int):
    """Get or create a <c> element for the given column."""
    existing = _find_cell_element(row_elem, col_ref)
    if existing is not None:
        return existing

    c = etree.SubElement(row_elem, f"{{{NS_X}}}c")
    c.attrib["r"] = f"{col_ref}{row_num}"
    return c


def _set_cell_formula(c_elem, formula: str, style_id: str | None = None):
    """Set a cell's formula. Remove any existing <v> (cached value) since
    Excel will recalculate on open (fullCalcOnLoad)."""
    # Remove existing f and v elements
    for tag in ("f", "v"):
        for child in list(c_elem):
            if child.tag == f"{{{NS_X}}}{tag}":
                c_elem.remove(child)

    # Set formula
    f_elem = etree.SubElement(c_elem, f"{{{NS_X}}}f")
    f_elem.text = formula

    # Set style if provided
    if style_id is not None:
        c_elem.attrib["s"] = style_id

    # Remove type attribute (formula cells don't use 't' for numbers/dates)
    if "t" in c_elem.attrib:
        # Only keep 't' if it's a string formula result
        pass


def _get_row(root, row_num: int):
    """Find a <row> element by row number."""
    for row in root.iter(f"{{{NS_X}}}row"):
        if row.attrib.get("r") == str(row_num):
            return row
    return None


def _excel_serial(serial: float) -> date:
    """Convert Excel serial number to Python date."""
    # Excel epoch: 1899-12-30 (serial 1 = 1900-01-01)
    epoch = date(1899, 12, 30)
    return epoch + timedelta(days=int(serial))


def _date_to_serial(d: date) -> int:
    """Convert Python date to Excel serial number."""
    epoch = date(1899, 12, 30)
    return (d - epoch).days


def fix_timeline_xml(input_path: str, output_path: str) -> None:
    """Fix timeline dates using direct XML manipulation to preserve cached values."""

    # First pass: use openpyxl to find the timeline sheet filename and reference values
    import openpyxl

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ts_name = None
    for n in wb.sheetnames:
        if "时间" in n or "timeline" in n.lower():
            ts_name = n
            break
    if ts_name is None:
        raise ValueError("Timeline sheet not found")

    # Get reference dates for each block
    ws = wb[ts_name]
    block_refs = {}
    for sumif_row, date_row, start_row in BLOCKS:
        c_ref = ws[f"C{start_row}"].value
        d_ref = ws[f"D{start_row}"].value
        c_date = ws[f"C{date_row}"].value
        if isinstance(c_ref, datetime):
            c_ref = c_ref.date() if hasattr(c_ref, "date") else c_ref
        if isinstance(d_ref, datetime):
            d_ref = d_ref.date() if hasattr(d_ref, "date") else d_ref
        if isinstance(c_date, datetime):
            c_date = c_date.date() if hasattr(c_date, "date") else c_date
        block_refs[date_row] = {
            "start_row": start_row,
            "c_ref": c_ref,
            "d_ref": d_ref,
            "c_date": c_date,
        }
    wb.close()

    # Find timeline sheet XML filename
    wb2 = openpyxl.load_workbook(input_path)
    ts_idx = wb2.sheetnames.index(ts_name)
    # openpyxl sheets are 1-indexed in filename
    sheet_file = f"xl/worksheets/sheet{ts_idx + 1}.xml"
    wb2.close()

    print(f"Timeline sheet: {ts_name} -> {sheet_file}")

    # Second pass: direct XML manipulation
    with zipfile.ZipFile(input_path, "r") as zin:
        # Read the timeline sheet XML
        sheet_xml = zin.read(sheet_file)
        root = etree.fromstring(sheet_xml)

        # Also read the styles to find the right style ID for date cells
        sheetdata = root.find(f"{{{NS_X}}}sheetData")

        total_fixed = 0

        for sumif_row, date_row, start_row in BLOCKS:
            year_row = date_row - 1
            month_row = date_row + 1
            refs = block_refs[date_row]

            # Find existing last column by scanning date_row for =D$start formula
            date_row_elem = _get_row(sheetdata, date_row)
            if date_row_elem is None:
                print(f"  Block date_row={date_row}: row not found, skipping")
                continue

            existing_last_col = None
            for ci in range(3, MAX_COL_IDX + 1):
                cl = get_column_letter(ci)
                c_elem = _find_cell_element(date_row_elem, cl)
                if c_elem is not None:
                    f_elem = c_elem.find(f"{{{NS_X}}}f")
                    if f_elem is not None and f_elem.text:
                        if f"D{start_row}" in f_elem.text.upper().replace("$", ""):
                            existing_last_col = cl
                            break

            if existing_last_col is None:
                print(f"  Block date_row={date_row}: no endpoint found, skipping")
                continue

            existing_last_idx = column_index_from_string(existing_last_col)

            # Get template style from D column
            d_cell = _find_cell_element(date_row_elem, "D")
            template_style = d_cell.attrib.get("s") if d_cell is not None else None

            # Get template styles for year and month rows
            year_row_elem = _get_row(sheetdata, year_row)
            month_row_elem = _get_row(sheetdata, month_row)

            year_d_cell = _find_cell_element(year_row_elem, "D") if year_row_elem else None
            year_style = year_d_cell.attrib.get("s") if year_d_cell is not None else template_style

            month_d_cell = _find_cell_element(month_row_elem, "D") if month_row_elem else None
            month_style = month_d_cell.attrib.get("s") if month_d_cell is not None else template_style

            # Step 1: Replace static dates (D to existing_last) with MIN formulas
            prev_col = "C"
            for ci in range(4, existing_last_idx + 1):
                cl = get_column_letter(ci)
                c_elem = _find_cell_element(date_row_elem, cl)
                is_endpoint = (cl == existing_last_col)

                # Check if this cell has a static value (no <f> element)
                has_formula = False
                if c_elem is not None:
                    f_elem = c_elem.find(f"{{{NS_X}}}f")
                    has_formula = f_elem is not None and f_elem.text is not None

                if not has_formula or is_endpoint:
                    if ci == 4:
                        formula = f"=MIN(DATE(YEAR({prev_col}{date_row}),12,31),D${start_row})"
                    else:
                        formula = f"=MIN(DATE(YEAR({prev_col}{date_row})+1,12,31),D${start_row})"

                    if c_elem is None:
                        c_elem = _get_or_create_cell(date_row_elem, cl, date_row)

                    _set_cell_formula(c_elem, formula, template_style)
                    total_fixed += 1

                prev_col = cl

            # Step 2: Extend date row from existing_last+1 to AX
            for ci in range(existing_last_idx + 1, MAX_COL_IDX + 1):
                cl = get_column_letter(ci)
                prev_cl = get_column_letter(ci - 1)
                formula = f"=MIN(DATE(YEAR({prev_cl}{date_row})+1,12,31),D${start_row})"

                c_elem = _get_or_create_cell(date_row_elem, cl, date_row)
                _set_cell_formula(c_elem, formula, template_style)
                total_fixed += 1

            # Step 3: Extend YEAR row (year_row = date_row - 1)
            if year_row_elem is not None:
                for ci in range(4, MAX_COL_IDX + 1):
                    cl = get_column_letter(ci)
                    c_elem = _find_cell_element(year_row_elem, cl)
                    has_formula = False
                    if c_elem is not None:
                        f_elem = c_elem.find(f"{{{NS_X}}}f")
                        has_formula = f_elem is not None

                    if not has_formula:
                        formula = f"=YEAR({cl}{date_row})"
                        if c_elem is None:
                            c_elem = _get_or_create_cell(year_row_elem, cl, year_row)
                        _set_cell_formula(c_elem, formula, year_style)
                        total_fixed += 1

            # Step 4: Extend MONTH row (month_row = date_row + 1)
            if month_row_elem is not None:
                prev_cl = "C"
                for ci in range(4, MAX_COL_IDX + 1):
                    cl = get_column_letter(ci)
                    c_elem = _find_cell_element(month_row_elem, cl)
                    has_formula = False
                    if c_elem is not None:
                        f_elem = c_elem.find(f"{{{NS_X}}}f")
                        has_formula = f_elem is not None

                    if not has_formula:
                        formula = f'=ROUND((DATEDIF({prev_cl}{date_row},{cl}{date_row},"d"))/30,0)'
                        if c_elem is None:
                            c_elem = _get_or_create_cell(month_row_elem, cl, month_row)
                        _set_cell_formula(c_elem, formula, month_style)
                        total_fixed += 1

                    prev_cl = cl

            # Step 5: Extend SUMIF formula ranges
            sumif_row_elem = _get_row(sheetdata, sumif_row)
            if sumif_row_elem is not None:
                for ci in range(3, MAX_COL_IDX + 1):
                    cl = get_column_letter(ci)
                    c_elem = _find_cell_element(sumif_row_elem, cl)
                    if c_elem is not None:
                        f_elem = c_elem.find(f"{{{NS_X}}}f")
                        if f_elem is not None and f_elem.text and "SUMIF" in f_elem.text:
                            old_formula = f_elem.text
                            new_formula = re.sub(
                                r"\$C\$\d+:\$[A-Z]+\$\d+",
                                f"$C${year_row}:${MAX_COL}${year_row}",
                                old_formula,
                                count=1,
                            )
                            new_formula = re.sub(
                                r",\$C\$\d+:\$[A-Z]+\$\d+",
                                f",$C${month_row}:${MAX_COL}${month_row}",
                                new_formula,
                                count=1,
                            )
                            f_elem.text = new_formula
                            # Remove cached value so Excel recalculates
                            for v in list(c_elem):
                                if v.tag == f"{{{NS_X}}}v":
                                    c_elem.remove(v)
                            total_fixed += 1

            print(f"  Block date_row={date_row}: C-{existing_last_col} -> C-{MAX_COL}, endpoint fixed")

        # Serialize modified XML
        modified_xml = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

        # Write output zip: replace only the timeline sheet, copy everything else
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == sheet_file:
                    zout.writestr(item, modified_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))

    print(f"\nTotal cells fixed/added: {total_fixed}")
    print(f"Saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        input_path = sys.argv[1]
        output_path = sys.argv[2]
    elif len(sys.argv) == 2:
        input_path = sys.argv[1]
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_fixed{ext}"
    else:
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
        sys.path.insert(0, ".")
        from financial_kg.storage.task_db import TaskDB
        from financial_kg.engine.excel_export import find_original_excel

        db = TaskDB()
        tasks = [t for t in db.list_tasks() if t.status == "done"]
        if not tasks:
            print("No tasks found")
            sys.exit(1)
        t = tasks[-1]
        input_path = find_original_excel(t.id, t.output_dir)
        if not input_path:
            print("Original Excel not found")
            sys.exit(1)
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_timeline_fixed{ext}"

    fix_timeline_xml(input_path, output_path)
