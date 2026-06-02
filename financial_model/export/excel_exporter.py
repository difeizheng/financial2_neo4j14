"""Excel 导出引擎 — AllResults → 格式化多Sheet Excel

将编排器的全部12张报表导出为一个格式化的 .xlsx 文件:
  - 项目摘要: 参数概览 + 关键指标
  - 10张数据表: 投资概算/资金筹措/折旧/成本/收入/利润表×2/现金流量表×3/资产负债表
  - 派生指标: IRR/NPV/DSCR/回收期

格式化:
  - 表头: 加粗 + 浅蓝底色 + 冻结首行
  - 金额: #,##0.00
  - 百分比: 0.00%
  - 比率: 0.00
  - 年份列作为首列

典型用法::

    from financial_model.engines.orchestrator import ModelOrchestrator

    results = ModelOrchestrator.from_excel_v17().run()
    path = export_excel(results, "output.xlsx")
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from financial_model.engines.orchestrator import AllResults


# ══════════════════════════════════════════════════════════
# 样式定义
# ══════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DATA_FONT = Font(size=10)
_NUMBER_FORMAT_MONEY = "#,##0.00"
_NUMBER_FORMAT_PCT = "0.00%"
_NUMBER_FORMAT_RATIO = "0.00"
_NUMBER_FORMAT_INT = "0"

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_SUMMARY_LABEL_FONT = Font(bold=True, size=11)
_SUMMARY_VALUE_FONT = Font(size=11)
_SUMMARY_LABEL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


# ══════════════════════════════════════════════════════════
# Sheet 配置
# ══════════════════════════════════════════════════════════

# 每张数据表的配置: (sheet_name, number_format)
# number_format 用于所有数据列 (index列除外)

class _SheetConfig:
    """单张Sheet的导出配置"""

    def __init__(
        self,
        title: str,
        number_format: str = _NUMBER_FORMAT_MONEY,
        pct_columns: set[str] | None = None,
        int_columns: set[str] | None = None,
    ) -> None:
        self.title = title
        self.number_format = number_format
        self.pct_columns = pct_columns or set()
        self.int_columns = int_columns or set()


# 数据表配置
_SHEET_CONFIGS: dict[str, _SheetConfig] = {
    "investment": _SheetConfig("投资概算表"),
    "depreciation": _SheetConfig(
        "折旧摊销表",
        pct_columns={"production_ratio"},
    ),
    "cost": _SheetConfig(
        "成本费用表",
        pct_columns={"production_ratio"},
    ),
    "revenue": _SheetConfig(
        "收入税金表",
        pct_columns={"production_ratio"},
    ),
    "pnl_total": _SheetConfig("利润表-全投资"),
    "pnl_equity": _SheetConfig("利润表-资本金"),
    "cf_total": _SheetConfig("现金流量表-全投资"),
    "cf_equity": _SheetConfig("现金流量表-资本金"),
    "cf_plan": _SheetConfig("现金流量表-财务计划"),
    "balance_sheet": _SheetConfig("资产负债表"),
}


# ══════════════════════════════════════════════════════════
# 公共 API
# ══════════════════════════════════════════════════════════


def export_excel(
    results: AllResults,
    path: str | Path,
    project_name: str = "",
) -> Path:
    """将 AllResults 导出为格式化 Excel

    Args:
        results: 编排器的完整运行结果
        path: 输出文件路径
        project_name: 项目名称 (显示在摘要页)

    Returns:
        输出文件的 Path 对象
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: 项目摘要
    _write_summary_sheet(wb, results, project_name)

    # Sheet 2: 资金筹措表 (financing 有两个子表)
    _write_financing_sheet(wb, results)

    # Sheet 3-12: 标准数据表
    _DATA_SHEET_BUILDERS: dict[str, tuple[pd.DataFrame, str]] = {
        "investment": (results.investment, "投资概算表"),
        "depreciation": (results.depreciation, "折旧摊销表"),
        "cost": (results.cost, "成本费用表"),
        "revenue": (results.revenue, "收入税金表"),
        "pnl_total": (results.pnl_total.data, "利润表-全投资"),
        "pnl_equity": (results.pnl_equity.data, "利润表-资本金"),
        "cf_total": (results.cf_total.data, "现金流量表-全投资"),
        "cf_equity": (results.cf_equity.data, "现金流量表-资本金"),
        "cf_plan": (results.cf_plan.data, "现金流量表-财务计划"),
        "balance_sheet": (results.balance_sheet.data, "资产负债表"),
    }

    for key, (df, title) in _DATA_SHEET_BUILDERS.items():
        config = _SHEET_CONFIGS.get(key, _SheetConfig(title))
        _write_data_sheet(wb, title, df, config)

    # 最后一个 Sheet: 派生指标
    _write_derived_metrics_sheet(wb, results)

    # 删除默认的空 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(path))
    return path


# ══════════════════════════════════════════════════════════
# Sheet 写入函数
# ══════════════════════════════════════════════════════════


def _write_summary_sheet(
    wb: Workbook,
    results: AllResults,
    project_name: str,
) -> None:
    """写入项目摘要 Sheet"""
    ws = wb.active
    ws.title = "项目摘要"

    # 标题
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = project_name or "抽水蓄能电站财务模型"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    row = 3

    # ── 关键指标 ──
    ws.merge_cells(f"A{row}:D{row}")
    _set_section_header(ws, row, "关键财务指标")
    row += 1

    dm = results.derived_metrics
    summary = dm.summary()
    for label, value in summary.items():
        ws.cell(row=row, column=1, value=label).font = _SUMMARY_LABEL_FONT
        ws.cell(row=row, column=1).fill = _SUMMARY_LABEL_FILL
        ws.cell(row=row, column=2, value=value).font = _SUMMARY_VALUE_FONT
        row += 1

    row += 1

    # ── 投资摘要 ──
    ws.merge_cells(f"A{row}:D{row}")
    _set_section_header(ws, row, "投资摘要")
    row += 1

    invest_total = float(results.investment["construction_investment"].sum())
    fin = results.financing

    invest_items: list[tuple[str, str]] = [
        ("建设投资(万元)", f"{invest_total:,.2f}"),
        ("建设期利息(万元)", f"{fin.construction_interest_total:,.2f}"),
        ("动态总投资(万元)", f"{fin.dynamic_total_investment:,.2f}"),
    ]
    for label, value in invest_items:
        ws.cell(row=row, column=1, value=label).font = _SUMMARY_LABEL_FONT
        ws.cell(row=row, column=1).fill = _SUMMARY_LABEL_FILL
        ws.cell(row=row, column=2, value=value).font = _SUMMARY_VALUE_FONT
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15


def _write_financing_sheet(
    wb: Workbook,
    results: AllResults,
) -> None:
    """写入资金筹措表 (年度汇总 + 还款计划)"""
    ws = wb.create_sheet("资金筹措表")

    # ── 年度汇总 ──
    ws.merge_cells("A1:J1")
    ws["A1"].value = "资金筹措表 — 年度汇总"
    ws["A1"].font = Font(bold=True, size=13)

    _write_dataframe_rows(ws, results.financing.annual_summary, start_row=3)

    # ── 还款计划 ──
    loan_start = len(results.financing.annual_summary) + 6
    ws.merge_cells(f"A{loan_start}:J{loan_start}")
    ws.cell(row=loan_start, column=1, value="还款计划").font = Font(bold=True, size=13)

    _write_dataframe_rows(ws, results.financing.loan_schedule, start_row=loan_start + 2)

    # 格式化
    _format_data_area(ws, start_row=3, ncols=len(results.financing.annual_summary.columns) + 1)
    _auto_width(ws)


def _write_data_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    config: _SheetConfig,
) -> None:
    """写入标准数据表 Sheet"""
    ws = wb.create_sheet(title)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(df.columns) + 1, 10))
    ws["A1"].value = title
    ws["A1"].font = Font(bold=True, size=13)

    # 数据
    _write_dataframe_rows(ws, df, start_row=3)

    # 格式化
    _format_data_area(ws, start_row=3, ncols=len(df.columns) + 1)

    # 特殊列格式
    _apply_column_formats(ws, df, config, start_row=4)

    # 冻结首行 (标题行)
    ws.freeze_panes = "A4"

    # 列宽
    _auto_width(ws)


def _write_derived_metrics_sheet(
    wb: Workbook,
    results: AllResults,
) -> None:
    """写入派生指标 Sheet"""
    ws = wb.create_sheet("派生指标")

    ws["A1"].value = "派生指标"
    ws["A1"].font = Font(bold=True, size=13)

    dm = results.derived_metrics
    row = 3

    items: list[tuple[str, str]] = [
        ("全投资IRR", _fmt_pct(dm.irr_total)),
        ("资本金IRR", _fmt_pct(dm.irr_equity)),
        ("全投资NPV(万元)", _fmt_num(dm.npv_total)),
        ("资本金NPV(万元)", _fmt_num(dm.npv_equity)),
        ("最低DSCR", _fmt_ratio(dm.dscr_min)),
        ("平均DSCR", _fmt_ratio(dm.dscr_avg)),
        ("静态回收期(年)", _fmt_years(dm.payback_static)),
        ("动态回收期(年)", _fmt_years(dm.payback_dynamic)),
        ("平均ROE", _fmt_pct(dm.roe_avg)),
        ("折现率", _fmt_pct(dm.discount_rate)),
        ("项目年限", str(dm.project_years)),
    ]

    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = _SUMMARY_LABEL_FONT
        ws.cell(row=row, column=1).fill = _SUMMARY_LABEL_FILL
        ws.cell(row=row, column=2, value=value).font = _SUMMARY_VALUE_FONT
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


# ══════════════════════════════════════════════════════════
# 辅助函数
# ══════════════════════════════════════════════════════════


def _write_dataframe_rows(
    ws: Any,
    df: pd.DataFrame,
    start_row: int,
) -> None:
    """将 DataFrame 写入 worksheet, 含 index 和 headers"""
    # 表头: index name + columns
    headers = [df.index.name or ""] + list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # 数据行
    for r_idx, (index_val, row_data) in enumerate(df.iterrows(), start_row + 1):
        # index 列
        cell = ws.cell(row=r_idx, column=1, value=index_val)
        cell.font = _DATA_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

        # 数据列
        for c_idx, val in enumerate(row_data.values, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            cell.number_format = _NUMBER_FORMAT_MONEY


def _format_data_area(
    ws: Any,
    start_row: int,
    ncols: int,
) -> None:
    """应用默认格式到数据区域"""
    # 表头行已由 _write_dataframe_rows 处理
    pass  # 样式在 _write_dataframe_rows 中已应用


def _apply_column_formats(
    ws: Any,
    df: pd.DataFrame,
    config: _SheetConfig,
    start_row: int,
) -> None:
    """根据配置应用列级数字格式"""
    nrows = len(df) + start_row - 1  # +start_row-1 因为 start_row 是数据起始

    for col_idx, col_name in enumerate(df.columns, 2):  # col 1 = index
        if col_name in config.pct_columns:
            fmt = _NUMBER_FORMAT_PCT
        elif col_name in config.int_columns:
            fmt = _NUMBER_FORMAT_INT
        else:
            fmt = config.number_format

        for row in range(start_row, nrows + 1):
            ws.cell(row=row, column=col_idx).number_format = fmt


def _auto_width(ws: Any) -> None:
    """自动调整列宽 (简易版, 基于最大字符数)"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                # 中文字符宽度 ×2
                char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def _set_section_header(ws: Any, row: int, text: str) -> None:
    """设置小节标题"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, color="4472C4")


def _fmt_pct(v: float | None) -> str:
    if v is None:
        return "N/A"
    return f"{v:.2%}"


def _fmt_num(v: float) -> str:
    return f"{v:,.2f}"


def _fmt_ratio(v: float | None) -> str:
    if v is None:
        return "N/A"
    return f"{v:.2f}"


def _fmt_years(v: float | None) -> str:
    if v is None:
        return "N/A"
    return f"{v:.1f}"
