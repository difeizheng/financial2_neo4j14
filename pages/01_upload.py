"""Page 1: Upload and parse an Excel financial model."""
from __future__ import annotations

import os
import sys
import shutil
import time
import uuid
import tempfile

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from financial_kg.parser.excel_reader import read_excel
from financial_kg.parser.cell_extractor import build_cell_graph
from financial_kg.parser.indicator_builder import build_indicators
from financial_kg.parser.relationship_builder import infer_relationships
from financial_kg.storage.json_store import save_graph, load_graph, verify_cell_count
from financial_kg.storage.task_db import TaskDB
from financial_kg.config import NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD, save_config
from financial_kg.viz.qa_chart import render_bar_chart_html, render_pie_chart_html
from financial_kg.qa.quality_score import compute_quality_diagnostics, QualityDiagnostics
from financial_kg.qa.stage_timer import StageTimer
from financial_kg.engine.excel_export import export_parsed_excel

st.set_page_config(layout="wide")

# ── Helper functions (must be defined before use in script mode) ─────────────

def _render_stage_timings(container, timings: list[dict]) -> None:
    """Render stage timing breakdown as a compact table."""
    with container:
        timing_rows = []
        for t in timings:
            bar = "█" * max(1, int(t["pct"] / 5))
            timing_rows.append({"阶段": t["stage"], "耗时": f"{t['duration_s']:.1f}s", "占比": f"{t['pct']}%", "进度": bar})
        if timing_rows:
            st.dataframe(pd.DataFrame(timing_rows), use_container_width=True, hide_index=True, height=min(len(timing_rows) * 35 + 38, 200))


def _export_parsed_excel_for_task(task) -> None:
    """导出解析后的 Excel（公式单元格保留原样，Excel 自动重算）。"""
    import tempfile
    cells_path = os.path.join(task.output_dir, f"{task.id}_cells.json")
    graph = load_graph(cells_path)
    original_path = os.path.join(task.output_dir, f"{task.id}_original.xlsx")

    if not os.path.exists(original_path):
        st.error(f"未找到原始 Excel 副本：{original_path}")
        return

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
        export_path = tmp_out.name

    with st.spinner(f"导出 {task.filename} ..."):
        export_parsed_excel(original_path, graph.cells, export_path)

    with open(export_path, "rb") as f:
        st.download_button(
            "⬇️ 下载完成",
            data=f,
            file_name=f"{task.filename.rsplit('.', 1)[0]}_parsed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_file_{task.id}",
        )


def _build_table_dag_data(graph) -> dict:
    """Build simplified DAG data for Table dependency visualization."""
    nodes = []
    edges = []
    palette = ["#89b4fa", "#a6e3a1", "#fab387", "#f38ba8", "#cba6f7"]
    sheet_colors: dict[str, str] = {}
    color_idx = 0
    for tbl in graph.tables.values():
        if tbl.sheet not in sheet_colors:
            sheet_colors[tbl.sheet] = palette[color_idx % len(palette)]
            color_idx += 1

    for tbl_id, tbl in graph.tables.items():
        nodes.append({
            "id": tbl_id,
            "label": tbl.name[:15],
            "category": tbl.sheet,
            "color": sheet_colors.get(tbl.sheet, "#BDBDBD"),
        })
        for target_id in tbl.feeds_into:
            if target_id in graph.tables:
                edges.append({"source": tbl_id, "target": target_id})

    return {"nodes": nodes, "edges": edges}


def _render_dep_dag_html(nodes: list[dict], edges: list[dict], title: str = "", height: str = "300px", echarts_cdn: str = "https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js") -> str:
    """Render a simplified DAG using ECharts graph with layered layout."""
    import json
    in_degree: dict[str, int] = {n["id"]: 0 for n in nodes}
    for e in edges:
        if e["target"] in in_degree:
            in_degree[e["target"]] += 1

    layers: dict[str, float] = {}
    queue = [nid for nid, deg in in_degree.items() if deg == 0]
    visited: set[str] = set()
    while queue:
        next_queue = []
        for nid in queue:
            if nid in visited:
                continue
            visited.add(nid)
            layers[nid] = len(layers)
            for e in edges:
                if e["source"] == nid and e["target"] in in_degree:
                    in_degree[e["target"]] -= 1
                    if in_degree[e["target"]] == 0:
                        next_queue.append(e["target"])
        queue = next_queue
    for nid in in_degree:
        if nid not in layers:
            layers[nid] = len(layers)

    max_layer = max(layers.values()) if layers else 1
    node_positions = {}
    layer_counts: dict[int, int] = {}
    for nid, layer in layers.items():
        layer_counts[layer] = layer_counts.get(layer, 0) + 1
    layer_positions: dict[int, int] = {}

    for nid, layer in layers.items():
        idx = layer_positions.get(layer, 0)
        layer_positions[layer] = idx + 1
        x = layer / max_layer * 800 if max_layer > 0 else 400
        count = layer_counts[layer]
        y = 30 + (idx / max(count, 1)) * 260
        node_positions[nid] = (x, y)

    echarts_nodes = []
    for n in nodes:
        pos = node_positions.get(n["id"], (400, 150))
        echarts_nodes.append({
            "id": n["id"],
            "name": n["label"],
            "x": pos[0],
            "y": pos[1],
            "itemStyle": {"color": n["color"]},
            "symbolSize": 40,
            "label": {"show": True, "fontSize": 11, "color": "#cdd6f4"},
        })

    echarts_edges = []
    for e in edges:
        echarts_edges.append({
            "source": e["source"],
            "target": e["target"],
            "lineStyle": {"curveness": 0.2, "color": "#585b70"},
        })

    opt = {
        "title": {"text": title, "textStyle": {"color": "#cdd6f4", "fontSize": 13}, "left": "center"},
        "tooltip": {"show": True, "backgroundColor": "#1e1e2e", "borderColor": "#313244", "textStyle": {"color": "#cdd6f4"}},
        "series": [{
            "type": "graph",
            "layout": "none",
            "data": echarts_nodes,
            "links": echarts_edges,
            "roam": True,
            "draggable": True,
        }],
    }
    opt_json = json.dumps(opt, ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ background: #181825; height: {height}; }}
  #chart {{ width: 100%; height: {height}; }}
</style>
</head>
<body>
<div id="chart"></div>
<script src="{echarts_cdn}"></script>
<script>
var chart = echarts.init(document.getElementById('chart'), 'dark', {{renderer: 'canvas'}});
chart.setOption({opt_json});
window.addEventListener('resize', function() {{ chart.resize(); }});
</script>
</body>
</html>"""


def _render_quality_detail(quality: QualityDiagnostics) -> None:
    """Render detailed quality diagnostics."""
    col1, col2 = st.columns(2)
    with col1:
        st.write("**基础统计**")
        st.write(f"- Sheet 总数: {quality.sheet_count}")
        if quality.empty_sheets:
            st.warning(f"空 Sheet: {', '.join(quality.empty_sheets)}")
        st.write(f"- 表头检测置信度: {quality.header_confidence:.0%}")
        st.write(f"- 公式单元格占比: {quality.formula_ratio:.1%}")
        st.write(f"- 数值常量占比: {quality.constant_ratio:.1%}")
        st.write(f"- 空白单元格占比: {quality.blank_ratio:.1%}")
    with col2:
        st.write("**图谱质量**")
        st.write(f"- Table 覆盖率: {quality.table_coverage:.1%}")
        st.write(f"- 未关联率: {quality.unlinked_ratio:.1%}")
        if quality.has_cycles:
            st.error(f"检测到 {quality.cycle_count} 个循环依赖！")
            with st.expander(f"循环涉及的 Cell（前 20 个）"):
                st.write(", ".join(quality.cycle_cells[:20]))
        else:
            st.success("无循环依赖")

        if quality.unlinked_hotspot:
            st.write("**未关联热点分布**")
            hotspot_data = sorted(quality.unlinked_hotspot.items(), key=lambda x: -x[1])[:10]
            hotspot_df = pd.DataFrame(hotspot_data, columns=["Sheet", "未关联数"]).set_index("Sheet")
            st.bar_chart(hotspot_df, height=200)


def _do_delete_task(db: TaskDB, t) -> None:
    """Delete a task from DB and remove its output files + snapshot directory."""
    output_dir = t.output_dir or "output"
    prefix = t.id
    snapshots_dir = os.path.join("snapshots", t.id)
    deleted_files = []

    for suffix in ["_cells.json", "_indicators.json", "_tables.json"]:
        fp = os.path.join(output_dir, f"{prefix}{suffix}")
        if os.path.isfile(fp):
            os.remove(fp)
            deleted_files.append(fp)

    for fp in db.list_snapshot_files(t.id):
        if os.path.isfile(fp):
            os.remove(fp)
            deleted_files.append(fp)

    if os.path.isdir(snapshots_dir):
        try:
            shutil.rmtree(snapshots_dir)
            deleted_files.append(snapshots_dir)
        except OSError:
            pass

    db.delete_task(t.id)
    if deleted_files:
        st.success(f"已删除任务 {t.id}（{len(deleted_files)} 个文件）")


def _do_neo4j_import(task, uri: str, user: str, pwd: str) -> None:
    """Execute Neo4j import with progress tracking."""
    if not pwd.strip():
        st.error("请填写 Neo4j 密码")
        return
    try:
        from financial_kg.storage.neo4j_store import Neo4jStore

        cells_path = os.path.join(task.output_dir, f"{task.id}_cells.json")
        with st.spinner("加载图谱..."):
            g = load_graph(cells_path)

        neo4j_progress = st.progress(0, text="连接 Neo4j...")
        step_msgs = [
            "导入 Cell 节点...", "导入 Indicator 节点...", "导入 Table 节点...",
            "导入 DEPENDS_ON 关系...", "导入 CALCULATES_FROM 关系...",
            "导入 FEEDS_INTO 关系...", "导入 BELONGS_TO 关系...",
        ]
        step_idx = [0]

        def _progress_cb(msg: str) -> None:
            pct = int((step_idx[0] / len(step_msgs)) * 100)
            neo4j_progress.progress(pct, text=msg)
            step_idx[0] += 1

        with Neo4jStore(uri, user, pwd) as store:
            counts = store.import_graph(g, task_id=task.id, progress_callback=_progress_cb)

        neo4j_progress.progress(100, text="导入完成！")
        st.success("Neo4j 导入成功")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Cell 节点", f"{counts.get('cells', 0):,}")
        c2.metric("Indicator 节点", f"{counts.get('indicators', 0):,}")
        c3.metric("Table 节点", f"{counts.get('tables', 0):,}")
        c4.metric("DEPENDS_ON 关系", f"{counts.get('depends_on', 0):,}")
    except Exception as e:
        st.error(f"Neo4j 导入失败：{e}")


def _do_neo4j_clear(uri: str, user: str, pwd: str) -> None:
    """Execute Neo4j clear with safety confirmation."""
    if not pwd.strip():
        st.error("请填写 Neo4j 密码")
        return

    confirm_key = "_neo4j_clear_confirm"
    if st.session_state.get(confirm_key):
        try:
            from financial_kg.storage.neo4j_store import Neo4jStore

            with Neo4jStore(uri, user, pwd) as store:
                store.clear_database()
            st.success("Neo4j 数据库已清空")
            st.session_state[confirm_key] = False
        except Exception as e:
            st.error(f"清空失败：{e}")
    else:
        st.session_state[confirm_key] = True
        st.warning("再次点击「清空 Neo4j 数据库」确认操作。这将删除数据库中的所有数据！")


st.title("📁 上传 Excel 财务模型")

db = TaskDB()

# ── Wizard state ─────────────────────────────────────────────────────────────
_WIZARD_KEY = "upload_wizard"
if _WIZARD_KEY not in st.session_state:
    st.session_state[_WIZARD_KEY] = {
        "step": 0,
        "filename": "",
        "file_size": 0,
        "uploaded_file": None,
        "sheet_cells": None,
        "graph": None,
        "task_id": "",
        "output_dir": "output",
        "timings": [],
        "quality": None,
        "paths": {},
        "total_raw": 0,
    }

wiz = st.session_state[_WIZARD_KEY]

# ── Step indicator ───────────────────────────────────────────────────────────
_step_names = ["文件选择", "配置选项", "解析处理", "结果看板"]
_step_icons = ["📄", "⚙️", "⏳", "📊"]
_step_cols = st.columns(4)
for i, (name, icon) in enumerate(zip(_step_names, _step_icons)):
    with _step_cols[i]:
        if i < wiz["step"]:
            st.success(f"{icon} {name}")
        elif i == wiz["step"]:
            st.info(f"{icon} {name}  (当前)")
        else:
            st.caption(f"{icon} {name}")
st.divider()

# ── Step 0: File selection ───────────────────────────────────────────────────
if wiz["step"] == 0:
    uploaded = st.file_uploader("选择 Excel 文件 (.xlsx)", type=["xlsx", "xls"])

    if uploaded:
        size_mb = uploaded.size / (1024 * 1024)
        st.info(f"**{uploaded.name}** — {size_mb:.1f} MB")

        if st.button("预览文件信息"):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(uploaded.read())
                tmp_path = tmp.name
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                sheet_rows = []
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    sheet_rows.append({
                        "Sheet": sn,
                        "行数": ws.max_row or 0,
                        "列数": ws.max_column or 0,
                    })
                wb.close()
                st.dataframe(sheet_rows, use_container_width=True, hide_index=True)
            finally:
                os.unlink(tmp_path)
            uploaded.seek(0)  # reset pointer after preview read

        if st.button("下一步：配置选项", type="primary"):
            wiz["uploaded_file"] = uploaded
            wiz["filename"] = uploaded.name
            wiz["file_size"] = uploaded.size
            wiz["step"] = 1
            st.rerun()

# ── Step 1: Configuration ────────────────────────────────────────────────────
if wiz["step"] == 1:
    st.subheader(f"文件: {wiz['filename']}")

    with st.expander("任务 ID", expanded=False):
        wiz["task_id"] = st.text_input("自定义任务 ID（留空自动生成）", value=wiz.get("task_id", ""))

    with st.expander("输出目录", expanded=False):
        wiz["output_dir"] = st.text_input("输出目录路径", value=wiz.get("output_dir", "output"))

    with st.expander("高级选项", expanded=False):
        st.caption("预留：未来可添加解析粒度、跳过空表等选项")

    col_back, col_next = st.columns([1, 1])
    if col_back.button("返回上一步"):
        wiz["step"] = 0
        st.rerun()
    if col_next.button("开始解析", type="primary"):
        wiz["step"] = 2
        st.rerun()

# ── Step 2: Parsing ─────────────────────────────────────────────────────────
if wiz["step"] == 2:
    timer = StageTimer()

    if not wiz.get("task_id", "").strip():
        wiz["task_id"] = str(uuid.uuid4())[:8]

    task_id = wiz["task_id"]
    output_dir = wiz["output_dir"]

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wiz["uploaded_file"].seek(0)  # ensure pointer at start
        tmp.write(wiz["uploaded_file"].read())
        tmp_path = tmp.name

    db.create_task(task_id, wiz["filename"], output_dir)
    db.update_task(task_id, status="running")

    progress_bar = st.progress(0, text="准备解析...")
    status_box = st.empty()
    timing_container = st.container()

    try:
        t0 = time.time()

        # Stage 1: Read Excel
        timer.start("读取 Excel")
        status_box.info("读取 Excel 文件...")
        sheet_cells = read_excel(tmp_path)
        total_raw = sum(len(v) for v in sheet_cells.values())
        wiz["total_raw"] = total_raw
        timer.stop()
        wiz["sheet_cells"] = sheet_cells
        progress_bar.progress(20, text=f"读取完成：{len(sheet_cells)} 个 sheet，{total_raw:,} 个单元格")
        _render_stage_timings(timing_container, timer.summary())

        # Stage 2: Build Cell Graph
        timer.start("构建 Cell 层")
        status_box.info("构建 Cell 层图谱...")

        def _cell_progress(sheet_name, done, total):
            pct = 20 + int((done / total) * 30)
            progress_bar.progress(pct, text=f"Cell 层: {sheet_name} ({done:,}/{total:,})")

        graph = build_cell_graph(sheet_cells, progress_callback=_cell_progress)
        timer.stop()
        wiz["graph"] = graph
        progress_bar.progress(50, text=f"Cell 层完成：{len(graph.cells):,} 节点，{graph.cell_graph.number_of_edges():,} 边")
        _render_stage_timings(timing_container, timer.summary())

        # Stage 3: Build Indicators + Tables
        timer.start("构建 Indicator/Table 层")
        status_box.info("构建 Indicator + Table 层...")
        build_indicators(sheet_cells, graph)
        timer.stop()
        wiz["graph"] = graph
        progress_bar.progress(70, text=f"Indicator: {len(graph.indicators):,}, Table: {len(graph.tables):,}")
        _render_stage_timings(timing_container, timer.summary())

        # Stage 4: Infer Relationships
        timer.start("推导关系")
        status_box.info("推导 Indicator/Table 关系...")
        infer_relationships(graph)
        timer.stop()
        progress_bar.progress(85)
        _render_stage_timings(timing_container, timer.summary())

        # Stage 5: Save JSON
        timer.start("保存 JSON")
        status_box.info("保存 JSON 文件...")
        paths = save_graph(graph, output_dir, task_id=task_id)
        timer.stop()
        wiz["paths"] = paths
        progress_bar.progress(90, text="保存原始 Excel 副本...")

        # Save original Excel copy for later export
        original_copy = os.path.join(output_dir, f"{task_id}_original.xlsx")
        import shutil
        shutil.copy2(tmp_path, original_copy)

        progress_bar.progress(100, text="保存完成")
        _render_stage_timings(timing_container, timer.summary())

        elapsed = time.time() - t0
        status_box.success(f"解析完成！总耗时 {elapsed:.1f}s")

        stats = graph.stats()
        db.update_task(
            task_id,
            status="done",
            cell_count=stats["total_cells"],
            indicator_count=stats["total_indicators"],
            table_count=stats["total_tables"],
            output_dir=output_dir,
        )

        quality = compute_quality_diagnostics(graph)
        wiz["quality"] = quality
        wiz["timings"] = timer.summary()
        wiz["step"] = 3
        st.rerun()

    except Exception as e:
        db.update_task(task_id, status="error", error_msg=str(e))
        status_box.error(f"解析失败：{e}")
    finally:
        os.unlink(tmp_path)

# ── Step 3: Results Dashboard ────────────────────────────────────────────────
if wiz["step"] == 3:
    graph = wiz["graph"]
    quality = wiz["quality"]
    stats = graph.stats()

    # Row 1: Key metrics
    mc = st.columns(5)
    score = quality.score
    score_icon = "✅" if score >= 80 else "⚠️" if score >= 60 else "❌"
    mc[0].metric(f"{score_icon} 数据质量", f"{score:.0f}/100", delta=f"链接率 {(1 - quality.unlinked_ratio) * 100:.0f}%")
    mc[1].metric("Cell 节点", f"{stats['total_cells']:,}")
    mc[2].metric("Indicator", f"{stats['total_indicators']:,}")
    mc[3].metric("Table", f"{stats['total_tables']:,}")
    unlinked = stats.get("unlinked_cells", 0)
    unlinked_pct = f"{unlinked / stats['total_cells'] * 100:.1f}%" if stats["total_cells"] else "0%"
    mc[4].metric("未关联 Cell", f"{unlinked:,}", delta=unlinked_pct)

    # Row 2: Sheet distribution + Cell type distribution
    c1, c2 = st.columns(2)
    with c1:
        with st.expander("Sheet 分布", expanded=True):
            sheet_counts = {}
            for sn in stats["sheets"]:
                sheet_counts[sn] = sum(1 for c in graph.cells.values() if c.sheet == sn)
            sorted_sheets = sorted(sheet_counts.items(), key=lambda x: -x[1])
            html = render_pie_chart_html(
                labels=[s[0][:20] for s in sorted_sheets],
                values=[s[1] for s in sorted_sheets],
                title="各 Sheet 单元格分布",
                height="300px",
                chart_type="doughnut",
            )
            components.html(html, height=320, scrolling=False)
    with c2:
        with st.expander("单元格类型分布", expanded=True):
            type_counts = {"string": 0, "number": 0, "formula": 0, "empty": 0, "other": 0}
            for c in graph.cells.values():
                dt = c.data_type
                if dt in type_counts:
                    type_counts[dt] += 1
                else:
                    type_counts["other"] += 1
            html = render_bar_chart_html(
                labels=["文本", "数值", "公式", "空值", "其他"],
                values=[type_counts[k] for k in type_counts],
                title="单元格类型分布",
                height="300px",
            )
            components.html(html, height=320, scrolling=False)

    # Row 3: Table Top 10 + Table dependency DAG
    c3, c4 = st.columns(2)
    with c3:
        with st.expander("Table 规模 Top 10", expanded=True):
            table_sizes = [
                {"name": t.name[:25], "indicators": len(t.indicator_ids)}
                for t in graph.tables.values()
            ]
            table_sizes.sort(key=lambda x: -x["indicators"])
            top10 = table_sizes[:10]
            if top10:
                html = render_bar_chart_html(
                    labels=[t["name"] for t in top10],
                    values=[t["indicators"] for t in top10],
                    title="Table Indicator 数量 Top 10",
                    height="300px",
                )
                components.html(html, height=320, scrolling=False)
            else:
                st.caption("无 Table 数据")
    with c4:
        with st.expander("Table 依赖关系", expanded=True):
            dag_data = _build_table_dag_data(graph)
            if dag_data["nodes"]:
                html = _render_dep_dag_html(
                    nodes=dag_data["nodes"],
                    edges=dag_data["edges"],
                    title="Table 依赖 DAG",
                    height="300px",
                )
                components.html(html, height=320, scrolling=False)
            else:
                st.caption("无 Table 依赖关系")

    # Row 4: Quality diagnostics detail
    with st.expander("数据质量详细诊断", expanded=False):
        _render_quality_detail(quality)

    # Row 5: Stage timing breakdown
    with st.expander("解析阶段耗时分析", expanded=False):
        timings = wiz["timings"]
        if timings:
            timing_df = pd.DataFrame(timings)
            st.dataframe(timing_df, use_container_width=True, hide_index=True)
            html = render_bar_chart_html(
                labels=[t["stage"] for t in timings],
                values=[t["duration_s"] for t in timings],
                title="各阶段耗时（秒）",
                height="250px",
            )
            components.html(html, height=270, scrolling=False)

    # Row 6: Output files + verification
    st.subheader("输出文件")
    for layer, path in wiz["paths"].items():
        size_kb = os.path.getsize(path) / 1024
        st.write(f"- **{layer}**: `{path}` ({size_kb:.0f} KB)")

    check = verify_cell_count(graph, wiz["total_raw"])
    status = "✅ 一致" if check["match"] else f"⚠️ 差异 {check['diff']:+d}"
    st.write(f"Cell 数量验证：{check['actual']:,} / {check['expected']:,}  {status}")

    st.session_state["current_task_id"] = wiz["task_id"]
    st.session_state["current_graph"] = graph

    st.info("解析完成！可前往「图谱浏览」页面查看交互图谱，或继续上传新文件。")
    if st.button("上传新文件"):
        wiz["step"] = 0
        wiz["uploaded_file"] = None
        wiz["sheet_cells"] = None
        wiz["graph"] = None
        wiz["quality"] = None
        st.rerun()

# ── History tasks ────────────────────────────────────────────────────────────
st.divider()
st.subheader("历史任务")
tasks = db.list_tasks()

if tasks:
    _DEL_CONFIRM_KEY = "task_delete_confirm"
    if _DEL_CONFIRM_KEY not in st.session_state:
        st.session_state[_DEL_CONFIRM_KEY] = None

    # Status filter
    status_filter = st.radio(
        "筛选状态",
        ["全部", "✅ 完成", "❌ 错误", "⏳ 运行中", "🕐 等待中"],
        horizontal=True,
        label_visibility="collapsed",
    )
    status_map = {"全部": None, "✅ 完成": "done", "❌ 错误": "error", "⏳ 运行中": "running", "🕐 等待中": "pending"}
    filter_status = status_map[status_filter]
    filtered_tasks = [t for t in tasks if t.status == filter_status] if filter_status else tasks

    if filtered_tasks:
        # Header row
        hc_id, hc_fn, hc_cell, hc_ind, hc_tbl, hc_time, hc_act = st.columns([5, 10, 6, 6, 5, 10, 4])
        hc_id.caption("ID"); hc_fn.caption("文件名"); hc_cell.caption("Cell")
        hc_ind.caption("Indicator"); hc_tbl.caption("Table"); hc_time.caption("创建时间"); hc_act.caption("操作")

        for t in filtered_tasks:
            icon = {"done": "✅", "running": "⏳", "error": "❌", "pending": "🕐"}.get(t.status, "?")
            c_id, c_fn, c_cell, c_ind, c_tbl, c_time, c_act = st.columns([5, 10, 6, 6, 5, 10, 4])
            c_id.text(f"{icon} {t.id}")
            c_fn.text(t.filename[:25])
            c_cell.text(f"{t.cell_count:,}" if t.cell_count else "—")
            c_ind.text(f"{t.indicator_count:,}" if t.indicator_count else "—")
            c_tbl.text(f"{t.table_count:,}" if t.table_count else "—")
            c_time.text(t.created_at[:19])
            with c_act:
                if t.status == "done":
                    if st.button("🔍", key=f"view_{t.id}", help="跳转到图谱浏览"):
                        st.session_state["current_task_id"] = t.id
                        st.session_state["current_graph"] = load_graph(
                            os.path.join(t.output_dir, f"{t.id}_cells.json")
                        )
                        st.switch_page("pages/02_explorer.py")
                    if st.button("📥", key=f"dl_{t.id}", help="下载解析后 Excel（公式单元格保留原样）"):
                        _export_parsed_excel_for_task(t)
                if st.button("🗑️", key=f"del_{t.id}", help=f"删除任务 {t.id}"):
                    st.session_state[_DEL_CONFIRM_KEY] = t.id

    # Delete confirmation
    if st.session_state[_DEL_CONFIRM_KEY]:
        t = next((x for x in tasks if x.id == st.session_state[_DEL_CONFIRM_KEY]), None)
        if t:
            st.warning(f"确认删除任务 **{t.id}**（{t.filename}）？输出文件将一并删除。")
            cc1, cc2, _ = st.columns([1, 1, 6])
            if cc1.button("确认删除", key=f"confirm_del_{t.id}", type="primary"):
                _do_delete_task(db, t)
                st.session_state[_DEL_CONFIRM_KEY] = None
                st.rerun()
            if cc2.button("取消", key=f"cancel_del_{t.id}"):
                st.session_state[_DEL_CONFIRM_KEY] = None
                st.rerun()
else:
    st.info("暂无历史任务")

# ── Neo4j Import ─────────────────────────────────────────────────────────────
st.divider()
with st.expander("Neo4j 导入配置", expanded=False):
    done_tasks = [t for t in db.list_tasks() if t.status == "done"]
    if not done_tasks:
        st.info("暂无已完成的任务可导入。")
    else:
        neo4j_task_label = st.selectbox(
            "选择要导入的任务",
            [f"{t.id} — {t.filename}" for t in done_tasks],
            key="neo4j_task_select",
        )
        selected_neo4j_task = next(
            t for t in done_tasks if f"{t.id} — {t.filename}" == neo4j_task_label
        )

        col_uri, col_user, col_pwd = st.columns(3)
        neo4j_uri = col_uri.text_input("Neo4j URI", value=NEO4J_URI, key="n4j_uri")
        neo4j_user = col_user.text_input("User", value=NEO4J_USER, key="n4j_user")
        neo4j_pwd = col_pwd.text_input("Password", value=NEO4J_PASSWORD, type="password", key="n4j_pwd")

        if st.button("保存 Neo4j 配置到 .env"):
            save_config(neo4j_uri=neo4j_uri, neo4j_user=neo4j_user, neo4j_password=neo4j_pwd)
            st.success("Neo4j 配置已保存")

        st.divider()
        col_import, col_clear = st.columns([1, 1])

        if col_import.button("导入到 Neo4j", type="primary"):
            _do_neo4j_import(selected_neo4j_task, neo4j_uri, neo4j_user, neo4j_pwd)

        if col_clear.button("清空 Neo4j 数据库", type="secondary"):
            _do_neo4j_clear(neo4j_uri, neo4j_user, neo4j_pwd)

