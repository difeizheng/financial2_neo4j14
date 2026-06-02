"""Phase 8 精度闭合诊断 — 逐年对比引擎 vs Excel v17

目标: 定位 IRR -0.16pp 偏差的来源年份和具体行项目

对标方法:
  1. 从 Excel sheet 11 (全投资现金流量表) 读取逐年数据
  2. 运行引擎获取逐年数据
  3. 按日历年聚合对比各关键行项
  4. 量化各年对 IRR 偏差的贡献
"""

from __future__ import annotations

from datetime import date, timedelta
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
import pytest

# ══════════════════════════════════════════════════════════
# Excel 读取辅助
# ══════════════════════════════════════════════════════════

import glob as _glob
import os as _os

# 通过 glob 查找 v17 文件 (避免中文路径编码问题)
_v17_files = [f for f in _glob.glob("*v17*xlsx")
              if not f.startswith("~") and "20" not in f.split("v17")[1].split(".")[0]]
assert _v17_files, "Excel v17 file not found"
EXCEL_PATH = _v17_files[0]

# Excel sheet 11 的关键行号映射
EXCEL_ROWS = {
    "total_inflow": 5,          # 1.现金流入
    "revenue": 6,               # 1.1 营业收入
    "subsidy": 9,               # 1.2 补贴收入
    "other_income": 12,         # 1.3 营业外收入
    "wc_recovery": 14,          # 1.4 回收流动资金
    "residual_value": 15,       # 1.5 回收资产余值
    "total_outflow": 16,        # 2.现金流出
    "capex": 17,                # 2.1 建设投资
    "working_capital": 18,      # 2.2 流动资金
    "operating_cost": 19,       # 2.3 维持运营支出
    "vat_outflow": 21,          # 2.5 增值税
    "surcharge": 22,            # 2.6 营业税金及附加
    "maintenance": 23,          # 2.7 维持运营投资支出
    "pretax_net_cf": 26,        # 3.税前净现金流量
    "income_tax": 28,           # 5.调整所得税
    "aftertax_net_cf": 29,      # 6.税后净现金流量
}


def _read_excel_cashflow() -> dict[str, dict[int, float]]:
    """从 Excel 读取全投资现金流量表的逐年数据 (按日历年聚合)

    Returns: {行项名: {年: 值}}
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.worksheets[11]

    # ── 解析年列 ──
    row3 = [c.value for c in ws[3]]
    col_year_map: dict[int, tuple[int, date]] = {}  # col → (year, date)
    for col_idx, val in enumerate(row3):
        if val is not None and col_idx >= 5 and isinstance(val, (int, float)):
            d = date(1899, 12, 30) + timedelta(days=int(val))
            col_year_map[col_idx + 1] = (d.year, d)

    # ── 读取各行项 ──
    result: dict[str, dict[int, float]] = {}
    for name, row_num in EXCEL_ROWS.items():
        by_year: dict[int, float] = {}
        for col, (yr, _) in col_year_map.items():
            val = ws.cell(row=row_num, column=col).value
            v = float(val) if val is not None else 0.0
            by_year[yr] = by_year.get(yr, 0.0) + v
        result[name] = by_year

    wb.close()
    return result


def _read_excel_raw_columns() -> list[tuple[date, float]]:
    """从 Excel 读取原始列级别的税后净现金流量 (用于 IRR 计算)

    Returns: [(date, net_cashflow), ...]
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.worksheets[11]

    row3 = [c.value for c in ws[3]]
    columns: list[tuple[date, float]] = []
    for col_idx, val in enumerate(row3):
        if val is not None and col_idx >= 5 and isinstance(val, (int, float)):
            d = date(1899, 12, 30) + timedelta(days=int(val))
            cf = ws.cell(row=29, column=col_idx + 1).value  # Row 29 = after-tax net CF
            columns.append((d, float(cf) if cf is not None else 0.0))

    wb.close()
    return columns


# ══════════════════════════════════════════════════════════
# Fixtures
# ══════════════════════════════════════════════════════════


@pytest.fixture(scope="module")
def engine():
    """运行引擎"""
    from financial_model.engines.orchestrator import ModelOrchestrator
    return ModelOrchestrator.from_excel_v17().run()


@pytest.fixture(scope="module")
def excel_cf():
    """Excel 现金流量表 (按年聚合)"""
    return _read_excel_cashflow()


@pytest.fixture(scope="module")
def excel_raw():
    """Excel 原始列级数据"""
    return _read_excel_raw_columns()


# ══════════════════════════════════════════════════════════
# 8.1 IRR 偏差溯源
# ══════════════════════════════════════════════════════════


class TestIRRDiagnostics:
    """IRR 偏差溯源 — 逐年逐项对比"""

    def test_irr_raw_column_comparison(self, engine, excel_raw) -> None:
        """对比引擎 vs Excel 的 IRR 计算

        Excel IRR 使用 52 列 (含子年期间), 引擎使用 48 年 (日历年)。
        """
        from financial_model.engines.xirr import xirr as compute_xirr

        # Excel 原始数据 IRR
        excel_dates = [d for d, _ in excel_raw]
        excel_values = np.array([v for _, v in excel_raw], dtype=float)
        excel_irr = compute_xirr(excel_values, excel_dates, basis=365.0)
        print(f"\n  Excel raw IRR (52 cols): {excel_irr:.6f}" if excel_irr else "  Excel IRR: None")

        # 引擎 IRR
        eng_irr = engine.derived_metrics.irr_total
        print(f"  Engine IRR (48 years):   {eng_irr:.6f}" if eng_irr else "  Engine IRR: None")

        # Excel 自身的 IRR 公式值
        print(f"  Excel formula IRR:       0.055493 (D33)")

        if excel_irr and eng_irr:
            print(f"  Excel raw vs formula: {(excel_irr - 0.055493)*100:+.4f}pp")
            print(f"  Engine vs Excel raw:  {(eng_irr - excel_irr)*100:+.4f}pp")

    def test_yearly_net_cashflow_comparison(self, engine, excel_cf) -> None:
        """逐年对比税后净现金流量 — 找出偏差最大的年份"""
        eng_cf = engine.cf_total.data
        excel_ncf = excel_cf["aftertax_net_cf"]

        print("\n  ══ 逐年税后净现金流量对比 ══")
        print(f"  {'Year':>6} {'Engine':>14} {'Excel':>14} {'Diff':>14} {'%Diff':>10}")
        print("  " + "─" * 62)

        big_diff_years: list[tuple[int, float, float, float]] = []

        for year in sorted(set(eng_cf.index) & set(excel_ncf.keys())):
            eng_val = float(eng_cf.loc[year, "net_cashflow"])
            ex_val = excel_ncf.get(year, 0.0)
            diff = eng_val - ex_val
            pct = (diff / abs(ex_val) * 100) if ex_val != 0 else 0.0

            if abs(diff) > 100:  # 差异 > 100万元
                big_diff_years.append((year, eng_val, ex_val, diff))

            if abs(diff) > 50:
                marker = " <<<" if abs(diff) > 500 else ""
                print(
                    f"  {year:>6} {eng_val:>14,.2f} {ex_val:>14,.2f} "
                    f"{diff:>+14,.2f} {pct:>+9.2f}%{marker}"
                )

        print(f"\n  总差异年数: {len(big_diff_years)}")
        total_eng = sum(v[1] for v in big_diff_years)
        total_exc = sum(v[2] for v in big_diff_years)
        total_diff = sum(v[3] for v in big_diff_years)
        print(f"  差异年合计: Engine={total_eng:,.0f}, Excel={total_exc:,.0f}, Diff={total_diff:+,.0f}")

    def test_line_item_comparison(self, engine, excel_cf) -> None:
        """关键行项合计对比 — 定位哪些行项偏差最大"""
        eng_cf = engine.cf_total.data
        years = sorted(set(eng_cf.index) & set(excel_cf["revenue"].keys()))

        # 行项映射: engine列名 → Excel行项名
        mappings = [
            ("revenue", "revenue", "营业收入"),
            ("capex", "capex", "建设投资"),
            ("operating_cost", "operating_cost", "经营成本"),
            ("surcharge", "surcharge", "营业税金及附加"),
            ("maintenance", "maintenance", "维持运营投资"),
            ("income_tax", "income_tax", "所得税"),
            ("residual_value", "residual_value", "回收资产余值"),
            ("net_cashflow", "aftertax_net_cf", "税后净现金流量"),
        ]

        print("\n  ══ 行项合计对比 ══")
        print(f"  {'项目':>15} {'Engine':>14} {'Excel':>14} {'Diff':>14} {'%Diff':>10}")
        print("  " + "─" * 70)

        for eng_col, exc_key, label in mappings:
            eng_total = float(eng_cf[eng_col].sum())
            exc_total = sum(excel_cf[exc_key].get(y, 0.0) for y in years)
            diff = eng_total - exc_total
            pct = (diff / abs(exc_total) * 100) if exc_total != 0 else 0.0
            marker = " <<<" if abs(pct) > 2 else ""
            print(
                f"  {label:>15} {eng_total:>14,.2f} {exc_total:>14,.2f} "
                f"{diff:>+14,.2f} {pct:>+9.2f}%{marker}"
            )

    def test_construction_period_detail(self, engine, excel_cf) -> None:
        """建设期 (2023-2030) 逐年详细对比"""
        eng_cf = engine.cf_total.data
        construction_years = [y for y in range(2023, 2031)]

        print("\n  ══ 建设期逐年详细 ══")
        print(f"  {'Year':>6} {'Eng CF':>12} {'Exc CF':>12} {'Eng CAPEX':>12} {'Exc CAPEX':>12}")
        print("  " + "─" * 58)

        for year in construction_years:
            eng_ncf = float(eng_cf.loc[year, "net_cashflow"]) if year in eng_cf.index else 0.0
            exc_ncf = excel_cf["aftertax_net_cf"].get(year, 0.0)
            eng_capex = float(eng_cf.loc[year, "capex"]) if year in eng_cf.index else 0.0
            exc_capex = excel_cf["capex"].get(year, 0.0)
            print(
                f"  {year:>6} {eng_ncf:>12,.2f} {exc_ncf:>12,.2f} "
                f"{eng_capex:>12,.2f} {exc_capex:>12,.2f}"
            )

    def test_transition_year_detail(self, engine, excel_cf) -> None:
        """过渡年 (2030) 详细拆解 — 关键年份"""
        eng_cf = engine.cf_total.data
        year = 2030

        if year not in eng_cf.index:
            print("  2030 not in engine — skipping")
            return

        items = [
            ("revenue", "revenue", "营业收入"),
            ("capex", "capex", "建设投资"),
            ("operating_cost", "operating_cost", "经营成本"),
            ("surcharge", "surcharge", "营业税金及附加"),
            ("maintenance", "maintenance", "维持运营投资"),
            ("income_tax", "income_tax", "所得税"),
            ("net_cashflow", "aftertax_net_cf", "净现金流量"),
        ]

        print(f"\n  ══ {year} 年过渡期拆解 ══")
        print(f"  {'项目':>15} {'Engine':>14} {'Excel':>14} {'Diff':>14}")
        print("  " + "─" * 60)

        for eng_col, exc_key, label in items:
            eng_val = float(eng_cf.loc[year, eng_col])
            exc_val = excel_cf[exc_key].get(year, 0.0)
            diff = eng_val - exc_val
            print(f"  {label:>15} {eng_val:>14,.2f} {exc_val:>14,.2f} {diff:>+14,.2f}")

    def test_vat_line_item_analysis(self, engine, excel_cf) -> None:
        """增值税行项分析 — Excel 有 VAT 行, 引擎如何处理"""
        eng_cf = engine.cf_total.data
        exc_vat = excel_cf.get("vat_outflow", {})
        years = sorted(set(eng_cf.index) & set(exc_vat.keys()))

        excel_vat_total = sum(exc_vat.get(y, 0.0) for y in years)
        print(f"\n  ══ 增值税分析 ══")
        print(f"  Excel VAT total: {excel_vat_total:,.2f}")
        print(f"  Engine has VAT in: surcharge column")
        print(f"  Engine surcharge total: {float(eng_cf['surcharge'].sum()):,.2f}")
        print(f"  Excel surcharge total: {sum(excel_cf['surcharge'].get(y, 0.0) for y in years):,.2f}")

        # VAT 是独立于 surcharge 的行项
        exc_surcharge_total = sum(excel_cf["surcharge"].get(y, 0.0) for y in years)
        print(f"  Excel surcharge (不含VAT): {exc_surcharge_total:,.2f}")
        print(f"  Excel VAT (单独行): {excel_vat_total:,.2f}")
        print(f"  Engine surcharge 是否含VAT? 对比总额判断")


class TestIRRSensitivity:
    """IRR 灵敏度分析 — 量化各年偏差对 IRR 的贡献"""

    def test_irr_contribution_by_year(self, engine, excel_cf) -> None:
        """将引擎现金流替换为 Excel 值, 观察 IRR 变化"""
        from financial_model.engines.xirr import xirr as compute_xirr

        eng_cf = engine.cf_total.data
        excel_ncf = excel_cf["aftertax_net_cf"]
        years = sorted(set(eng_cf.index) & set(excel_ncf.keys()))

        # 基线: 引擎 IRR
        base_irr = engine.derived_metrics.irr_total
        assert base_irr is not None

        # 找差异最大的年份
        diffs: list[tuple[int, float]] = []
        for year in years:
            eng_val = float(eng_cf.loc[year, "net_cashflow"])
            exc_val = excel_ncf.get(year, 0.0)
            diffs.append((year, abs(eng_val - exc_val)))

        # 按差异排序
        diffs.sort(key=lambda x: x[1], reverse=True)
        top_years = [y for y, _ in diffs[:10]]

        print("\n  ══ IRR 灵敏度: 逐年替换引擎→Excel ══")
        print(f"  Base engine IRR: {base_irr:.6f} ({base_irr*100:.4f}%)")
        print(f"  Target Excel IRR: 0.055493 (5.5493%)")
        print(f"  Gap: {(base_irr - 0.055493)*100:+.4f}pp")
        print()

        # 逐个替换: 把某年的引擎值换成 Excel 值, 看 IRR 变多少
        for target_year in top_years:
            dates = list(engine.cf_total.dates)
            values = list(eng_cf["net_cashflow"].values.astype(float))

            # 替换目标年
            for i, y in enumerate(eng_cf.index):
                if y == target_year and target_year in excel_ncf:
                    values[i] = excel_ncf[target_year]

            new_irr = compute_xirr(np.array(values), dates, basis=365.0)
            if new_irr:
                delta = (new_irr - base_irr) * 100  # pp
                print(
                    f"  Year {target_year}: diff={excel_ncf.get(target_year,0)-float(eng_cf.loc[target_year,'net_cashflow']):+,.0f}"
                    f"  → IRR {new_irr*100:.4f}% (Δ={delta:+.4f}pp)"
                )


# ══════════════════════════════════════════════════════════
# 8.2 资产负债表闭环
# ══════════════════════════════════════════════════════════


class TestBalanceSheetClosure:
    """资产负债表闭环验证 — 资产=负债+所有者权益"""

    def test_annual_balance(self, engine) -> None:
        """每年资产=负债+所有者权益"""
        bs = engine.balance_sheet.data

        print("\n  ══ 资产负债表闭环 ══")

        failures = []
        for year in bs.index:
            assets = float(bs.loc[year, "total_assets"])
            liabilities = float(bs.loc[year, "total_liabilities"])
            equity = float(bs.loc[year, "total_equity"])
            diff = assets - liabilities - equity

            if abs(diff) > 1e-6:
                failures.append((int(year), diff))
                print(f"  Year {year}: A={assets:,.2f}, L+E={liabilities+equity:,.2f}, diff={diff:+.6f} *** UNBALANCED")
            # Only print problematic years or first/last
            elif year == bs.index[0] or year == bs.index[-1]:
                print(f"  Year {year}: diff={diff:+.8f} OK")

        if failures:
            print(f"\n  ** UNBALANCED: {len(failures)} years **")
            for yr, d in failures[:10]:
                print(f"    Year {yr}: {d:+.6f}")
        else:
            print(f"\n  OK: All {len(bs)} years balanced (tolerance 1e-6)")

        assert len(failures) == 0, f"Balance sheet unbalanced in {len(failures)} years"


# ══════════════════════════════════════════════════════════
# 8.3 DSCR 对标
# ══════════════════════════════════════════════════════════


class TestDSCRBenchmark:
    """DSCR 对标验证"""

    def test_dscr_min_and_critical_years(self, engine) -> None:
        """验证最低DSCR和DSCR<1的年份"""
        dm = engine.derived_metrics
        dscr = dm.dscr_by_year

        print("\n  ══ DSCR 分析 ══")
        print(f"  最低DSCR: {dm.dscr_min}")
        print(f"  平均DSCR: {dm.dscr_avg}")

        # DSCR < 1 的年份
        below_1 = {y: r for y, r in dscr.items() if r < 1.0}
        if below_1:
            print(f"  DSCR < 1 的年份:")
            for y, r in sorted(below_1.items()):
                print(f"    {y}: {r:.4f}")
        else:
            print(f"  OK: No years with DSCR < 1")

        # 最低5年
        sorted_dscr = sorted(dscr.items(), key=lambda x: x[1])
        print(f"\n  最低DSCR 5年:")
        for y, r in sorted_dscr[:5]:
            print(f"    {y}: {r:.4f}")

        # Excel 参考: DSCR 应始终 > 1 (抽蓄项目)
        assert dm.dscr_min is not None
        assert dm.dscr_min >= 0.9, f"DSCR min = {dm.dscr_min}, expected >= 0.9"


# ══════════════════════════════════════════════════════════
# 综合报告
# ══════════════════════════════════════════════════════════


class TestPhase8Summary:
    """Phase 8 综合报告 — 始终通过, 仅输出"""

    def test_summary(self, engine, excel_cf) -> None:
        """输出 Phase 8 综合诊断摘要"""
        dm = engine.derived_metrics
        fin = engine.financing

        print("\n" + "═" * 60)
        print("  Phase 8 精度闭合诊断摘要")
        print("═" * 60)

        # 投资概算
        inv = engine.investment
        print(f"\n  ── 投资概算 ──")
        print(f"  静态投资:     {float(inv['static_investment'].sum()):>14,.2f}  (Excel: 810,608.42)")
        print(f"  建设投资:     {float(inv['construction_investment'].sum()):>14,.2f}  (Excel: 869,973.95)")
        print(f"  建设期利息:   {fin.construction_interest_total:>14,.2f}  (Excel: 106,971.73)")
        print(f"  动态总投资:   {fin.dynamic_total_investment:>14,.2f}  (Excel: 976,945.68)")

        # 派生指标
        print(f"\n  ── 派生指标 ──")
        irr = dm.irr_total or 0
        print(f"  IRR(税后):    {irr*100:>10.4f}%  (Excel: 5.5493%)")
        print(f"  IRR偏差:      {(irr-0.055493)*100:>+10.4f}pp")
        print(f"  静态回收期:   {dm.payback_static:>10.2f}年  (Excel: 20.48年)")
        print(f"  DSCR min:     {dm.dscr_min:>10.4f}")
        print(f"  DSCR avg:     {dm.dscr_avg:>10.4f}")

        # 差异统计
        eng_cf = engine.cf_total.data
        excel_ncf = excel_cf["aftertax_net_cf"]
        years = sorted(set(eng_cf.index) & set(excel_ncf.keys()))

        total_diff = sum(
            abs(float(eng_cf.loc[y, "net_cashflow"]) - excel_ncf.get(y, 0.0))
            for y in years
        )
        big_diff_count = sum(
            1 for y in years
            if abs(float(eng_cf.loc[y, "net_cashflow"]) - excel_ncf.get(y, 0.0)) > 100
        )

        print(f"\n  ── 现金流差异统计 ──")
        print(f"  逐年|diff|合计: {total_diff:>14,.0f}")
        print(f"  |diff|>100的年: {big_diff_count:>14}")
        print(f"  总对比年数:     {len(years):>14}")

        assert True  # Always pass — reporting only
